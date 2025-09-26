#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
check_VRR_QMS.py

Scan a model.ini and verify:
- If any assigned value points to a file path that contains "edid" and ends with ".bin"
- Infer VRR/QMS support from the EDID filename substrings:
    *_VRR*      -> isSupportVRR must be true
    *_NO_VRR*   -> isSupportVRR must be false
    *_QMS*      -> isSupportQMS must be true
    *_NO_QMS*   -> isSupportQMS must be false
- Compare against [SrcFunc] flags in the same model.ini:
    isSupportVRR = true/false
    isSupportQMS = true/false
- Produce console output and, optionally, append results to an Excel sheet
  with headers: Rules, Result, condition_1, condition_2, ...

Usage:
    python3 check_VRR_QMS.py --model-ini path/to/model.ini [--root ROOT_DIR] \
        [--report kipling.xlsx] [-v]

Notes:
- If --report is provided and the file already exists, results are appended.
- Sheet/tab naming follows your convention: derive PID from model filename prefix.
  E.g., "11_WW_TV_..." -> sheet "PID_11". If no numeric prefix is detected -> "others".
"""

import argparse
import os
import re
import sys
from typing import Dict, List, Optional, Tuple

def smart_read_text(path: str, encodings=("utf-8", "utf-16", "latin-1")) -> str:
    last_err = None
    for enc in encodings:
        try:
            with open(path, "r", encoding=enc) as f:
                return f.read()
        except UnicodeDecodeError as e:
            last_err = e
            continue
        except FileNotFoundError:
            raise
    if last_err:
        raise last_err
    raise RuntimeError(f"Unable to read file: {path}")

def parse_bool(s: str) -> Optional[bool]:
    s = s.strip().strip(';').strip('"').strip("'").lower()
    if s in ("true", "1", "yes", "on"):
        return True
    if s in ("false", "0", "no", "off"):
        return False
    return None

_ASSIGN_RE = re.compile(r'^\s*([A-Za-z0-9_]+)\s*=\s*"(.*?)"\s*;?\s*$')
_SECTION_RE = re.compile(r'^\s*\[([^\]]+)\]\s*$')

def parse_model_ini_for_edids(text: str) -> Tuple[Dict[str, Optional[bool]], Dict[str, Optional[bool]], List[Dict[str, str]]]:
    """
    Returns:
        (srcfunc_vrr, srcfunc_qms, edid_entries)
        - srcfunc_vrr: {"isSupportVRR": True/False/None}
        - srcfunc_qms: {"isSupportQMS": True/False/None}
        - edid_entries: list of dicts with keys:
            section, key, value(path), filename
    """
    current_section = None
    vrr_flag: Optional[bool] = None
    qms_flag: Optional[bool] = None
    edid_entries: List[Dict[str, str]] = []

    for raw_line in text.splitlines():
        # remove inline comments starting with '#' or '//' (keep ';' as value terminator)
        line = raw_line.split('#', 1)[0].split('//', 1)[0]

        sec_m = _SECTION_RE.match(line)
        if sec_m:
            current_section = sec_m.group(1).strip()
            continue

        m = _ASSIGN_RE.match(line)
        if not m:
            continue
        key, val = m.group(1), m.group(2)

        # capture SrcFunc flags
        if current_section and current_section.strip().lower() in ("srcfunc", "vrr"):
            if key.strip() == "isSupportVRR":
                vrr_flag = parse_bool(val)
            elif key.strip() == "isSupportQMS":
                qms_flag = parse_bool(val)

        # record EDID bin paths
        path_lc = val.lower()
        if "edid" in path_lc and path_lc.endswith(".bin"):
            filename = os.path.basename(val)
            edid_entries.append({
                "section": current_section or "",
                "key": key,
                "value": val,
                "filename": filename,
            })

    return ({"isSupportVRR": vrr_flag}, {"isSupportQMS": qms_flag}, edid_entries)

def infer_flags_from_filename(filename: str) -> Tuple[Optional[bool], Optional[bool]]:
    """
    Infer (vrr, qms) from filename tokens.
    Priority: explicit *_NO_* beats *_VRR/_QMS* if both exist (rare but defend).
    """
    up = filename.upper()
    vrr: Optional[bool] = None
    qms: Optional[bool] = None

    if "_NO_VRR" in up:
        vrr = False
    elif "_VRR" in up:
        vrr = True

    if "_NO_QMS" in up:
        qms = False
    elif "_QMS" in up:
        qms = True

    return vrr, qms

def sheet_name_from_model_path(model_path: str) -> str:
    base = os.path.basename(model_path)
    m = re.match(r'^(\d+)_', base)
    if m:
        return f"PID_{m.group(1)}"
    return "others"

def build_rows(model_ini_path: str,
               src_vrr: Optional[bool],
               src_qms: Optional[bool],
               edid_entries: List[Dict[str, str]]) -> List[List[str]]:
    rows: List[List[str]] = []
    for e in edid_entries:
        vrr_infer, qms_infer = infer_flags_from_filename(e["filename"])
        # Determine pass/fail per EDID item
        checks: List[str] = []
        result = "PASS"

        # VRR check (only if filename implies a value)
        if vrr_infer is not None:
            if src_vrr is None:
                result = "FAIL"
                checks.append(f"VRR implied by name ({vrr_infer}) but isSupportVRR missing")
            elif src_vrr != vrr_infer:
                result = "FAIL"
                checks.append(f"VRR implied by name ({vrr_infer}) != isSupportVRR ({src_vrr})")
            else:
                checks.append(f"VRR OK: implied {vrr_infer} == isSupportVRR")
        else:
            checks.append("VRR not implied by filename")

        # QMS check
        if qms_infer is not None:
            if src_qms is None:
                result = "FAIL"
                checks.append(f"QMS implied by name ({qms_infer}) but isSupportQMS missing")
            elif src_qms != qms_infer:
                result = "FAIL"
                checks.append(f"QMS implied by name ({qms_infer}) != isSupportQMS ({src_qms})")
            else:
                checks.append(f"QMS OK: implied {qms_infer} == isSupportQMS")
        else:
            checks.append("QMS not implied by filename")

        rule = "EDID filename tokens must match [SrcFunc] flags"
        row = [
            rule,
            result,
            f"model_ini={model_ini_path}",
            f"section={e['section']} key={e['key']}",
            f"edid_path={e['value']}",
            f"infer_VRR={vrr_infer}",
            f"isSupportVRR={src_vrr}",
            f"infer_QMS={qms_infer}",
            f"isSupportQMS={src_qms}",
        ] + checks
        rows.append(row)
    return rows

def write_or_append_xlsx(xlsx_path: str, sheet: str, rows: List[List[str]]) -> None:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    headers = ["Rules", "Result", "condition_1", "condition_2", "condition_3",
               "condition_4", "condition_5", "condition_6", "condition_7",
               "condition_8", "condition_9", "condition_10"]

    # Expand rows to match header length (leave extra checks after headers if any)
    norm_rows = []
    for r in rows:
        base = r[:len(headers)]
        if len(base) < len(headers):
            base += [""] * (len(headers) - len(base))
        norm_rows.append(base)

    if os.path.exists(xlsx_path):
        wb = load_workbook(xlsx_path)
    else:
        wb = Workbook()

    ws = wb[sheet] if sheet in wb.sheetnames else wb.create_sheet(title=sheet)

    # If empty sheet, write header
    if ws.max_row == 1 and all(c.value is None for c in ws[1]):
        ws.append(headers)

    for r in norm_rows:
        ws.append(r)

    # Formatting: bold header, wrap text, vertical top, uniform column width
    # Bold header
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Wrap & vertical top
    align = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = align

    # Uniform column widths
    for idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(idx)].width = 28

    # Freeze header row and add auto-filter (to match your preferred report UX)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

    # Set active sheet to this one
    wb.active = wb.index(ws)

    wb.save(xlsx_path)

def main():
    ap = argparse.ArgumentParser(description="Check EDID filename vs isSupportVRR/QMS flags in model.ini")
    ap.add_argument("--model-ini", required=True, help="Path to model.ini")
    ap.add_argument("--root", default=".", help="Root dir to resolve relative paths (unused for now, kept for consistency)")
    ap.add_argument("--report", help="If set, append results to this Excel file")
    # Backward compatibility with older param name, if any scripts still pass it.
    ap.add_argument("--report-xlsx", help=argparse.SUPPRESS)
    ap.add_argument("-v", "--verbose", action="store_true", help="Verbose logging")
    args = ap.parse_args()

    # Normalize report path: prefer --report, fallback to --report-xlsx
    report_path = args.report or args.report_xlsx

    model_path = args.model_ini
    try:
        text = smart_read_text(model_path)
    except Exception as e:
        print(f"[ERROR] Cannot read model.ini: {model_path}: {e}", file=sys.stderr)
        sys.exit(2)

    src_vrr_dict, src_qms_dict, edids = parse_model_ini_for_edids(text)
    src_vrr = src_vrr_dict["isSupportVRR"]
    src_qms = src_qms_dict["isSupportQMS"]

    if args.verbose:
        print(f"[INFO] Parsed [SrcFunc]: isSupportVRR={src_vrr}, isSupportQMS={src_qms}")
        print(f"[INFO] Found {len(edids)} EDID .bin entries")

    rows = build_rows(model_path, src_vrr, src_qms, edids)

    # Console summary
    total = len(rows)
    fails = sum(1 for r in rows if r[1] == "FAIL")
    passes = total - fails
    print("== EDID VRR/QMS Consistency Check ==")
    print(f"model.ini : {model_path}")
    print(f"EDID bins : {total} (PASS={passes}, FAIL={fails})")
    for r in rows:
        print(f"- [{r[1]}] {r[0]} | {r[3]} | {r[4]} | {r[5]} {r[6]} | {r[7]} {r[8]}")

    # Excel report
    if report_path:
        sheet = sheet_name_from_model_path(model_path)
        try:
            write_or_append_xlsx(report_path, sheet, rows)
            print(f"[INFO] Report written to: {report_path} (sheet: {sheet})")
        except Exception as e:
            print(f"[ERROR] Failed to write report: {e}", file=sys.stderr)
            sys.exit(3)

if __name__ == "__main__":
    main()
