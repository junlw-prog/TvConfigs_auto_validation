#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Check PQ-related assets in model.ini

Rules:
1) PQ_OSD must equal "/tvconfigs/PQ_OSD/OSDTable.ini" AND file must exist
2) ICM must equal "/tvconfigs/PQ/ICM.bin" AND file must exist
3) DBC must equal "/tvconfigs/PQ/DBC.ini" AND file must exist
4) PQ_PANEL_COLOR must point to a path ending with .ini (case-insensitive) AND file must exist
Result = PASS only if all rules satisfied.
Exports Excel (kipling.xlsx by default) with headers: Rules, Result, condition_1..5
Sheet name inferred from model.ini filename prefix: PID_<num> or 'others'.
"""

import argparse, os, re
from typing import Optional, Dict

# -----------------------------
# Utilities for report
# -----------------------------

def _sheet_name_for_model(model_ini_path: str) -> str:
    base = os.path.basename(model_ini_path or "")
    m = re.match(r"^(\d+)_", base)
    if m:
        return f"PID_{int(m.group(1))}"
    return "others"

def _ensure_openpyxl():
    try:
        import openpyxl  # noqa
    except ImportError:
        raise SystemExit(
            "[ERROR] 需要 openpyxl 以支援報表輸出與附加。\n"
            "  安裝： pip install --user openpyxl\n"
        )

def export_report(res: Dict[str, str], xlsx_path: str = "kipling.xlsx", num_condition_cols: int = 5) -> None:
    _ensure_openpyxl()
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    COMMON_WIDTH = 80
    COMMON_ALIGN = Alignment(wrap_text=True, vertical="top")
    BOLD = Font(bold=True)

    def _na(s: str) -> str:
        s = (s or "").strip()
        return s if s else "N/A"

    sheet_name = _sheet_name_for_model(res.get("model_ini", ""))

    # 開啟或新建 xlsx
    try:
        wb = load_workbook(xlsx_path)
    except Exception:
        wb = Workbook()

    # 建立或取得 sheet
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)
        headers = ["Rules", "Result"] + [f"condition_{i}" for i in range(1, num_condition_cols + 1)]
        ws.append(headers)

    # 準備資料
    rules = (
        "1) PQ_OSD = /tvconfigs/PQ_OSD/OSDTable.ini\n"
        "2) ICM = /tvconfigs/PQ/ICM.bin\n"
        "3) DBC = /tvconfigs/PQ/DBC.ini\n"
        "4) PQ_PANEL_COLOR -> *.ini\n"
        "5) All above files exist"
    )
    result   = "PASS" if res.get("passed", False) else "FAIL"

    conds = [
        f"PQ_OSD = { _na(res.get('pq_osd', '')) }\nexists: {res.get('pq_osd_exists', 'N/A')}",
        f"ICM = { _na(res.get('icm', '')) }\nexists: {res.get('icm_exists', 'N/A')}",
        f"DBC = { _na(res.get('dbc', '')) }\nexists: {res.get('dbc_exists', 'N/A')}",
        f"PQ_PANEL_COLOR = { _na(res.get('pq_panel_color', '')) }\nendswith .ini: {res.get('pq_panel_color_is_ini', 'N/A')}\nexists: {res.get('pq_panel_color_exists', 'N/A')}",
        _na(res.get('summary', '')),
    ]

    row_values = [rules, result] + conds[:num_condition_cols]
    ws.append(row_values)
    last_row = ws.max_row

    total_cols = 2 + num_condition_cols
    for col_idx in range(1, total_cols + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = COMMON_WIDTH

    for cell in ws[1]:  # header
        cell.font = BOLD
        cell.alignment = COMMON_ALIGN

    for cell in ws[last_row]:
        cell.alignment = COMMON_ALIGN

    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        try:
            wb.remove(wb["Sheet"])
        except Exception:
            pass

    wb.save(xlsx_path)

# -----------------------------
# Core parsing / validation
# -----------------------------

def _read_text(path: str) -> str:
    for enc in ("utf-8", "latin-1", "utf-16"):
        try:
            with open(path, "r", encoding=enc) as f:
                return f.read()
        except UnicodeDecodeError:
            continue
        except FileNotFoundError:
            raise
    with open(path, "r") as f:
        return f.read()

def _strip_comment(line: str) -> str:
    line = line.split("#", 1)[0]
    line = line.split(";", 1)[0]
    return line.strip()

def _resolve_tvconfigs_path(root: str, tvconfigs_like: str) -> str:
    if tvconfigs_like.startswith("/tvconfigs/"):
        rel = tvconfigs_like[len("/tvconfigs/"):]
        return os.path.normpath(os.path.join(root, rel))
    if tvconfigs_like.startswith("./") or tvconfigs_like.startswith("../"):
        return os.path.normpath(os.path.join(root, tvconfigs_like))
    if tvconfigs_like.startswith("/"):  # absolute path
        return tvconfigs_like
    return os.path.normpath(os.path.join(root, tvconfigs_like))

def _parse_key_value(txt: str, key: str) -> Optional[str]:
    pattern = rf"^\s*{re.escape(key)}\s*=\s*\"?([^\"]+)\"?\s*$"
    for raw in txt.splitlines():
        line = _strip_comment(raw)
        if not line:
            continue
        m = re.match(pattern, line, re.IGNORECASE)
        if m:
            return m.group(1).strip()
    return None

def check_pq_assets(model_ini: str, root: str, verbose: bool = False) -> Dict[str, str]:
    txt = _read_text(model_ini)

    pq_osd = _parse_key_value(txt, "PQ_OSD")
    icm = _parse_key_value(txt, "ICM")
    dbc = _parse_key_value(txt, "DBC")
    pq_panel_color = _parse_key_value(txt, "PQ_PANEL_COLOR")

    # Expected literals for three keys
    exp_pq_osd = "/tvconfigs/PQ_OSD/OSDTable.ini"
    exp_icm    = "/tvconfigs/PQ/ICM.bin"
    exp_dbc    = "/tvconfigs/PQ/DBC.ini"

    ok_pq_osd_value = (pq_osd == exp_pq_osd)
    ok_icm_value    = (icm == exp_icm)
    ok_dbc_value    = (dbc == exp_dbc)

    # Resolve paths and existence
    pq_osd_path = _resolve_tvconfigs_path(root, pq_osd) if pq_osd else None
    icm_path    = _resolve_tvconfigs_path(root, icm) if icm else None
    dbc_path    = _resolve_tvconfigs_path(root, dbc) if dbc else None
    pq_panel_color_path = _resolve_tvconfigs_path(root, pq_panel_color) if pq_panel_color else None

    pq_osd_exists = bool(pq_osd_path and os.path.exists(pq_osd_path))
    icm_exists    = bool(icm_path and os.path.exists(icm_path))
    dbc_exists    = bool(dbc_path and os.path.exists(dbc_path))

    pq_panel_color_is_ini = bool(pq_panel_color and pq_panel_color.strip().lower().endswith(".ini"))
    pq_panel_color_exists = bool(pq_panel_color_path and os.path.exists(pq_panel_color_path))

    passed = all([
        ok_pq_osd_value, ok_icm_value, ok_dbc_value,
        pq_panel_color_is_ini,
        pq_osd_exists, icm_exists, dbc_exists, pq_panel_color_exists
    ])

    if verbose:
        print(f"[DBG] pq_osd={pq_osd} expect={exp_pq_osd} exist={pq_osd_exists} -> {_safe(pq_osd_path)}")
        print(f"[DBG] icm={icm} expect={exp_icm} exist={icm_exists} -> {_safe(icm_path)}")
        print(f"[DBG] dbc={dbc} expect={exp_dbc} exist={dbc_exists} -> {_safe(dbc_path)}")
        print(f"[DBG] pq_panel_color={pq_panel_color} .ini?={pq_panel_color_is_ini} exist={pq_panel_color_exists} -> {_safe(pq_panel_color_path)}")

    # Build summary text
    missing = []
    if not ok_pq_osd_value: missing.append("PQ_OSD not expected value")
    if not ok_icm_value: missing.append("ICM not expected value")
    if not ok_dbc_value: missing.append("DBC not expected value")
    if not pq_panel_color_is_ini: missing.append("PQ_PANEL_COLOR not *.ini")
    if not pq_osd_exists: missing.append("PQ_OSD file missing")
    if not icm_exists: missing.append("ICM file missing")
    if not dbc_exists: missing.append("DBC file missing")
    if not pq_panel_color_exists: missing.append("PQ_PANEL_COLOR file missing")

    return {
        "passed": passed,
        "model_ini": model_ini,
        "pq_osd": pq_osd or "",
        "pq_osd_exists": str(pq_osd_exists),
        "icm": icm or "",
        "icm_exists": str(icm_exists),
        "dbc": dbc or "",
        "dbc_exists": str(dbc_exists),
        "pq_panel_color": pq_panel_color or "",
        "pq_panel_color_is_ini": str(pq_panel_color_is_ini),
        "pq_panel_color_exists": str(pq_panel_color_exists),
        "summary": "; ".join(missing) if missing else "All checks passed",
    }

def _safe(p: Optional[str]) -> str:
    return p if p else "(None)"

# -----------------------------
# Main
# -----------------------------

def main():
    parser = argparse.ArgumentParser(description="Check PQ assets in model.ini")
    parser.add_argument("--model-ini", required=True, help="path to model ini")
    parser.add_argument("--root", required=True, help="tvconfigs root path (maps /tvconfigs/*)")
    parser.add_argument("-v", "--verbose", action="store_true")
    parser.add_argument("--report", action="store_true", help="export to kipling.xlsx (append) ")
    parser.add_argument("--report-xlsx", metavar="FILE", help="export to specific xlsx (append) ")

    args = parser.parse_args()
    model_ini = args.model_ini
    root = os.path.abspath(args.root)

    res = check_pq_assets(model_ini, root, verbose=args.verbose)

    print(f"[INFO] model_ini : {model_ini}")
    print(f"[INFO] PQ_OSD : {res['pq_osd']}  exists={res['pq_osd_exists']}")
    print(f"[INFO] ICM : {res['icm']}  exists={res['icm_exists']}")
    print(f"[INFO] DBC : {res['dbc']}  exists={res['dbc_exists']}")
    print(f"[INFO] PQ_PANEL_COLOR : {res['pq_panel_color']}  .ini?={res['pq_panel_color_is_ini']}  exists={res['pq_panel_color_exists']}")
    print(f"Result: {'PASS' if res['passed'] else 'FAIL'}")
    if res.get("summary"): print(f"Summary: {res['summary']}")

    if args.report or args.report_xlsx:
        xlsx_path = args.report_xlsx if args.report_xlsx else "kipling.xlsx"
        export_report(res, xlsx_path=xlsx_path, num_condition_cols=5)

if __name__ == "__main__":
    main()
