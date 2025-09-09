
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
check_pq_osd_dolby_bright_strict.py

需求：
只檢查兩個節點的 MEMC_Level 必須為 0：
  1) [PictureModeData_Default] -> [dolby_bright] -> MEMC_Level
  2) [PictureModeData_VO]      -> [dolby_bright] -> MEMC_Level
（小節名 [dolby_bright] 不分大小寫，允許前後空白；鍵名 MEMC_Level 大小寫固定）

流程：
- 從 model.ini 取得 PQ_OSD（大小寫敏感）→ 解析為實體路徑（以 --root 映射 /tvconfigs/*）。
- 解析 OSDTable 檔文字，定位對應區塊與子區塊，擷取 MEMC_Level 的數值。
- 兩個值皆為 0 → PASS；否則 FAIL。
- 將結果以一列附加到 xlsx（表頭：Rules, Result, condition_1..；不自動補 N/A；分頁依 PID_x/others）。
"""

import argparse
import os
import re
from typing import List, Optional, Tuple

# -----------------------------
# Report helpers
# -----------------------------

def _ensure_openpyxl():
    try:
        import openpyxl  # noqa
    except ImportError:
        raise SystemExit(
            "[ERROR] 需要 openpyxl 以支援報表輸出與附加。\n"
            "  安裝： pip install --user openpyxl\n"
        )

def _sheet_name_for_model(model_ini_path: str) -> str:
    base = os.path.basename(model_ini_path or "")
    m = re.match(r"^(\d+)_", base)
    if m:
        return f"PID_{int(m.group(1))}"
    return "others"

def export_report_row(model_ini: str, rules: str, passed: bool, conditions: List[str], xlsx_path: str):
    """
    寫入一行報表：Rules, Result, condition_1..condition_10
    - 不自動補 N/A；缺值留空白。
    - 既有工作簿附加；頁簽按 PID_x/others；首列粗體、欄寬 80、自動換行、垂直靠上。
    """
    _ensure_openpyxl()
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    COMMON_WIDTH = 80
    COMMON_ALIGN = Alignment(wrap_text=True, vertical="top")
    BOLD = Font(bold=True)

    sheet_name = _sheet_name_for_model(model_ini)

    # 開啟或新建 xlsx
    try:
        wb = load_workbook(xlsx_path)
    except Exception:
        wb = Workbook()

    # 建立或取得 sheet
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)
        headers = ["Rules", "Result"] + [f"condition_{i}" for i in range(1, 11)]
        ws.append(headers)

    # 填資料（不足 10 個條件會自動留白）
    row = [rules or "", "PASS" if passed else "FAIL"]
    row.extend(conditions[:10])
    if len(conditions) < 10:
        row.extend([""] * (10 - len(conditions)))

    ws.append(row)
    last_row = ws.max_row

    # 樣式
    total_cols = len(ws[1])
    for col_idx in range(1, total_cols + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = COMMON_WIDTH

    for cell in ws[1]:  # header
        cell.font = BOLD
        cell.alignment = COMMON_ALIGN

    for cell in ws[last_row]:
        cell.alignment = COMMON_ALIGN

    # 移除預設 Sheet
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        try:
            wb.remove(wb["Sheet"])
        except Exception:
            pass

    wb.save(xlsx_path)

# -----------------------------
# Parsing helpers
# -----------------------------

def _read_text(path: str) -> str:
    for enc in ("utf-8", "latin-1", "utf-16"):
        try:
            with open(path, "r", encoding=enc) as f:
                return f.read()
        except UnicodeDecodeError:
            continue
        except FileNotFoundError:
            raise
    with open(path, "r") as f:
        return f.read()

def _strip_comment(line: str) -> str:
    # 去掉 # 或 ; 之後的註解
    line = line.split("#", 1)[0]
    line = line.split(";", 1)[0]
    return line.strip()

def _resolve_tvconfigs_path(root: str, tvconfigs_like: str) -> str:
    """
    把 "/tvconfigs/xxx/yyy.ini" 映射為 "<root>/xxx/yyy.ini"
    其他相對路徑: 以 root 為基底；非 /tvconfigs 開頭之絕對路徑維持不動
    """
    if tvconfigs_like.startswith("/tvconfigs/"):
        rel = tvconfigs_like[len("/tvconfigs/"):]
        return os.path.normpath(os.path.join(root, rel))
    if tvconfigs_like.startswith("./") or tvconfigs_like.startswith("../"):
        return os.path.normpath(os.path.join(root, tvconfigs_like))
    if tvconfigs_like.startswith("/"):
        return tvconfigs_like
    return os.path.normpath(os.path.join(root, tvconfigs_like))

def find_pq_osd_value_case_sensitive(model_ini_path: str) -> Optional[str]:
    """
    僅匹配大小寫敏感鍵名 'PQ_OSD'。
    允許值被雙引號包住；回傳去除引號後的字串。
    """
    txt = _read_text(model_ini_path)
    for raw in txt.splitlines():
        line = _strip_comment(raw)
        if not line:
            continue
        m = re.match(r'^\s*PQ_OSD\s*=\s*"?([^"]+)"?\s*$', line)  # no IGNORECASE
        if m:
            return m.group(1).strip()
    return None

# -----------------------------
# OSD parser for the two strict paths
# -----------------------------

def _iter_top_blocks(txt: str, top_name: str) -> List[Tuple[int, int]]:
    """
    回傳所有 top-level 區塊 [<top_name>] 的大括號內容範圍 (start_idx, end_idx)，不含最外層大括號。
    top_name 比對大小寫敏感。
    """
    spans: List[Tuple[int, int]] = []
    # 找如： ^\s*\[PictureModeData_Default\]\s*$ 然後找後面的 {...} 成對
    for m in re.finditer(rf'^\s*\[{re.escape(top_name)}\]\s*$', txt, flags=re.MULTILINE):
        start = m.end()
        brace_open = re.search(r'\{', txt[start:])
        if not brace_open:
            continue
        block_start = start + brace_open.start()
        depth = 0
        i = block_start
        while i < len(txt):
            ch = txt[i]
            if ch == '{':
                depth += 1
            elif ch == '}':
                depth -= 1
                if depth == 0:
                    inner_start = block_start + 1
                    inner_end = i
                    spans.append((inner_start, inner_end))
                    break
            i += 1
    return spans

def _find_sub_block(txt: str, start: int, end: int, sub_name_ci: str) -> Optional[Tuple[int, int]]:
    """
    在指定區間 (start, end) 中找一個子區塊 [sub_name_ci]（區塊名不分大小寫），
    回傳其內部大括號內容範圍 (inner_start, inner_end)，不含最外層大括號。
    """
    body = txt[start:end]
    # 不分大小寫：(?i)
    for m in re.finditer(rf'(?i)^\s*\[{re.escape(sub_name_ci)}\]\s*$', body, flags=re.MULTILINE):
        after = m.end()
        brace_open = re.search(r'\{', body[after:])
        if not brace_open:
            continue
        block_start = after + brace_open.start()
        depth = 0
        i = block_start
        while i < len(body):
            ch = body[i]
            if ch == '{':
                depth += 1
            elif ch == '}':
                depth -= 1
                if depth == 0:
                    inner_start = start + block_start + 1
                    inner_end = start + i
                    return (inner_start, inner_end)
            i += 1
    return None

def _extract_memc_level_from_range(txt: str, start: int, end: int) -> Optional[int]:
    """
    在 (start, end) 範圍內尋找第一個 'MEMC_Level=<int>'（大小寫固定）並回傳其值；找不到回傳 None。
    """
    body = txt[start:end]
    for line in body.splitlines():
        line = line.strip()
        m = re.match(r'^MEMC_Level\s*=\s*(-?\d+)\s*$', line)
        if m:
            try:
                return int(m.group(1))
            except ValueError:
                return None
    return None

def strict_two_paths_check(osd_ini_path: str) -> Tuple[bool, Optional[int], Optional[int]]:
    """
    僅檢查：
      A = [PictureModeData_Default] -> [dolby_bright] -> MEMC_Level
      B = [PictureModeData_VO]      -> [dolby_bright] -> MEMC_Level
    兩者都存在且都為 0 才 PASS。否則 FAIL。
    回傳: (passed, A_value, B_value)；若任一不存在對應值，該值為 None。
    """
    if not os.path.exists(osd_ini_path):
        return (False, None, None)
    txt = _read_text(osd_ini_path)

    # Default
    a_val: Optional[int] = None
    spans_def = _iter_top_blocks(txt, "PictureModeData_Default")
    if spans_def:
        sub = _find_sub_block(txt, spans_def[0][0], spans_def[0][1], "dolby_bright")
        if sub:
            a_val = _extract_memc_level_from_range(txt, sub[0], sub[1])

    # VO
    b_val: Optional[int] = None
    spans_vo = _iter_top_blocks(txt, "PictureModeData_VO")
    if spans_vo:
        sub = _find_sub_block(txt, spans_vo[0][0], spans_vo[0][1], "dolby_bright")
        if sub:
            b_val = _extract_memc_level_from_range(txt, sub[0], sub[1])

    passed = (a_val == 0) and (b_val == 0)
    return (passed, a_val, b_val)

# -----------------------------
# Main
# -----------------------------

def main():
    parser = argparse.ArgumentParser(description="From model.ini -> PQ_OSD -> strict check on [PictureModeData_Default/PictureModeData_VO]->[dolby_bright]->MEMC_Level==0, then export Excel.")
    parser.add_argument("--model-ini", required=True, help="path to model ini (e.g., model/1_xxx.ini)")
    parser.add_argument("--root", required=True, help="tvconfigs project root (maps /tvconfigs/* to here)")
    parser.add_argument("-v", "--verbose", action="store_true", help="verbose logs")
    parser.add_argument("--report", action="store_true", help="export report to xlsx (default: kipling.xlsx)")
    parser.add_argument("--report-xlsx", metavar="FILE", help="export report to specific xlsx file")

    args = parser.parse_args()

    model_ini = args.model_ini
    if not os.path.exists(model_ini):
        raise SystemExit(f"[ERROR] model ini not found: {model_ini}")

    root = os.path.abspath(os.path.normpath(args.root))

    if args.verbose:
        print(f"[INFO] model_ini: {model_ini}")
        print(f"[INFO] root     : {root}")

    # 1) 取得 PQ_OSD（大小寫敏感）
    pq_osd_value = find_pq_osd_value_case_sensitive(model_ini)
    if args.verbose:
        print(f"[INFO] PQ_OSD: {pq_osd_value or '(not found)'}")

    # 2) 解析成實際檔案路徑
    osd_path = _resolve_tvconfigs_path(root, pq_osd_value) if pq_osd_value else None
    exists = os.path.exists(osd_path) if osd_path else False
    if args.verbose:
        print(f"[INFO] OSD file: {osd_path or '(N/A)'}  exists={exists}")

    # 3) 執行嚴格檢查
    if not exists:
        ok, a_val, b_val = (False, None, None)
    else:
        ok, a_val, b_val = strict_two_paths_check(osd_path)

    # 4) Console 輸出
    print(f"[OSD] {osd_path if exists else '(missing)'}")
    print(f"  Default/dolby_bright MEMC_Level: {a_val if a_val is not None else '(not found)'}")
    print(f"  VO/dolby_bright      MEMC_Level: {b_val if b_val is not None else '(not found)'}")
    print(f"[RESULT] {'PASS' if ok else 'FAIL'}")

    # 5) 報表輸出
    if args.report or args.report_xlsx:
        xlsx_path = args.report_xlsx if args.report_xlsx else "kipling.xlsx"
        rules = "OSD: [PictureModeData_Default/VO]->[dolby_bright]->MEMC_Level must both be 0"
        conds: List[str] = []
        conds.append(f"PQ_OSD = {pq_osd_value or ''}")
        conds.append(f"OSD file exists = {exists}")
        conds.append(f"Default/dolby_bright MEMC_Level = {'' if a_val is None else a_val}")
        conds.append(f"VO/dolby_bright MEMC_Level = {'' if b_val is None else b_val}")

        export_report_row(model_ini, rules, bool(exists and ok), conds, xlsx_path)
        sheet = _sheet_name_for_model(model_ini)
        print(f"[INFO] Report appended to: {xlsx_path} (sheet: {sheet}, +1 row)")

if __name__ == "__main__":
    main()
