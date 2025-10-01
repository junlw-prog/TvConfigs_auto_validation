#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import argparse
import os
import re
from typing import Dict, List, Optional, Tuple, Any

# -----------------------------
# Report helpers (style aligned with tv_multi_standard_validation.py)
# -----------------------------

def _sheet_name_for_model(model_ini_path: str) -> str:
    base = os.path.basename(model_ini_path or "")
    m = re.match(r"^(\d+)_", base)
    if m:
        return f"PID_{int(m.group(1))}"
    return "others"

def _ensure_openpyxl():
    try:
        import openpyxl  # noqa
    except ImportError:
        raise SystemExit(
            "[ERROR] 需要 openpyxl 以支援報表輸出與附加。\n"
            "  安裝： pip install --user openpyxl\n"
        )

def export_report(res: dict, xlsx_path: str = "kipling.xlsx", num_condition_cols: int = 5) -> None:
    """
    表頭固定: Rules, Result, condition_1, condition_2, ...
    欄位無值填 'N/A'，所有欄同寬、換行、垂直置頂。依 model.ini 前綴分頁。
    """
    _ensure_openpyxl()
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    COMMON_WIDTH = 80
    COMMON_ALIGN = Alignment(wrap_text=True, vertical="top")
    BOLD = Font(bold=True)

    def _na(s: Optional[str]) -> str:
        s = (s or "").strip()
        return s if s else "N/A"

    sheet_name = _sheet_name_for_model(res.get("model_ini", ""))

    # 開啟或新建 xlsx
    try:
        wb = load_workbook(xlsx_path)
    except Exception:
        wb = Workbook()

    # 建立或取得 sheet
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)
        headers = ["Rules", "Result"] + [f"condition_{i}" for i in range(1, num_condition_cols + 1)]
        ws.append(headers)

    # 準備資料
    rules = f"4. 標準測試 image\n" \
            f"    - TvServIni → DEFAULT_PICTURE_MODE=9\n" \
            f"    - TvServIni → DEFAULT_DOLBY_PICTURE_MODE=1\n" \
            f"    - TvServIni → SUPPORT_DOLBY_CERT=false"
    result = "PASS" if res.get("passed") else "FAIL"

    conds = [
        f"TvServIni = {_na(res.get('tvserv_ini_path'))}",
        f"DEFAULT_PICTURE_MODE = {_na(res.get('vals', {}).get('DEFAULT_PICTURE_MODE'))}",
        f"DEFAULT_DOLBY_PICTURE_MODE = {_na(res.get('vals', {}).get('DEFAULT_DOLBY_PICTURE_MODE'))}",
        f"SUPPORT_DOLBY_CERT = {_na(res.get('vals', {}).get('SUPPORT_DOLBY_CERT'))}",
    ][:num_condition_cols]

    ws.append([rules, result] + conds)
    last_row = ws.max_row

    # 給儲存格指派上色
    rules_color = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
    failed_color = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")
    # 上色
    first_cell = ws.cell(row=last_row, column=1)  # 欄位1對應的是 'A' 列
    first_cell.fill = rules_color
    if result == "FAIL":
        ws.cell(row=last_row, column=2).fill = failed_color
    if conds[0] == "TvServIni = N/A":
        ws.cell(row=last_row, column=3).fill = failed_color
    if conds[1] != "DEFAULT_PICTURE_MODE = 9" :
        ws.cell(row=last_row, column=4).fill = failed_color
    if conds[2] != "DEFAULT_DOLBY_PICTURE_MODE = 1":
        ws.cell(row=last_row, column=5).fill = failed_color
    if conds[3] != "SUPPORT_DOLBY_CERT = false":
        ws.cell(row=last_row, column=6).fill = failed_color


    # 欄寬/置頂/換行
    total_cols = 2 + num_condition_cols
    for col_idx in range(1, total_cols + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COMMON_WIDTH

    for cell in ws[1]:
        cell.font = BOLD
        cell.alignment = COMMON_ALIGN
    for cell in ws[last_row]:
        cell.alignment = COMMON_ALIGN

    # 移除預設 Sheet
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        try:
            wb.remove(wb["Sheet"])
        except Exception:
            pass

    wb.save(xlsx_path)

# -----------------------------
# Core parsing helpers (path mapping & file reading)
# -----------------------------

def _read_text(path: str) -> str:
    for enc in ("utf-8", "latin-1", "utf-16"):
        try:
            with open(path, "r", encoding=enc) as f:
                return f.read()
        except UnicodeDecodeError:
            continue
        except FileNotFoundError:
            raise
    with open(path, "r") as f:
        return f.read()

def _strip_comment(line: str) -> str:
    # 去掉 # 或 ; 後面的註解
    line = line.split("#", 1)[0]
    line = line.split(";", 1)[0]
    return line.strip()

def _resolve_tvconfigs_path(root: str, tvconfigs_like: str) -> str:
    """
    把 "/tvconfigs/xxx/yyy.ini" 映射為 "<root>/xxx/yyy.ini"
    其他相對路徑: 以 root 為基底；非 /tvconfigs 絕對路徑維持不動
    """
    if tvconfigs_like.startswith("/tvconfigs/"):
        rel = tvconfigs_like[len("/tvconfigs/"):]
        return os.path.normpath(os.path.join(root, rel))
    if tvconfigs_like.startswith("./") or tvconfigs_like.startswith("../"):
        return os.path.normpath(os.path.join(root, tvconfigs_like))
    if tvconfigs_like.startswith("/"):
        return tvconfigs_like
    return os.path.normpath(os.path.join(root, tvconfigs_like))

# -----------------------------
# Model.ini → TvServIni
# -----------------------------

def parse_model_ini_for_tvserv(model_ini_path: str, root: str) -> Optional[str]:
    """
    從 model.ini 找:
      TvServIni = "<path>"
    回傳對應到檔案系統的路徑（已映射到 --root）
    """
    txt = _read_text(model_ini_path)
    tvserv_path = None
    for raw in txt.splitlines():
        line = _strip_comment(raw)
        if not line:
            continue
        m = re.match(r'^\s*TvServIni\s*=\s*"?([^"]+)"?\s*$', line, re.IGNORECASE)
        if m:
            tvserv_path = _resolve_tvconfigs_path(root, m.group(1).strip())
            break
    return tvserv_path

# -----------------------------
# tvserv_ini parsing
# -----------------------------

def parse_tvserv_kv(tvserv_ini_path: str) -> Dict[str, str]:
    """
    解析簡單 key=value，忽略註解與空白。大小寫不敏感地讀 key，但回存標準大寫鍵名。
    """
    if not tvserv_ini_path or not os.path.exists(tvserv_ini_path):
        return {}

    txt = _read_text(tvserv_ini_path)
    out: Dict[str, str] = {}
    for raw in txt.splitlines():
        line = _strip_comment(raw)
        if not line or "=" not in line:
            continue
        key, val = line.split("=", 1)
        key = key.strip()
        val = val.strip()
        # 保留原始大小寫比較安全，但我們把關注的 3 個鍵映射到固定大寫
        ku = key.upper()
        if ku in {"DEFAULT_PICTURE_MODE", "DEFAULT_DOLBY_PICTURE_MODE", "SUPPORT_DOLBY_CERT"}:
            out[ku] = val
    return out

# -----------------------------
# Check logic
# -----------------------------

REQUIRED = {
    "DEFAULT_PICTURE_MODE": "9",
    "DEFAULT_DOLBY_PICTURE_MODE": "1",
    "SUPPORT_DOLBY_CERT": "false",
}

def evaluate(vals: Dict[str, str]) -> Tuple[bool, List[str]]:
    """
    回傳 (passed, notes)
    notes 會列出缺失或不符的項目
    """
    notes: List[str] = []
    passed = True
    for k, expect in REQUIRED.items():
        if k not in vals:
            notes.append(f"缺少 {k}")
            passed = False
        else:
            actual = vals[k]
            if str(actual).strip().lower() != str(expect).strip().lower():
                notes.append(f"{k} 不符 (expect {expect}, got {actual})")
                passed = False
    return passed, notes

def run(
    model_ini: str,
    root: str = ".",
    standard: Optional[str] = None,
    verbose: bool = False,
    conditions: str = "",
    report_xlsx: Optional[str] = None,
    ctx: Any = None,
    **kwargs,                         # 吸收多餘參數避免 TypeError
) -> Dict[str, Any]:
    #root = os.path.abspath(os.path.normpath(args.root))
    if verbose:
        print(f"[INFO] model_ini: {model_ini}")
        print(f"[INFO] root     : {root}")

    # 1) 取 TvServIni
    tvserv_path = parse_model_ini_for_tvserv(model_ini, root)
    if verbose:
        print(f"[INFO] TvServIni : {tvserv_path or '(not found in model.ini)'}")

    vals: Dict[str, str] = {}
    notes: List[str] = []

    if not tvserv_path:
        notes.append("model.ini 未找到 TvServIni")
        passed = False
    elif not os.path.exists(tvserv_path):
        notes.append(f"TvServIni 指向檔案不存在: {tvserv_path}")
        passed = False
    else:
        # 2) 解析 tvserv_ini，抽出三個鍵
        vals = parse_tvserv_kv(tvserv_path)
        passed, more = evaluate(vals)
        notes.extend(more)

    # Console 輸出（僅 PASS/FAIL）
    print(f"Result  : {'PASS' if passed else 'FAIL'}")

    # 3) 報表
    res = {
        "passed": passed,
        "model_ini": model_ini,
        "tvserv_ini_path": tvserv_path or "",
        "vals": vals,
        "notes": notes,
    }

    if report_xlsx:
        out_xlsx = f"{report_xlsx}.xlsx" if not report_xlsx.endswith(".xlsx") else report_xlsx
        #export_report(res, xlsx_path=out_xlsx, num_condition_cols=max(1, len(conditions)))
        export_report(res, xlsx_path=out_xlsx, num_condition_cols=max(4, len(conditions.split(",")) if conditions else 4),)
        print(f"[INFO] Report appended to: {out_xlsx} (sheet: {_sheet_name_for_model(model_ini)})")

# -----------------------------
# Main
# -----------------------------

def main():
    p = argparse.ArgumentParser(
        description="Check TvServIni for picture modes and Dolby cert flag, with optional Excel report."
    )
    p.add_argument("--model-ini", required=True, help="path to model ini (e.g., model/1_xxx.ini)")
    p.add_argument("--root", required=True, help="tvconfigs project root (maps /tvconfigs/* to here)")
    p.add_argument("--report", action="store_true", help="export report to xlsx (default: kipling.xlsx)")
    p.add_argument("--report-xlsx", metavar="FILE", help="export report to specific xlsx file")
    p.add_argument("-v", "--verbose", action="store_true", help="verbose logs")
    args = p.parse_args()

    model_ini = args.model_ini
    if not os.path.exists(model_ini):
        raise SystemExit(f"[ERROR] model ini not found: {model_ini}")

    root = os.path.abspath(os.path.normpath(args.root))
    if args.verbose:
        print(f"[INFO] model_ini: {model_ini}")
        print(f"[INFO] root     : {root}")

    # 1) 取 TvServIni
    tvserv_path = parse_model_ini_for_tvserv(model_ini, root)
    if args.verbose:
        print(f"[INFO] TvServIni : {tvserv_path or '(not found in model.ini)'}")

    vals: Dict[str, str] = {}
    notes: List[str] = []

    if not tvserv_path:
        notes.append("model.ini 未找到 TvServIni")
        passed = False
    elif not os.path.exists(tvserv_path):
        notes.append(f"TvServIni 指向檔案不存在: {tvserv_path}")
        passed = False
    else:
        # 2) 解析 tvserv_ini，抽出三個鍵
        vals = parse_tvserv_kv(tvserv_path)
        passed, more = evaluate(vals)
        notes.extend(more)

    # Console 輸出（僅 PASS/FAIL）
    print(f"Result  : {'PASS' if passed else 'FAIL'}")

    # 3) 報表
    if args.report or args.report_xlsx:
        xlsx_path = args.report_xlsx if args.report_xlsx else "kipling.xlsx"
        res = {
            "passed": passed,
            "model_ini": model_ini,
            "tvserv_ini_path": tvserv_path or "",
            "vals": vals,
            "notes": notes,
        }
        export_report(res, xlsx_path=xlsx_path)
        sheet = _sheet_name_for_model(model_ini)
        print(f"[INFO] Report appended to: {xlsx_path} (sheet: {sheet})")

if __name__ == "__main__":
    main()


