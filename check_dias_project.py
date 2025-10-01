
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
check_dias_project.py
---------------------
功能：
- 讀取指定的 model.ini，搜尋是否有「DIAS_Project = true;」設定，且該行 *不* 以 # 為註解起始。
- 找到 → PASS；找不到 → FAIL。
- 輸出找到的設定原始行內容（未找到則顯示 N/A）。
- 產出 XLSX 報表（與 check_dias_4k60.py / check_dias_5k.py 一致的風格）：
  - 工作表依 model.ini 檔名前綴數字 → PID_<N>；否則 others
  - 表頭：Rules, Result, condition_1..condition_N（動態欄位，不自動補 N/A）
  - 欄寬統一、換行、垂直靠上；既有活頁簿附加資料

相容：Python 3.8+
"""

import argparse
import re
import sys
import os
from pathlib import Path

# -----------------------------
# Report helpers (aligned style, no dedicated Model/Panel columns, no N/A padding)
# -----------------------------

def _ensure_openpyxl():
    try:
        import openpyxl  # noqa
    except ImportError:
        raise SystemExit(
            "[ERROR] 需要 openpyxl 以支援報表輸出。\n"
            "安裝： pip install --user openpyxl\n"
        )

def _sheet_name_for_model(model_ini_path: str) -> str:
    """
    根據 model.ini 檔名的前綴決定頁簽：
      - 前綴是數字 N → 'PID_N'
      - 其他 → 'others'
    例: '1_EU_XXX.ini' → 'PID_1'
    """
    base = os.path.basename(model_ini_path or "")
    m = re.match(r"^(\d+)_", base)
    if m:
        return f"PID_{int(m.group(1))}"
    return "others"

def export_report(res: dict, xlsx_path: str = "kipling.xlsx", conditions: list = None) -> None:
    """
    動態輸出欄位，不做 N/A padding：
    表頭固定前兩欄: Rules, Result；
    後續依據實際條件數生成 condition_1..condition_N。
    若工作表已存在且表頭的 condition 欄不足本次需要，會自動擴增。
    """
    _ensure_openpyxl()
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    COMMON_WIDTH = 80
    COMMON_ALIGN = Alignment(wrap_text=True, vertical="top")
    BOLD = Font(bold=True)

    sheet_name = _sheet_name_for_model(res.get("model_ini", ""))
    conds = list(conditions or [])

    # 開啟或新建 xlsx
    try:
        wb = load_workbook(xlsx_path)
    except Exception:
        wb = Workbook()

    # 建立或取得 sheet
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # 若是空表，補 header
        if ws.max_row < 1:
            ws.append(["Rules", "Result"])
    else:
        ws = wb.create_sheet(title=sheet_name)
        ws.append(["Rules", "Result"])  # 先放兩欄，稍後依需要擴增

    # 依需要擴增 header 的 condition 欄位
    header = [c.value for c in ws[1]]
    current_cond_cols = max(0, len(header) - 2)
    needed_cond_cols = len(conds)
    if needed_cond_cols > current_cond_cols:
        new_header = ["Rules", "Result"] + [f"condition_{i}" for i in range(1, needed_cond_cols + 1)]
        ws.delete_rows(1)
        ws.append(new_header)

    # 準備資料
    rules  = res.get("rules", "DIAS_Project flag must be true (uncommented)")
    result = "PASS" if res.get("passed", False) else "FAIL"
    row_values = [rules, result] + conds

    # 寫入 row
    ws.append(row_values)
    last_row = ws.max_row

    # 套用樣式：欄寬、換行、垂直靠上（含表頭）
    total_cols = ws.max_column
    for col_idx in range(1, total_cols + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = COMMON_WIDTH

    for cell in ws[1]:
        cell.font = BOLD
        cell.alignment = COMMON_ALIGN

    for cell in ws[last_row]:
        cell.alignment = COMMON_ALIGN

    # 移除預設 Sheet
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        try:
            wb.remove(wb["Sheet"])
        except Exception:
            pass

    wb.save(xlsx_path)

# -----------------------------
# Core logic
# -----------------------------

# 只要該行不是以 '#' 開頭，且包含 DIAS_Project = true; 即視為有效（大小寫不敏感，允許空白）
DIAS_TRUE_RE = re.compile(r"\bDIAS_Project\s*=\s*true\s*;", re.IGNORECASE)

def find_dias_true(model_ini: Path) -> (bool, str):
    """
    搜尋未被 # 註解的行是否含有 'DIAS_Project = true;'。
    回傳 (found, matched_line_text or 'N/A')
    """
    try:
        with model_ini.open("r", encoding="utf-8", errors="ignore") as f:
            for raw in f:
                line = raw.rstrip("\n")
                # 忽略以 # 起始（含前置空白）的註解行
                if line.lstrip().startswith("#"):
                    continue
                if DIAS_TRUE_RE.search(line):
                    return True, line
    except FileNotFoundError:
        raise FileNotFoundError(f"Model ini not found: {model_ini}")
    return False, "N/A"

# -----------------------------
# Main
# -----------------------------

def main():
    ap = argparse.ArgumentParser(
        description="檢查 model.ini 是否宣告未註解之 'DIAS_Project = true;'，並輸出 XLSX 報表。"
    )
    ap.add_argument("--model-ini", required=True, help="model/*.ini 路徑")
    ap.add_argument("--report", action="store_true", help="輸出報表到 kipling.xlsx（若未提供 --report-xlsx）")
    ap.add_argument("--report-xlsx", metavar="FILE", help="自訂輸出報表路徑（.xlsx）")
    args = ap.parse_args()

    model_ini = Path(args.model_ini).resolve()

    try:
        found, matched = find_dias_true(model_ini)
    except Exception as e:
        print(f"[FAIL] 開啟/解析 model.ini 失敗：{e}")
        if args.report or args.report_xlsx:
            res = {
                "passed": False,
                "model_ini": str(model_ini),  # 僅用於決定分頁名稱
                "rules": "DIAS_Project flag must be true",
            }
            conditions = [
                "Matched line: N/A",
                "Errors: " + str(e),
            ]
            export_report(res, xlsx_path=(args.report_xlsx or "kipling.xlsx"), conditions=conditions)
        return 2

    print("=== DIAS_Project 檢查 ===")
    print(f"Model INI     : {model_ini}")
    print(f"Matched line  : {matched}")
    if found:
        print("[PASS] 找到未註解的 'DIAS_Project = true;'。")
    else:
        print("[FAIL] 未找到未註解的 'DIAS_Project = true;'。")

    # 報表輸出（對齊：無 Model/Panel 專屬欄位、動態 condition）
    if args.report or args.report_xlsx:
        res = {
            "passed": bool(found),
            "model_ini": str(model_ini),  # 用於分頁命名
            "rules": "DIAS_Project flag must be true (uncommented)",
        }
        conds = [
            f"Matched line: {matched}",
        ]
        export_report(res, xlsx_path=(args.report_xlsx or "kipling.xlsx"), conditions=conds)

    return 0 if found else 1

if __name__ == "__main__":
    sys.exit(main())
