import argparse
import os
import re
from typing import Dict, List, Tuple, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font

# Import sheet naming rule from the reference checker
try:
    import importlib.util, sys
    ref_path = os.path.join(os.path.dirname(__file__), "check_dvbs_satellite_flag.py")
    spec = importlib.util.spec_from_file_location("ref_checker", ref_path)
    ref_checker = importlib.util.module_from_spec(spec)
    sys.modules["ref_checker"] = ref_checker
    spec.loader.exec_module(ref_checker)  # type: ignore
    _sheet_name_for_model = ref_checker._sheet_name_for_model  # type: ignore
except Exception:
    # Fallback if import fails
    def _sheet_name_for_model(model_ini_path: str) -> str:
        base = os.path.basename(model_ini_path or "")
        m = re.match(r"^(\d+)_", base)
        if m:
            return f"PID_{int(m.group(1))}"
        return "others"


def _read_text(path: str) -> str:
    for enc in ("utf-8", "latin-1", "utf-16"):
        try:
            with open(path, "r", encoding=enc) as f:
                return f.read()
        except UnicodeDecodeError:
            continue
        except FileNotFoundError:
            return ""
    return ""


def _resolve_tvconfigs_path(root: str, tvconfigs_like: str) -> str:
    """Mirror the path resolution style of the reference checker."""
    if tvconfigs_like.startswith("/tvconfigs/"):
        rel = tvconfigs_like[len("/tvconfigs/"):]
        return os.path.normpath(os.path.join(root, rel))
    if tvconfigs_like.startswith("./") or tvconfigs_like.startswith("../"):
        return os.path.normpath(os.path.join(root, tvconfigs_like))
    if tvconfigs_like.startswith("/"):
        return tvconfigs_like
    return os.path.normpath(os.path.join(root, tvconfigs_like))


def _find_blocks_with_spans(xml_text: str) -> List[Tuple[Tuple[int, int], str]]:
    """
    Return list of ((start, end), block_text) for each <COUNTRY_TVCONFIG_MAP>...</COUNTRY_TVCONFIG_MAP> block.
    Case-insensitive, DOTALL.
    """
    blocks = []
    pattern = re.compile(r"<\s*COUNTRY_TVCONFIG_MAP\s*>.*?<\s*/\s*COUNTRY_TVCONFIG_MAP\s*>", re.IGNORECASE | re.DOTALL)
    for m in pattern.finditer(xml_text):
        blocks.append(((m.start(), m.end()), m.group(0)))
    return blocks


def _extract_tag(txt: str, tag: str) -> str:
    m = re.search(rf"<\s*{tag}\s*>(.*?)<\s*/\s*{tag}\s*>", txt, re.IGNORECASE | re.DOTALL)
    return (m.group(1).strip() if m else "")


def _line_number_at(text: str, index: int) -> int:
    """1-based line number for a given byte index in text."""
    return text.count("\n", 0, index) + 1


def check_country_tvconfig(xml_path: str, treat_dvb_prefix_as_dvb: bool = True, verbose: bool = False) -> Dict:
    """
    Perform the DVB HbbTV check:
      - Count DVB blocks.
      - For each DVB block, TV_CONFIG must include 'tv.config.dvb_hbbtv'. Otherwise record failure and line number.
    Returns a dictionary with details for reporting.
    """
    notes: List[str] = []
    if not os.path.exists(xml_path):
        notes.append("countryTvSysMap.xml not found")
        return {
            "xml_path": xml_path,
            "dvb_blocks": 0,
            "failed": [],
            "failed_count": 0,
            "error_lines": [],
            "notes": "; ".join(notes) if notes else "",
        }

    xml_text = _read_text(xml_path)
    if not xml_text.strip():
        notes.append("XML is empty or cannot be read")
        return {
            "xml_path": xml_path,
            "dvb_blocks": 0,
            "failed": [],
            "failed_count": 0,
            "error_lines": [],
            "notes": "; ".join(notes) if notes else "",
        }

    blocks = _find_blocks_with_spans(xml_text)
    if verbose:
        print(f"[INFO] Parsed COUNTRY_TVCONFIG_MAP blocks: {len(blocks)}")

    dvb_blocks = 0
    failed: List[Tuple[str, str, str, int]] = []  # (country, tv_system, tv_config, line_no)
    for (start, end), frag in blocks:
        tv_system = _extract_tag(frag, "TV_SYSTEM").upper()
        is_dvb = (tv_system == "DVB") or (treat_dvb_prefix_as_dvb and tv_system.startswith("DVB_"))
        if not is_dvb:
            continue
        dvb_blocks += 1

        tv_config_match = re.search(r"<\s*TV_CONFIG\s*>(.*?)<\s*/\s*TV_CONFIG\s*>", frag, re.IGNORECASE | re.DOTALL)
        tv_config_val = tv_config_match.group(1).strip() if tv_config_match else ""
        if "tv.config.dvb_hbbtv" not in tv_config_val:
            # Compute the absolute line number of <TV_CONFIG> open tag within the full file text
            tag_match = re.search(r"<\s*TV_CONFIG\b", frag, re.IGNORECASE)
            line_no = _line_number_at(xml_text, start + (tag_match.start() if tag_match else 0))
            failed.append((_extract_tag(frag, "COUNTRY_NAME") or "", tv_system, tv_config_val, line_no))

    error_lines = sorted({ln for *_, ln in failed})
    return {
        "xml_path": xml_path,
        "dvb_blocks": dvb_blocks,
        "failed": failed,
        "failed_count": len(failed),
        "error_lines": error_lines,
        "notes": "; ".join(notes) if notes else "",
    }


def export_report(res: Dict, model_ini: Optional[str], xlsx_path: str) -> None:
    """
    Write or append a row to an Excel sheet, following the style used by the reference checker:
      - Sheet name derives from model ini (PID_xxx or 'others').
      - Wrap text, align top, bold header. Wider first columns.
    Columns:
      A: Model.ini
      B: XML
      C: DVB 區塊數
      D: 錯誤的line number
      E: 失敗數
      F: 失敗項彙總
      G: Notes
    """
    COMMON_ALIGN = Alignment(wrap_text=True, vertical="top")
    BOLD = Font(bold=True)

    sheet_name = _sheet_name_for_model(model_ini or "")

    # Open workbook or create
    try:
        wb = load_workbook(xlsx_path)
    except Exception:
        wb = Workbook()

    # Use or create sheet
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(title=sheet_name)

    # Header
    headers = ["Model.ini", "XML", "DVB 區塊數", "錯誤的line number", "失敗數", "失敗項彙總", "Notes"]
    if ws.max_row == 1 and all(c.value is None for c in ws[1]):
        ws.append(headers)
        for cell in ws[1]:
            cell.font = BOLD
            cell.alignment = COMMON_ALIGN

    # Build row contents
    model_col = model_ini or ""
    xml_col = res.get("xml_path", "")
    dvb_count_col = res.get("dvb_blocks", 0)
    err_lines = res.get("error_lines", [])
    err_lines_col = ", ".join(str(n) for n in err_lines) if err_lines else ""
    fail_count_col = res.get("failed_count", 0)

    # Build failure summary: Country(tv_system): tv_config
    failed: List[Tuple[str, str, str, int]] = res.get("failed", [])
    if failed:
        parts = [f"{(country or '?') }({tv}): {tvconf or '(empty)'} [L{ln}]" for country, tv, tvconf, ln in failed]
        fail_summary_col = "\n".join(parts)
    else:
        fail_summary_col = ""

    notes_col = res.get("notes", "")

    ws.append([model_col, xml_col, dvb_count_col, err_lines_col, fail_count_col, fail_summary_col, notes_col])
    last_row = ws.max_row
    for cell in ws[last_row]:
        cell.alignment = COMMON_ALIGN

    # Column widths (approximate to reference script's style)
    # Wider for Model/XML and Summary
    widths = [40, 60, 12, 20, 10, 80, 60]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = w

    # Remove default sheet if present and more than one sheet exists
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        try:
            wb.remove(wb["Sheet"])
        except Exception:
            pass

    wb.save(xlsx_path)


def main():
    parser = argparse.ArgumentParser(description="Check DVB HbbTV tv.config in countryTvSysMap.xml and export Excel.")
    parser.add_argument("--root", required=True, help="The tvconfigs root folder (the path that maps to '/tvconfigs').")
    parser.add_argument("--model-ini", required=True, help="Model ini path (used only for sheet naming).")
    parser.add_argument("--xml", default="/tvconfigs/TvSysMap/countryTvSysMap.xml",
                        help="XML path (abs or relative or starting with /tvconfigs/). Default: /tvconfigs/TvSysMap/countryTvSysMap.xml")
    parser.add_argument("--report-xlsx", default="kipling.xlsx", help="Output Excel path (append). Default: kipling.xlsx")
    parser.add_argument("-v", "--verbose", action="store_true", help="Verbose logging.")
    args = parser.parse_args()

    xml_path = _resolve_tvconfigs_path(args.root, args.xml)
    if args.verbose:
        print(f"[INFO] Using XML: {xml_path}")

    res = check_country_tvconfig(xml_path=xml_path, treat_dvb_prefix_as_dvb=True, verbose=args.verbose)

    # Print console summary
    print("==== DVB HbbTV Check Summary ====")
    print(f"XML path         : {res['xml_path']}")
    print(f"DVB blocks       : {res['dvb_blocks']}")
    print(f"Failures         : {res['failed_count']}")
    if res['failed']:
        print("Failed items:")
        for country, tv_sys, tv_conf, line_no in res['failed']:
            print(f" - {country or '?'} ({tv_sys}) @L{line_no}: {tv_conf or '(empty)'}")
    if res['notes']:
        print(f"Notes            : {res['notes']}")

    export_report(res, args.model_ini, args.report_xlsx)
    print(f"[INFO] Report appended to: {args.report_xlsx} (sheet: {_sheet_name_for_model(args.model_ini)})")


if __name__ == "__main__":
    main()