#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
check_factory_menu_params.py

參考 tv_multi_standard_validation.py 的「路徑解析方式」與「報表格式」，
從指定的 model.ini 讀出以下三個參數，並將值寫入 Excel 報表：
- RtkFacktoryMenuComboKey
- RtkFacktoryMenuPackage
- RtkFacktoryMenuActivity

說明：
1) 路徑解析：支援將 "/tvconfigs/..." 映射到 --root 之下的對應路徑；
   其他相對路徑也會視為相對於 --root 解析。
2) 報表：沿用「Rules, Result, condition_1, ...」表頭，PID_前綴分頁（1→PID_1），
   若檔案存在則附加資料，欄寬/自動換行/垂直靠上與參考檔一致。
3) 判定：三個參數都找到 → PASS；任一缺少 → FAIL。

使用方法：
python3.8 check_factory_menu_params.py --model-ini model/1_xxx.ini --root . --report

作者：Auto-generated
"""

import argparse
import os
import re
from typing import Dict, List, Tuple, Optional

# -----------------------------
# Utilities for report (對齊 tv_multi_standard_validation.py 風格)
# -----------------------------

def _sheet_name_for_model(model_ini_path: str) -> str:
    """
    根據 model.ini 檔名的前綴決定頁簽：
      - 前綴是數字 N → 'PID_N'
      - 其他 → 'others'
    例: '1_EU_XXX.ini' → 'PID_1'
    """
    base = os.path.basename(model_ini_path or "")
    m = re.match(r"^(\d+)_", base)
    if m:
        return f"PID_{int(m.group(1))}"
    return "others"


def _ensure_openpyxl():
    try:
        import openpyxl  # noqa
    except ImportError:
        raise SystemExit(
            "[ERROR] 需要 openpyxl 以支援報表輸出與附加。\n"
            "  安裝： pip install --user openpyxl\n"
        )


def export_report(res: dict, xlsx_path: str = "kipling.xlsx", num_condition_cols: int = 5) -> None:
    """
    欄位無值時以 'N/A' 填入。依 model.ini 檔名前綴分頁（PID_1、PID_2…；非數字→others），既有資料則附加。
    表頭固定為: Rules, Result, condition_1, condition_2, condition_3, ...
    統一：所有欄位同寬、同為自動換行、垂直置頂（包含表頭）。
    """
    _ensure_openpyxl()
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    COMMON_WIDTH = 80
    COMMON_ALIGN = Alignment(wrap_text=True, vertical="top")
    BOLD = Font(bold=True)

    def _na(s: str) -> str:
        s = (s or "").strip()
        return s if s else "N/A"

    sheet_name = _sheet_name_for_model(res.get("model_ini", ""))

    # 開啟或新建 xlsx
    try:
        wb = load_workbook(xlsx_path)
    except Exception:
        wb = Workbook()

    # 建立或取得 sheet
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)
        headers = ["Rules", "Result"] + [f"condition_{i}" for i in range(1, num_condition_cols + 1)]
        ws.append(headers)

    # 準備資料
    rules    = "Check RtkFacktoryMenu* keys exist"
    result   = "PASS" if res.get("passed", False) else "FAIL"

    combo    = _na(res.get("combo_key", ""))
    pkg      = _na(res.get("package", ""))
    activity = _na(res.get("activity", ""))
    model    = _na(res.get("model_ini", ""))

    # condition values（與參考檔案一致：最多 num_condition_cols 欄）
    conds = [
        f"RtkFacktoryMenuComboKey = {combo}",     # condition_1
        f"RtkFacktoryMenuPackage  = {pkg}",       # condition_2
        f"RtkFacktoryMenuActivity = {activity}",  # condition_3
        f"model.ini = {model}",                   # condition_4（補充資訊，不影響判斷）
        "N/A",                                    # condition_5（保留欄）
    ][:num_condition_cols]

    # 寫入 row
    row_values = [rules, result] + conds
    ws.append(row_values)
    last_row = ws.max_row

    # 套用樣式：欄寬、換行、垂直靠上
    total_cols = 2 + num_condition_cols
    for col_idx in range(1, total_cols + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = COMMON_WIDTH

    for cell in ws[1]:  # header
        cell.font = BOLD
        cell.alignment = COMMON_ALIGN

    for cell in ws[last_row]:
        cell.alignment = COMMON_ALIGN

    # 移除預設 Sheet
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        try:
            wb.remove(wb["Sheet"])
        except Exception:
            pass

    wb.save(xlsx_path)


# -----------------------------
# Core parsing
# -----------------------------

def _read_text(path: str) -> str:
    for enc in ("utf-8", "latin-1", "utf-16"):
        try:
            with open(path, "r", encoding=enc) as f:
                return f.read()
        except UnicodeDecodeError:
            continue
        except FileNotFoundError:
            raise
    # 若皆失敗，讓原始例外往外拋
    with open(path, "r") as f:
        return f.read()


def _strip_comment(line: str) -> str:
    # 去掉 # 或 ; 之後的註解
    line = line.split("#", 1)[0]
    line = line.split(";", 1)[0]
    return line.strip()


def _resolve_tvconfigs_path(root: str, tvconfigs_like: str) -> str:
    """
    把 "/tvconfigs/xxx/yyy.ini" 映射為 "<root>/xxx/yyy.ini"
    其他相對路徑: 以 root 為基底
    絕對路徑（非 /tvconfigs 開頭）維持不動
    """
    if tvconfigs_like.startswith("/tvconfigs/"):
        rel = tvconfigs_like[len("/tvconfigs/"):]
        return os.path.normpath(os.path.join(root, rel))
    if tvconfigs_like.startswith("./") or tvconfigs_like.startswith("../"):
        return os.path.normpath(os.path.join(root, tvconfigs_like))
    if tvconfigs_like.startswith("/"):
        # 非 /tvconfigs 絕對路徑：原樣返回
        return tvconfigs_like
    # 其他當作相對於 root
    return os.path.normpath(os.path.join(root, tvconfigs_like))


def parse_factory_menu_keys(model_ini_path: str) -> Dict[str, Optional[str]]:
    """
    從 model.ini 內解析三個參數值（忽略大小寫、允許前後引號與空白）：
      - RtkFacktoryMenuComboKey
      - RtkFacktoryMenuPackage
      - RtkFacktoryMenuActivity
    同時也兼容「RtkFactoryMenuXXX」拼寫（避免 Facktory/Factory 混用）。
    """
    txt = _read_text(model_ini_path)

    # 建立一組正則，依序搜尋；第一個命中即採用
    patterns = {
        "combo_key": [
            r'^\s*RtkFacktoryMenuComboKey\s*=\s*"?([^"\r\n]+)"?\s*$',
            r'^\s*RtkFactoryMenuComboKey\s*=\s*"?([^"\r\n]+)"?\s*$',
        ],
        "package": [
            r'^\s*RtkFacktoryMenuPackage\s*=\s*"?([^"\r\n]+)"?\s*$',
            r'^\s*RtkFactoryMenuPackage\s*=\s*"?([^"\r\n]+)"?\s*$',
        ],
        "activity": [
            r'^\s*RtkFacktoryMenuActivity\s*=\s*"?([^"\r\n]+)"?\s*$',
            r'^\s*RtkFactoryMenuActivity\s*=\s*"?([^"\r\n]+)"?\s*$',
        ],
    }

    results: Dict[str, Optional[str]] = {"combo_key": None, "package": None, "activity": None}

    for raw in txt.splitlines():
        line = _strip_comment(raw)
        if not line:
            continue
        for key, p_list in patterns.items():
            if results[key] is not None:
                continue
            for p in p_list:
                m = re.match(p, line, re.IGNORECASE)
                if m:
                    results[key] = m.group(1).strip()
                    break

    return results


def build_result(model_ini: str, keys: Dict[str, Optional[str]]) -> Dict:
    """
    彙整報表欄位。
    """
    passed = all(keys.get(k) for k in ("combo_key", "package", "activity"))

    return {
        "passed": bool(passed),
        "model_ini": model_ini,
        "combo_key": keys.get("combo_key") or "",
        "package": keys.get("package") or "",
        "activity": keys.get("activity") or "",
    }


# -----------------------------
# Main
# -----------------------------

def main():
    parser = argparse.ArgumentParser(description="Check RtkFacktoryMenu* params from model.ini and export to Excel report (kipling.xlsx).")
    parser.add_argument("--model-ini", required=True, help="path to model ini (e.g., model/1_xxx.ini)")
    parser.add_argument("--root", required=True, help="tvconfigs project root (maps /tvconfigs/* to here)")
    parser.add_argument("-v", "--verbose", action="store_true", help="verbose logs")

    # 報表參數：--report 與 --report-xlsx（任一存在即輸出；未指定路徑則用 kipling.xlsx）
    parser.add_argument("--report", action="store_true", help="export report to xlsx (default: kipling.xlsx)")
    parser.add_argument("--report-xlsx", metavar="FILE", help="export report to specific xlsx file")

    args = parser.parse_args()

    model_ini = args.model_ini
    if not os.path.exists(model_ini):
        raise SystemExit(f"[ERROR] model ini not found: {model_ini}")

    root = os.path.abspath(os.path.normpath(args.root))

    if args.verbose:
        print(f"[INFO] model_ini: {model_ini}")
        print(f"[INFO] root     : {root}")

    # 解析三個參數
    keys = parse_factory_menu_keys(model_ini)
    if args.verbose:
        print("[INFO] Parsed keys:",
              f"ComboKey={keys.get('combo_key') or '(N/A)'};",
              f"Package={keys.get('package') or '(N/A)'};",
              f"Activity={keys.get('activity') or '(N/A)'}")

    # 產生結果
    res = build_result(model_ini, keys)

    # Console 輸出
    print(f"Result  : {'PASS' if res['passed'] else 'FAIL'}")
    print(f"ComboKey: {res['combo_key'] or '(N/A)'}")
    print(f"Package : {res['package'] or '(N/A)'}")
    print(f"Activity: {res['activity'] or '(N/A)'}")

    # 報表輸出
    if args.report or args.report_xlsx:
        xlsx_path = args.report_xlsx if args.report_xlsx else "kipling.xlsx"
        export_report(res, xlsx_path=xlsx_path)
        sheet = _sheet_name_for_model(model_ini)
        print(f"[INFO] Report appended to: {xlsx_path} (sheet: {sheet})")


if __name__ == "__main__":
    main()
