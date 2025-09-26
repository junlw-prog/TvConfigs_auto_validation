# -*- coding: utf-8 -*-
"""
check_pq_osdtable_dolby_colorspace.py

功能:
- 從 model.ini 讀取 PQ_OSD 指向的 OSDTable.ini
- 檢查 OSDTable.ini 內
  [Dolby_Dark] -> ColorSpace == 0
  [Dolby_IQ]   -> ColorSpace == 0
  兩者皆為 0 才判定 PASS，否則 FAIL

重要路徑規則:
- 若 PQ_OSD = "/tvconfigs/PQ_OSD/OSDTable.ini"
  實際檔案位置 = <root_dir>/PQ_OSD/OSDTable.ini
  (也就是把 "/tvconfigs/" 去掉，剩下的相對於 root_dir)

輸出:
- 預設將結果印在 console
- 若提供 --report，則將結果寫入 kipling.xlsx
  表頭: Rules, Result, condition_1, condition_2
  (格式: 首列粗體、所有欄位等寬、文字自動換行、垂直靠上)

使用範例:
python3.8 check_pq_osdtable_dolby_colorspace.py \
  --model-ini model/21_WW_xxx.ini --root . --report -v
"""

import os
import re
import sys
import argparse
from pathlib import Path
from typing import Dict, Tuple, Optional

# ========== 工具函式 ==========

def smart_read_text(path: Path, encodings=("utf-8", "latin-1", "utf-16")) -> str:
    last_err = None
    for enc in encodings:
        try:
            return path.read_text(encoding=enc)
        except UnicodeDecodeError as e:
            last_err = e
            continue
        except Exception as e:
            last_err = e
            continue
    if last_err:
        raise last_err
    raise RuntimeError(f"Unknown read error: {path}")

def parse_simple_ini(text: str) -> Dict[str, Dict[str, str]]:
    data: Dict[str, Dict[str, str]] = {}
    current = None
    section_re = re.compile(r'^\s*\[(?P<name>[^\]]+)\]\s*$')
    kv_re = re.compile(r'^\s*(?P<key>[^=:#]+?)\s*[:=]\s*(?P<val>.*?)\s*$')

    for line in text.splitlines():
        line_stripped = line.strip()
        if not line_stripped or line_stripped.startswith(("#", ";")):
            continue
        m = section_re.match(line)
        if m:
            current = m.group("name").strip()
            data.setdefault(current, {})
            continue
        m = kv_re.match(line)
        if m and current is not None:
            key = m.group("key").strip().lower().replace(" ", "")
            val = m.group("val").strip()
            if "#" in val:
                val = val.split("#", 1)[0].strip()
            if ";" in val:
                val = val.split(";", 1)[0].strip()
            data[current][key] = val
    return data

def extract_value(d: Dict[str, Dict[str, str]], section: str, key: str) -> Optional[str]:
    sec = d.get(section)
    if not sec:
        return None
    return sec.get(key.lower().replace(" ", ""))

def find_pq_osd_path_from_model(model_ini_text: str) -> Optional[str]:
    # 先用簡易解析器嘗試在與 "misc_pq_map_cfg" 名稱相近的 section 找 pq_osd
    ini = parse_simple_ini(model_ini_text)
    cand_sections = [k for k in ini.keys() if k.strip().lower() in {"misc_pq_map_cfg", "misc_pq", "misc_pq_map"}]
    for sec in cand_sections:
        v = ini[sec].get("pq_osd")
        if v:
            return v.strip().strip('"').strip("'")
    # 退回全文 regex
    m = re.search(r'(?im)^\s*PQ_OSD\s*=\s*(?P<path>.+?)\s*$', model_ini_text)
    if m:
        return m.group("path").strip().strip('"').strip("'")
    return None

def to_abs_under_root(root: Path, candidate: str) -> Path:
    """
    映射規則:
    - 以 "/tvconfigs/" 開頭 => 去掉此前綴，當作 root 下的相對路徑
      例: "/tvconfigs/PQ_OSD/OSDTable.ini" -> <root>/PQ_OSD/OSDTable.ini
    - 以 "/" 開頭但不是 "/tvconfigs/" => 去掉前導 "/"，掛到 root 下
    - 其他 => 視為相對於 root
    """
    c = candidate.strip()
    if c.startswith("/tvconfigs/"):
        rel = c[len("/tvconfigs/"):]
        return (root / rel).resolve()
    if c.startswith("/"):
        return (root / c.lstrip("/")).resolve()
    return (root / c).resolve()

# ========== 報表輸出 ==========

def append_report_row(xlsx_path: Path, sheet_name: str, row: Dict[str, str]) -> None:
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except Exception as e:
        raise RuntimeError("需要 openpyxl 來輸出 xlsx，請先安裝: pip install openpyxl") from e

    headers = ["Rules", "Result", "condition_1", "condition_2"]
    if xlsx_path.exists():
        wb = load_workbook(str(xlsx_path))
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
        # 若是空白 sheet，補上表頭
        if ws.max_row == 1 and ws.max_column == 1 and ws["A1"].value is None:
            ws.append(headers)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)

    ws = wb[sheet_name]
    values = [row.get(h, "N/A") for h in headers]
    ws.append(values)
    last_row = ws.max_row

    # 給儲存格指派上色
    rules_color = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
    failed_color = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")
    # 上色
    first_cell = ws.cell(row=last_row, column=1)  # 欄位1對應的是 'A' 列
    first_cell.fill = rules_color
    if row["Result"] == "FAIL":
        ws.cell(row=last_row, column=2).fill = failed_color
    if row["condition_1"] == "[Dolby_Dark] ColorSpace = N/A":
        ws.cell(row=last_row, column=3).fill = failed_color
    if row["condition_2"] == "[Dolby_IQ]   ColorSpace = N/A":
        ws.cell(row=last_row, column=4).fill = failed_color

    # 統一格式
    col_width = 38
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[chr(64 + col)].width = col_width
    for r in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in r:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(str(xlsx_path))

# ========== 主流程 ==========

def check_colorspace(osd_ini_path: Path, verbose=False) -> Tuple[Optional[int], Optional[int], str]:
    if not osd_ini_path.exists():
        return None, None, "FAIL"  # 檔案不存在 => FAIL
    text = smart_read_text(osd_ini_path)
    ini = parse_simple_ini(text)

    dark_cs_str = extract_value(ini, "Dolby_Dark", "ColorSpace")
    iq_cs_str   = extract_value(ini, "Dolby_IQ",   "ColorSpace")

    def to_int_or_none(s: Optional[str]) -> Optional[int]:
        if s is None:
            return None
        s = s.strip().strip('"').strip("'")
        try:
            return int(s, 0)
        except Exception:
            return None

    dark_cs = to_int_or_none(dark_cs_str)
    iq_cs = to_int_or_none(iq_cs_str)

    if verbose:
        print(f"[VERBOSE] OSDTable: {osd_ini_path}")
        print(f"[VERBOSE] [Dolby_Dark] ColorSpace -> {dark_cs_str} (parsed {dark_cs})")
        print(f"[VERBOSE] [Dolby_IQ]   ColorSpace -> {iq_cs_str} (parsed {iq_cs})")

    if dark_cs is None or iq_cs is None:
        verdict = "FAIL"  # 找不到其中之一也視為 FAIL（按你的新規則）
    else:
        verdict = "PASS" if (dark_cs == 0 and iq_cs == 0) else "FAIL"
    return dark_cs, iq_cs, verdict

def get_sheet_name(model_ini: Path) -> str:
    base = model_ini.stem
    prefix = base.split("_", 1)[0]
    if prefix.isdigit():
        return f"PID_{prefix}"
    return "PQ_OSDTable_Dolby_ColorSpace"

def main():
    ap = argparse.ArgumentParser(description="Check [Dolby_Dark]/[Dolby_IQ] ColorSpace == 0 in OSDTable.ini")
    ap.add_argument("--model-ini", required=True, help="model.ini 路徑")
    ap.add_argument("--root", default=".", help="專案根目錄，預設為當前目錄")
    ap.add_argument("--report", action="store_true", help="輸出報表 kipling.xlsx")
    ap.add_argument("-v", action="store_true", help="顯示詳細判斷過程")
    args = ap.parse_args()

    root = Path(args.root).resolve()
    model_ini = Path(args.model_ini).resolve()

    if not model_ini.exists():
        print(f"[ERROR] model.ini 不存在: {model_ini}")
        sys.exit(2)

    model_text = smart_read_text(model_ini)
    pq_osd_val = find_pq_osd_path_from_model(model_text)
    rules = "9. 先確認 OSDTable.ini 使用的檔案是那一個\n" \
            "    - 確認 所有的 Dolby_xxx 相關的區塊 ColorSpace 是否都有 off"

    if not pq_osd_val:
        if args.v:
            print("[VERBOSE] PQ_OSD 未在 model.ini 中宣告")
        verdict = "FAIL"
        cond1 = "[Dolby_Dark] ColorSpace = N/A"
        cond2 = "[Dolby_IQ]   ColorSpace = N/A"
        print("Rules:", rules)
        print("Result:", verdict)
        print("condition_1:", cond1)
        print("condition_2:", cond2)
        if args.report:
            xlsx_path = Path("kipling.xlsx").resolve()
            sheet = get_sheet_name(model_ini)
            append_report_row(xlsx_path, sheet, {"Rules": rules, "Result": verdict, "condition_1": cond1, "condition_2": cond2})
            print(f"[INFO] 已寫入報表: {xlsx_path} (sheet: {sheet})")
        sys.exit(1)

    osd_ini_path = to_abs_under_root(root, pq_osd_val)

    if args.v:
        print(f"[VERBOSE] 解析 PQ_OSD: {pq_osd_val} -> {osd_ini_path}")

    dark_cs, iq_cs, verdict = check_colorspace(osd_ini_path, verbose=args.v)

    cond1 = f"[Dolby_Dark] ColorSpace = {dark_cs if dark_cs is not None else 'N/A'}"
    cond2 = f"[Dolby_IQ]   ColorSpace = {iq_cs if iq_cs is not None else 'N/A'}"

    print("Rules:", rules)
    print("Result:", verdict)
    print("condition_1:", cond1)
    print("condition_2:", cond2)

    if args.report:
        row = {"Rules": rules, "Result": verdict, "condition_1": cond1, "condition_2": cond2}
        xlsx_path = Path("kipling.xlsx").resolve()
        sheet = get_sheet_name(model_ini)
        append_report_row(xlsx_path, sheet, row)
        print(f"[INFO] 已寫入報表: {xlsx_path} (sheet: {sheet})")

if __name__ == "__main__":
    main()
