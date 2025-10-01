#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
check_dap_virtualizer_mode.py

功能：
- 參考 tv_multi_standard_validation.py 的路徑映射、xlsx輸出格式、sheet 命名規則
- 在 xlsx 內不輸出 model.ini 欄位
- 從 model.ini 取得 DAP_Sound_Param 的值，打開指向檔案後搜尋 virtualizer_mode 參數：
    * virtualizer_mode == 1 → PASS
    * virtualizer_mode == 0 → FAIL
- 其他狀況（檔案不存在、未找到鍵值、解析失敗）皆視為 FAIL，並在 Notes / Missing 欄位說明

Python 3.8+
依賴：openpyxl（僅在 --report 或 --report-xlsx 使用時需要）
"""
import argparse
import os
import re
from typing import Dict, List, Tuple, Optional


# -----------------------------
# Utilities for report (參考既有專案風格)
# -----------------------------

def _sheet_name_for_model(model_ini_path: str) -> str:
    """
    根據 model.ini 檔名的前綴決定頁簽：
      - 前綴是數字 N → 'PID_N'
      - 其他 → 'others'
    例: '1_EU_XXX.ini' → 'PID_1'
    """
    base = os.path.basename(model_ini_path or "")
    m = re.match(r"^(\d+)_", base)
    if m:
        return f"PID_{int(m.group(1))}"
    return "others"


def _ensure_openpyxl():
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        raise SystemExit(
            "[ERROR] 需要 openpyxl 以支援報表輸出與附加。\n"
            "  安裝： pip install --user openpyxl\n"
        )


def export_report(res: dict, xlsx_path: str = "kipling.xlsx", num_condition_cols: int = 5) -> None:
    """
    欄位無值時以 'N/A' 填入。依 model.ini 檔名前綴分頁（PID_1、PID_2…；非數字→others），既有資料則附加。
    表頭固定為: Rules, Result, condition_1, condition_2, condition_3, ...
    統一：所有欄位同寬、同為自動換行、垂直置頂（包含表頭）。
    *本報表不輸出 model.ini 欄位*
    """
    _ensure_openpyxl()
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    COMMON_WIDTH = 80
    COMMON_ALIGN = Alignment(wrap_text=True, vertical="top")
    BOLD = Font(bold=True)

    def _na(s: Optional[str]) -> str:
        s = (s or "").strip()
        return s if s else "N/A"

    sheet_name = _sheet_name_for_model(res.get("model_ini_path", ""))

    # 開啟或新建 xlsx
    try:
        wb = load_workbook(xlsx_path)
    except Exception:
        wb = Workbook()

    # 建立或取得 sheet
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)
        headers = ["Rules", "Result"] + [f"condition_{i}" for i in range(1, num_condition_cols + 1)]
        ws.append(headers)

    # 準備資料
    rules = "1) 解析 DAP_Sound_Param → 2) 開啟該檔案 → 3) virtualizer_mode 是否為 1 ?"
    result = "PASS" if res.get("passed", False) else "FAIL"

    dap_path = _na(res.get("dap_path_resolved"))
    virt_val = res.get("virtualizer_mode_value")
    virt_text = "N/A" if virt_val is None else str(virt_val)
    decision = _na(res.get("decision", ""))
    notes = _na(res.get("notes", ""))
    missing = _na(", ".join(res.get("missing", []) or []))

    conds = [
        f"DAP_Sound_Param = {dap_path}",       # condition_1
        f"virtualizer_mode = {virt_text}",     # condition_2
        f"Decision = {decision}",              # condition_3
        f"Notes = {notes}",                    # condition_4
        f"Missing = {missing}",                # condition_5
    ][:num_condition_cols]

    # 寫入 row
    row_values = [rules, result] + conds
    ws.append(row_values)
    last_row = ws.max_row

    # 套用樣式：欄寬、換行、垂直靠上
    total_cols = 2 + num_condition_cols
    for col_idx in range(1, total_cols + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = COMMON_WIDTH

    for cell in ws[1]:  # header
        cell.font = BOLD
        cell.alignment = COMMON_ALIGN

    for cell in ws[last_row]:
        cell.alignment = COMMON_ALIGN

    # 移除預設 Sheet
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        try:
            wb.remove(wb["Sheet"])
        except Exception:
            pass

    wb.save(xlsx_path)


# -----------------------------
# Core parsing / validation
# -----------------------------

def _read_text(path: str) -> str:
    for enc in ("utf-8", "latin-1", "utf-16"):
        try:
            with open(path, "r", encoding=enc) as f:
                return f.read()
        except UnicodeDecodeError:
            continue
        except FileNotFoundError:
            raise
    with open(path, "r") as f:
        return f.read()


def _strip_comment(line: str) -> str:
    # 去掉 # 或 ; 之後的註解
    line = line.split("#", 1)[0]
    line = line.split(";", 1)[0]
    return line.strip()


def _resolve_tvconfigs_path(root: str, tvconfigs_like: str) -> str:
    """
    把 "/tvconfigs/xxx/yyy.ini" 映射為 "<root>/xxx/yyy.ini"
    其他相對路徑: 以 root 為基底
    絕對路徑（非 /tvconfigs 開頭）維持不動
    """
    if tvconfigs_like.startswith("/tvconfigs/"):
        rel = tvconfigs_like[len("/tvconfigs/"):]
        return os.path.normpath(os.path.join(root, rel))
    if tvconfigs_like.startswith("./") or tvconfigs_like.startswith("../"):
        return os.path.normpath(os.path.join(root, tvconfigs_like))
    if tvconfigs_like.startswith("/"):
        # 非 /tvconfigs 絕對路徑：原樣返回
        return tvconfigs_like
    # 其他當作相對於 root
    return os.path.normpath(os.path.join(root, tvconfigs_like))


def _find_key_value_in_ini_text(text: str, key: str) -> Optional[str]:
    """
    在一般 ini/kv 檔案中找 key=value；忽略註解與空白；大小寫不敏感；允許有引號。
    回傳 value（未去引號）。若未找到回傳 None。
    """
    key_re = re.compile(r'^\s*' + re.escape(key) + r'\s*=\s*"?([^"\n\r]+)"?\s*$', re.IGNORECASE)
    for raw in text.splitlines():
        line = _strip_comment(raw)
        if not line:
            continue
        m = key_re.match(line)
        if m:
            return m.group(1).strip()
    return None


def parse_model_ini_for_dap(model_ini_path: str, root: str) -> Optional[str]:
    """
    從 model.ini 找 DAP_Sound_Param = "<path>" 並映射到 --root
    """
    txt = _read_text(model_ini_path)
    val = _find_key_value_in_ini_text(txt, "DAP_Sound_Param")
    if val is None:
        return None
    return _resolve_tvconfigs_path(root, val)


def read_virtualizer_mode(dap_param_path: str) -> Optional[int]:
    """
    在 DAP_Sound_Param 指向的檔案中找 virtualizer_mode；可接受：
      - virtualizer_mode=1
      - virtualizer_mode = 0
      - 允許大小寫忽略/引號
    成功則回傳 int(0/1)，找不到回傳 None。
    """
    if not dap_param_path or not os.path.exists(dap_param_path):
        return None
    txt = _read_text(dap_param_path)
    val = _find_key_value_in_ini_text(txt, "virtualizer_mode")
    if val is None:
        return None
    # 取出數字部分
    m = re.match(r"^\s*([01])\s*$", val)
    if not m:
        # 若值不是 0/1，仍嘗試以 int 解析（容錯），失敗則 None
        try:
            return int(val.strip())
        except Exception:
            return None
    return int(m.group(1))


def build_result(args, model_ini: str, dap_path: Optional[str], virt_val: Optional[int]) -> Dict:
    """
    產生報表所需的結果結構
    """
    missing: List[str] = []
    notes: List[str] = []

    if dap_path is None:
        notes.append("model.ini 未找到 DAP_Sound_Param")
    elif not os.path.exists(dap_path):
        missing.append(dap_path)

    decision = ""
    passed = False
    if virt_val is None:
        decision = "virtualizer_mode 未找到或格式異常 → FAIL"
        passed = False
    elif virt_val == 1:
        decision = "virtualizer_mode == 1 → PASS"
        passed = True
    elif virt_val == 0:
        decision = "virtualizer_mode == 0 → FAIL"
        passed = False
    else:
        decision = f"virtualizer_mode 非 0/1（{virt_val}）→ FAIL"
        passed = False

    return {
        "passed": bool(passed),
        "model_ini_path": model_ini,            # 僅用於決定 sheet 名稱，不輸出到表格
        "dap_path_resolved": dap_path or "",
        "virtualizer_mode_value": virt_val,
        "decision": decision,
        "notes": "; ".join(notes),
        "missing": missing,
    }


# -----------------------------
# Main
# -----------------------------

def main():
    parser = argparse.ArgumentParser(description="Check DAP virtualizer_mode from model.ini → DAP_Sound_Param (Excel report style aligned with tv_multi_standard_validation.py).")
    parser.add_argument("--model-ini", required=True, help="path to model ini (e.g., model/1_xxx.ini)")
    parser.add_argument("--root", required=True, help="tvconfigs project root (maps /tvconfigs/* to here)")
    parser.add_argument("-v", "--verbose", action="store_true", help="verbose logs")

    # 報表參數
    parser.add_argument("--report", action="store_true", help="export report to xlsx (default: kipling.xlsx)")
    parser.add_argument("--report-xlsx", metavar="FILE", help="export report to specific xlsx file")

    args = parser.parse_args()

    model_ini = args.model_ini
    if not os.path.exists(model_ini):
        raise SystemExit(f"[ERROR] model ini not found: {model_ini}")

    root = os.path.abspath(os.path.normpath(args.root))

    if args.verbose:
        print(f"[INFO] model_ini: {model_ini}")
        print(f"[INFO] root     : {root}")

    # 解析 model.ini → DAP_Sound_Param
    dap_path = parse_model_ini_for_dap(model_ini, root)
    if args.verbose:
        print(f"[INFO] DAP_Sound_Param → {dap_path if dap_path else '(not found)'}")

    # 讀取 virtualizer_mode
    virt_val = read_virtualizer_mode(dap_path) if dap_path else None
    if args.verbose:
        print(f"[INFO] virtualizer_mode = {virt_val if virt_val is not None else '(not found)'}")

    # 結果
    res = build_result(args, model_ini, dap_path, virt_val)

    # Console summary
    print(f"Result  : {'PASS' if res['passed'] else 'FAIL'}")
    print(f"Decision: {res['decision']}")
    if res.get("notes"):
        print(f"Notes   : {res['notes']}")
    if res.get("missing"):
        print(f"Missing : {', '.join(res['missing'])}")

    # 報表輸出
    if args.report or args.report_xlsx:
        xlsx_path = args.report_xlsx if args.report_xlsx else "kipling.xlsx"
        export_report(res, xlsx_path=xlsx_path)
        sheet = _sheet_name_for_model(model_ini)
        print(f"[INFO] Report appended to: {xlsx_path} (sheet: {sheet})")


if __name__ == "__main__":
    main()
