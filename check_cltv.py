#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
check_cltv.py

依照 check_cltv.py 的檢查邏輯：
  1) 讀取 model.ini，尋找 LaunchCLTVByCountry = "<path>"
  2) 將 /tvconfigs/ 前綴路徑映射到 --root 目錄下的實際檔案路徑
  3) 檢查該檔案是否存在

報表格式「參考 tv_multi_standard_validation.py」：
  - 首列表頭固定順序：Rules, Result, condition_1, condition_2, condition_3, ...
  - 表頭粗體、所有欄位同寬、換行、垂直靠上（含表頭與資料列）
  - Result 僅輸出 PASS / FAIL / N/A（不含 "Result = " 前綴）
  - 不輸出 Model.ini 欄位（如需可自行在 conditions 中加入）
"""
import argparse
import os
import re
from typing import Optional

# -----------------------------
# Utilities for report
# -----------------------------
def _sheet_name_for_model(model_ini_path: str) -> str:
    import os, re
    base = os.path.basename(model_ini_path or "")
    m = re.match(r"^(\d+)_", base)
    if m:
        return f"PID_{int(m.group(1))}"
    return "others"


def _ensure_openpyxl():
    try:
        import openpyxl  # noqa
    except ImportError:
        raise SystemExit(
            "[ERROR] 需要 openpyxl 以支援報表輸出與附加。\n"
            "  安裝： pip install --user openpyxl\n"
        )


def export_report(res: dict, xlsx_path: str = "kipling.xlsx", num_condition_cols: int = 5) -> None:
    """
    表頭固定為: Rules, Result, condition_1, condition_2, condition_3, ...
    欄位無值時以 'N/A' 填入。依 model.ini 檔名前綴分頁（PID_1、PID_2…；非數字→others），既有資料則附加。
    全欄統一樣式：同寬、換行、垂直置頂（包含表頭）。
    """
    _ensure_openpyxl()
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    COMMON_WIDTH = 80
    COMMON_ALIGN = Alignment(wrap_text=True, vertical="top")
    BOLD = Font(bold=True)

    def _na(s: str) -> str:
        s = (s or "").strip()
        return s if s else "N/A"

    sheet_name = _sheet_name_for_model(res.get("model_ini", ""))

    # 開啟或新建 xlsx
    try:
        wb = load_workbook(xlsx_path)
    except Exception:
        wb = Workbook()

    # 建立或取得 sheet（表頭固定順序）
    header = ["Rules", "Result"] + [f"condition_{i}" for i in range(1, num_condition_cols + 1)]
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row < 1:
            ws.append(header)
    else:
        ws = wb.create_sheet(title=sheet_name)
        ws.append(header)

    # 取值
    rules      = _na(res.get("rules", ""))
    result     = _na(res.get("result", ""))
    conditions = [ _na(x) for x in (res.get("conditions", []) or []) ]
    """
    # 補足 condition_* 欄位數
    if len(conditions) < num_condition_cols:
        conditions += ["N/A"] * (num_condition_cols - len(conditions))
    else:
        conditions = conditions[:num_condition_cols]
    """
    # 寫入一列
    row_values = [rules, result] + conditions
    ws.append(row_values)
    last_row = ws.max_row

    # ── 統一樣式：所有欄位同寬 & 換行 & 垂直置頂（含表頭） ──
    total_cols = 2 + num_condition_cols
    for col_idx in range(1, total_cols + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = COMMON_WIDTH

    # 表頭樣式
    for cell in ws[1]:
        cell.font = BOLD
        cell.alignment = COMMON_ALIGN

    # 資料列樣式（最新一列）
    for col_idx in range(1, total_cols + 1):
        ws.cell(row=last_row, column=col_idx).alignment = COMMON_ALIGN

    # 移除預設空白 Sheet（若存在且非唯一）
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        try:
            wb.remove(wb["Sheet"])
        except Exception:
            pass

    wb.save(xlsx_path)


# -----------------------------
# Core logic
# -----------------------------
def _read_text_lines(path: str):
    for enc in ("utf-8", "latin-1", "utf-16"):
        try:
            with open(path, "r", encoding=enc, errors="ignore") as f:
                return f.readlines()
        except UnicodeDecodeError:
            continue
        except FileNotFoundError:
            raise
    with open(path, "r") as f:
        return f.readlines()


def _strip_comment(line: str) -> str:
    # 僅去掉行首 # 註解
    return line if not line.lstrip().startswith("#") else ""


def _resolve_tvconfigs_path(root: str, tvconfigs_like: str) -> str:
    """
    把 "/tvconfigs/xxx/yyy" 映射為 "<root>/xxx/yyy"
    其他相對路徑: 以 root 為基底
    絕對路徑（非 /tvconfigs 開頭）維持不動
    """
    tvconfigs_like = tvconfigs_like.strip()
    if tvconfigs_like.startswith("/tvconfigs/"):
        rel = tvconfigs_like[len("/tvconfigs/"):]
        return os.path.normpath(os.path.join(root, rel))
    if tvconfigs_like.startswith("./") or tvconfigs_like.startswith("../"):
        return os.path.normpath(os.path.join(root, tvconfigs_like))
    if tvconfigs_like.startswith("/"):
        # 非 /tvconfigs 絕對路徑：原樣返回
        return tvconfigs_like
    # 其他當作相對於 root
    return os.path.normpath(os.path.join(root, tvconfigs_like))


def parse_model_ini_for_launch_cltv(model_ini_path: str) -> Optional[str]:
    """
    從 model.ini 找：第一條未被 # 註解的 LaunchCLTVByCountry = "<value>"
    回傳原始 value（可含 /tvconfigs 前綴）；若未找到回傳 None。
    """
    lines = _read_text_lines(model_ini_path)
    for raw in lines:
        line = _strip_comment(raw)
        if not line:
            continue
        if "LaunchCLTVByCountry" in line and "=" in line:
            m = re.search(r'LaunchCLTVByCountry\s*=\s*"([^"]*)"', line)
            if m:
                return m.group(1).strip()
            # 容忍未加引號
            m2 = re.search(r'LaunchCLTVByCountry\s*=\s*(\S+)', line)
            if m2:
                return m2.group(1).strip()
            # 格式不正確時，回傳空字串以供後續 FAIL 判定
            return ""
    return None


def main():
    parser = argparse.ArgumentParser(description="Check LaunchCLTVByCountry in model.ini and export report (tv_multi_standard_validation.py style).")
    parser.add_argument("--root", required=True, help="專案根目錄路徑（將 /tvconfigs/* 映射到此）")
    parser.add_argument("--model-ini", required=True, help="model.ini 檔案路徑（例如 model/1_xxx.ini）")
    parser.add_argument("--report", action="store_true", help="輸出報表到 xlsx（預設檔名 kipling.xlsx）")
    parser.add_argument("--report-xlsx", metavar="FILE", help="指定報表 xlsx 檔案名稱")
    parser.add_argument("--conditions", type=int, default=5, help="condition_* 欄位數，預設 5")
    parser.add_argument("-v", "--verbose", action="store_true", help="顯示詳細過程")
    args = parser.parse_args()

    model_ini = args.model_ini
    if not os.path.exists(model_ini):
        raise SystemExit(f"[ERROR] model ini not found: {model_ini}")

    root = os.path.abspath(os.path.normpath(args.root))

    if args.verbose:
        print(f"[INFO] model_ini: {model_ini}")
        print(f"[INFO] root     : {root}")

    # 1) 擷取設定值
    raw_value = parse_model_ini_for_launch_cltv(model_ini)

    if raw_value is None:
        if args.verbose:
            print("[INFO] LaunchCLTVByCountry 未宣告（或只有註解）→ N/A")
        result = "N/A"
        resolved = ""
        exists_text = "N/A"
    else:
        if raw_value == "":
            if args.verbose:
                print("[WARN] LaunchCLTVByCountry 格式錯誤或值為空 → FAIL")
            result = "FAIL"
            resolved = ""
            exists_text = "N/A"
        else:
            # 2) 映射到實際路徑
            resolved = _resolve_tvconfigs_path(root, raw_value)
            # 3) 檢查檔案存在
            exists = os.path.exists(resolved)
            exists_text = "Yes" if exists else "No"
            result = "PASS" if exists else "FAIL"

    # 螢幕輸出（比對步驟）
    print("=== LaunchCLTVByCountry Check ===")
    print(f"Model.ini : {model_ini}")
    print(f"Setting   : {raw_value if raw_value is not None else 'N/A'}")
    print(f"Resolved  : {resolved if resolved else 'N/A'}")
    print(f"Exists?   : {exists_text}")
    print(f"Result    : {result}")

    # 準備報表資料（符合新格式）
    rules = "LaunchCLTVByCountry declared?\nResolved path exists?"
    conditions = [
        f'LaunchCLTVByCountry = {raw_value if raw_value is not None and raw_value != "" else "N/A"}',  # condition_1
        #f"Resolved Path = {resolved if resolved else 'N/A'}",                                          # condition_2
        f"File Exists = {exists_text}",                                                                # condition_3
        # 如需加入更多資訊，可在此繼續擴充 condition_4, condition_5, ...
    ]

    res = {
        "result": result,       # PASS / FAIL / N/A
        "rules": rules,         # Rules 欄位內容
        "model_ini": model_ini, # 用於決定分頁（PID_1..others），不會直接輸出欄位
        "conditions": conditions,
    }

    # 報表輸出
    if args.report or args.report_xlsx:
        xlsx_path = args.report_xlsx if args.report_xlsx else "kipling.xlsx"
        export_report(res, xlsx_path=xlsx_path, num_condition_cols=args.conditions)
        sheet = _sheet_name_for_model(model_ini)
        print(f"[INFO] Report appended to: {xlsx_path} (sheet: {sheet})")


if __name__ == "__main__":
    main()
