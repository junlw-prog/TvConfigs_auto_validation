#!/usr/bin/env python3.8
# -*- coding: utf-8 -*-
"""
  pid1_config_check_v3.py
  獨立檢核：PID 1 (Model_1 / Board_1)

  功能概述
  1) 讀取 sys/*.ini（可選：device/*/*/sys/*/device_sys.ini）
  2) 驗證 Model_1/Board_1 指向的檔案存在（Board_1 可選；若定義必須存在）
  3) 從 model.ini 與其引用的 country/、TvSysMap/ 檔蒐集國家碼
  4) 規則：僅允許歐洲國家（預設 EU+EFTA+GB+CH），且不得包含非 DVB 國家（預設 deny-list）
  5) 額外資訊：CLTV 是否整合、多制式切換（on/off）
  6) tvSysMap 檢查：
     6.1 [VOLUME_CURVE_CFG] 區塊內 value 解析成路徑並驗證存在
     6.2 是否存在 <TvSystem type="DVB"> / <TvSystem type="DVB_CO"> / <TvSystem type="DTMB">
     6.3 在上述 TvSystem 節點內，inputSource 值需至少包含 DVBT/DVBC/DVBS 之一（且非 NULL/None）
! 7) **制式檢查新流程**：model.ini → tvSysMapCfgs.xml → CountryTvSysMapXML → countryTvSysMap.xml，
!    以 <COUNTRY_NAME>/<TV_SYSTEM> 為準，對 EU/EFTA/GB/CH 國家強制 `DVB` 或 `DVB_CO`
! 8) 輸出 Excel（格式化）或 CSV 報表；可用退出碼配合 CI
"""
import argparse
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple
import re
from collections import Counter

# ---------- 規則/常數 ----------
# 歐洲國家 allow-list（預設 EU + EFTA + GB + CH）
EU_ALLOWED_DEFAULT: Set[str] = {
    # EU
    "AT","BE","BG","HR","CY","CZ","DK","EE","FI","FR","DE","GR","HU","IE","IT",
    "LV","LT","LU","MT","NL","PL","PT","RO","SK","SI","ES","SE",
    # EFTA
    "IS","LI","NO","CH",
    # 其他常見歐洲市場
    "GB",
}

# 明確非 DVB 的國家（可用 CLI 覆蓋）
NON_DVB_DENYLIST_DEFAULT: Set[str] = {
    "US","CA","MX",       # ATSC
    "JP","KR",            # ISDB / ATSC3
    "CN","HK","MO",       # DTMB
    "BR"                  # ISDB-Tb（巴西）
}

# 抓 /tvconfigs/... 的路徑片段（直到空白/引號/; 結束）
TV_PATH_RE = re.compile(r'(/tvconfigs/[^\s";]+)')

# 寬鬆偵測 CLTV 線索
CLTV_PAT = re.compile(r'(CLTV|livechannels|live channels)', re.IGNORECASE)

# 寬鬆偵測多制式切換鍵（允許鍵名前後含其他字串；鍵名需同時含 MULTI 與 STANDARD/STD）
MULTISTD_LINE_PAT = re.compile(
    r'^\s*([A-Z0-9_.-]*MULTI[A-Z0-9_.-]*(STANDARD|STD)[A-Z0-9_.-]*)\s*=\s*(.+)$',
    re.IGNORECASE
)

# 抽 2 碼國家碼（仍保留，用於行內短碼）
COUNTRY_TOKEN_RE = re.compile(r'\b([A-Z]{2})\b')
# 布林解析
BOOL_TRUE = {"1","true","yes","on","enable","enabled"}
BOOL_FALSE = {"0","false","no","off","disable","disabled"}

# tvSysMap 解析常數
TVSYS_ALLOWED_TYPES = {"DVB", "DVB_CO", "DTMB"}
DVB_INPUT_TOKENS = {"DVBT", "DVBC", "DVBS"}
NULL_TOKENS = {"", "NULL", "NONE", "N/A"}

# ----------- 國名→ISO 兩碼映射 -----------
NAME_TO_ALPHA2 = {
    # 你提供的 tvSysMapCountry.ini
    "USA": "US",
    "MEXICO": "MX",
    "CANADA": "CA",
    "AUSTRALIA": "AU",
    "THAILAND": "TH",
    "INDONESIA": "ID",
    "VIETNAM": "VN",
    "MALAYSIA": "MY",
    "PHILIPPINES": "PH",
    "SOUTH_KOREA": "KR",
    "INDIA": "IN",
    "SPAIN": "ES",
    "FRANCE": "FR",
    "GERMANY": "DE",
    "ITALY": "IT",
    "SWEDEN": "SE",
    "CZECH": "CZ",
    "UNITED_KINGDOM": "GB",
    "HUNGARY": "HU",
    "NORWAY": "NO",
    "FINLAND": "FI",
    "BRAZIL": "BR",
    "COLOMBIA": "CO",
    "URUGUAY": "UY",
    "CHILE": "CL",
    # 其他常見（可逐步擴充）
    "UNITED_STATES": "US",
    "UK": "GB",
}

# ---------- 小工具 ----------
def _sanitize_tv_path(p: str) -> str:
    return p.rstrip('",; \t\r\n)')

def _resolve_to_project(root: Path, tv_path: str) -> Optional[Path]:
    tv_path = _sanitize_tv_path(tv_path)
    if tv_path.startswith("/tvconfigs/"):
        return (root / tv_path[len("/tvconfigs/"):]).resolve()
    # 相對路徑也容忍（較少見）
    return (root / tv_path.lstrip("./")).resolve()

def _read_text(path: Path) -> str:
    try:
        return path.read_text(encoding="utf-8", errors="ignore")
    except Exception:
        return ""

def _read_lines(path: Path) -> List[str]:
    try:
        return path.read_text(encoding="utf-8", errors="ignore").splitlines()
    except Exception:
        return []
"""
def _extract_country_codes_from_text(text: str) -> Set[str]:
    #print(f"_extract_country_codes_from_text: {text}")
    codes = set(COUNTRY_TOKEN_RE.findall(text))
    noise = {"ON","OFF","IN","TV","OS","PQ","AV","EU"}  # 常見高頻非國碼
    return {c for c in codes if c not in noise}
"""

def _extract_country_codes_from_text(text: str) -> Set[str]:
    """
    從一段文字抽取「國家集合」，**輸出一律為 ISO 兩碼**。
    規則：
      1) 先以逗號/換行/分號切段，token 允許底線（UNITED_KINGDOM）
      2) token 若在 NAME_TO_ALPHA2 → 取映射；若是兩碼大寫 → 直接採用
      3) 仍保留舊有掃描（兩碼短詞），避免遺漏
    """
    results: Set[str] = set()
    # 先粗切（適合 tvSysMapCountry.ini 一行一個、逗號結尾的格式）
    raw_tokens = re.split(r'[,;\r\n]+', text)
    for tok in raw_tokens:
        t = tok.strip().upper().replace(" ", "_")
        if not t:
            continue
        if t in NAME_TO_ALPHA2:
            results.add(NAME_TO_ALPHA2[t])
            continue
        if re.fullmatch(r"[A-Z]{2}", t):
            # 過濾常見非國碼短詞
            if t not in {"ON","OF","NO","IN","TV","OS","PQ","AV","EU"}:
                results.add(t)
    # 再補：舊有兩碼掃描（避免文字段落中的短碼遺漏）
    for c in COUNTRY_TOKEN_RE.findall(text):
        if c not in {"ON","OF","NO","IN","TV","OS","PQ","AV","EU"}:
            results.add(c)
    return results

def _bool_state_from_str(s: str) -> Optional[bool]:
    v = s.strip().strip('"\';').lower()
    if v in BOOL_TRUE: return True
    if v in BOOL_FALSE: return False
    return None

def _tokenize_values(s: str) -> List[str]:
    """
    將字串以逗號/分號/管線/空白/斜線分割成 token，移除引號與空白。
    """
    raw = re.split(r'[,\s;/|]+', s.strip().strip('"\'')) if s else []
    return [t for t in (tok.strip().strip('"\'' ) for tok in raw) if t]


# ---------- 解析 sys/*.ini / device_sys.ini ----------
def find_pid_model_paths(root: Path, pid_index: int = 1) -> List[Tuple[Path, Optional[Path], Optional[Path]]]:
    """
    回傳 [(sys_ini, model_ini_abs or None, board_ini_abs or None), ...]
    從 root/sys/*.ini 解析 Model_<pid_index> / Board_<pid_index>
    """
    results: List[Tuple[Path, Optional[Path], Optional[Path]]] = []
    for sys_ini in sorted((root / "sys").glob("*.ini")):
        model_path: Optional[Path] = None
        board_path: Optional[Path] = None
        try:
            with sys_ini.open("r", encoding="utf-8", errors="ignore") as f:
                for line in f:
                    m = re.match(rf'^\s*Model_{pid_index}\s*=\s*"(?P<p>[^"]+)"', line)
                    if m:
                        p = m.group("p")
                        model_path = _resolve_to_project(root, p)
                        continue
                    b = re.match(rf'^\s*Board_{pid_index}\s*=\s*"(?P<b>[^"]+)"', line)
                    if b:
                        bp = b.group("b")
                        board_path = _resolve_to_project(root, bp)
        except Exception:
            pass
        results.append((sys_ini, model_path, board_path))
    return results

def find_device_pid_model_paths(root: Path, pid_index: int = 1) -> List[Tuple[Path, Optional[Path], Optional[Path]]]:
    """
    回傳 [(device_sys.ini, model_ini_abs or None, board_ini_abs or None), ...]
    從 root 的上一層往下找 device_sys.ini，用於覆寫檢視
    """
    results: List[Tuple[Path, Optional[Path], Optional[Path]]] = []
    for dev_sys in (root.parent).rglob("device_sys.ini"):
        m_path: Optional[Path] = None
        b_path: Optional[Path] = None
        try:
            with dev_sys.open("r", encoding="utf-8", errors="ignore") as f:
                for line in f:
                    m = re.match(rf'^\s*Model_{pid_index}\s*=\s*"(?P<p>[^"]+)"', line)
                    if m:
                        p = m.group("p")
                        m_path = _resolve_to_project(root, p)
                        continue
                    b = re.match(rf'^\s*Board_{pid_index}\s*=\s*"(?P<b>[^"]+)"', line)
                    if b:
                        bp = b.group("b")
                        b_path = _resolve_to_project(root, bp)
        except Exception:
            pass
        results.append((dev_sys, m_path, b_path))
    return results

"""
# ---------- 從 model.ini 與引用檔蒐集國家碼 ----------
def gather_pid_countries_from_model(model_ini: Path, root: Path) -> Set[str]:
    #從 model.ini 本身與其行內引用的 /tvconfigs/country/、/tvconfigs/TvSysMap/ 檔案抓國家碼。
    countries: Set[str] = set()
    txt = _read_text(model_ini)
    for line in txt.splitlines():
        low = line.lower()
        if "country_path" not in low and "tvsysmap" not in low:
            continue
        countries |= _extract_country_codes_from_text(line)
        for m in TV_PATH_RE.findall(line):
            if ("/tvconfigs/country/" in m) or ("/tvconfigs/TvSysMap/" in m):
                p = _resolve_to_project(root, m)
                if p and p.exists() and p.is_file():
                    countries |= _extract_country_codes_from_text(_read_text(p))
        print('in gather_pid_countries_from_model ', countries)
    return countries
"""
def gather_pid_countries_from_model(model_ini: Path, root: Path) -> Set[str]:
    """
    從 model.ini 本身與其行內引用的 /tvconfigs/country/、/tvconfigs/TvSysMap/ 檔案抓國家碼。
    """
    countries: Set[str] = set()
    txt = _read_text(model_ini)
    
    print(f"\n▼ 開始解析檔案: {model_ini.name}")  # 增加明顯的解析起點標記
    
    for line_num, line in enumerate(txt.splitlines(), 1):
        low = line.lower()
        if "country_path" not in low and "tvsysmap" not in low:
            continue
            
        print(f"\n▏ 第 {line_num} 行符合條件: {line.strip()}")
        
        # 從本行直接提取
        line_codes = _extract_country_codes_from_text(line)
        print(f"  行內提取國家碼: {line_codes}")
        countries |= line_codes
        
        # 從關聯檔案提取
        for m in TV_PATH_RE.findall(line):
            if ("/tvconfigs/country/" in m) or ("/tvconfigs/TvSysMap/" in m):
                p = _resolve_to_project(root, m)
                if p and p.exists() and p.is_file():
                    print(f"  找到關聯檔案: {p.name}")
                    file_codes = _extract_country_codes_from_text(_read_text(p))
                    print(f"  檔案內提取國家碼: {file_codes}")
                    countries |= file_codes
                    
        print(f"  目前累積國家碼: {sorted(countries)}")  # 顯示即時進度
    
    print(f"▲ 最終國家碼集合: {sorted(countries)}\n")
    return countries


# ---------- CLTV / 多制式偵測 ----------
def detect_cltv_and_multistd(model_ini: Path) -> Tuple[Optional[bool], Optional[bool]]:
    """
    回傳：(cltv_enabled | None, multi_std_enabled | None)
    - CLTV：只要行內含 CLTV 或 livechannels 就視為「有整合」（True）
    - Multi-Std：若找到 MULTI...STD 類鍵，解析布林；否則回 None
    """
    cltv = None
    multi = None
    try:
        with model_ini.open("r", encoding="utf-8", errors="ignore") as f:
            for line in f:
                if cltv is None and CLTV_PAT.search(line):
                    cltv = True
                if multi is None:
                    m = MULTISTD_LINE_PAT.match(line)
                    if m:
                        multi = _bool_state_from_str(m.group(3))
    except Exception:
        pass
    return (cltv, multi)


# ---------- 從 model.ini 取得 TvSysMap 檔，並檢查其中 [VOLUME_CURVE_CFG] ----------
SECTION_RE = re.compile(r'^\s*\[([^\]]+)\]\s*$', re.IGNORECASE)
KV_RE = re.compile(r'^\s*([^;#=\s][^=]{0,120}?)\s*=\s*(.*?)\s*$')
PLACEHOLDER_VALUES = {"", "none", "null", "0", "off", "disable", "disabled"}

def find_tvsysmap_files_from_model(model_ini: Path, root: Path) -> List[Path]:
    files: List[Path] = []
    txt = _read_text(model_ini)
    for line in txt.splitlines():
        if "tvsysmap" not in line.lower():
            continue
        for m in TV_PATH_RE.findall(line):
            if "/tvconfigs/TvSysMap/" in m:
                p = _resolve_to_project(root, m)
                if p and p.exists() and p.is_file():
                    files.append(p)
    return files

def _resolve_candidate_path(root: Path, value: str) -> Optional[Path]:
    """從 VOLUME_CURVE_CFG 的 value 解析可能的檔案路徑。支援 /tvconfigs/ 與相對路徑。"""
    v = value.strip().strip('"\'')

    # 移除 inline 註解（; 之後）
    if ";" in v:
        v = v.split(";", 1)[0].strip()
    if v.lower() in PLACEHOLDER_VALUES:
        return None
    # 先抓 /tvconfigs/ 片段
    tvs = TV_PATH_RE.findall(v)
    if tvs:
        return _resolve_to_project(root, tvs[0])
    # 否則當作相對於 root 的路徑
    return (root / v.lstrip("./")).resolve()

def extract_volcurve_paths_from_tvsysmap(tvsysmap_path: Path, root: Path) -> Tuple[List[Path], List[Path]]:
    """
    讀取一個 TvSysMap 檔，擷取 [VOLUME_CURVE_CFG] 區塊所有 value 的檔案路徑並檢查存在性。
    回傳：(全部解析出的路徑, 缺檔路徑)
    """
    lines = _read_lines(tvsysmap_path)
    in_section = False
    refs: List[Path] = []
    for raw in lines:
        msec = SECTION_RE.match(raw)
        if msec:
            in_section = (msec.group(1).strip().upper() == "VOLUME_CURVE_CFG")
            continue
        if not in_section:
            continue
        mkv = KV_RE.match(raw)
        if not mkv:
            continue
        value = mkv.group(2)
        p = _resolve_candidate_path(root, value)
        if p:
            refs.append(p)
    missing = [p for p in refs if not p.exists()]
    return refs, missing


# ---------- 解析/檢查 TvSystem 與 inputSource ----------
def _try_parse_xml(text: str):
    """
    嘗試用 xml.etree.ElementTree 解析；失敗則回 None。
    """
    try:
        import xml.etree.ElementTree as ET
        return ET.fromstring(text)
    except Exception:
        return None

def _scan_tvsystem_blocks_regex(text: str) -> List[Tuple[str, str]]:
    """
    以 regex 掃描 <TvSystem ...> ... </TvSystem> 區塊。
    回傳清單：[(attributes_string, inner_text), ...]
    """
    blocks = []
    for m in re.finditer(r'<\s*TvSystem\b([^>]*)>(.*?)</\s*TvSystem\s*>', text, flags=re.IGNORECASE | re.DOTALL):
        attrs = m.group(1) or ""
        inner = m.group(2) or ""
        blocks.append((attrs, inner))
    return blocks

def _extract_attr(attrs: str, name: str) -> Optional[str]:
    m = re.search(rf'\b{name}\s*=\s*"(.*?)"', attrs, flags=re.IGNORECASE)
    if m: return m.group(1)
    m = re.search(rf"\b{name}\s*=\s*'(.*?)'", attrs, flags=re.IGNORECASE)
    if m: return m.group(1)
    return None

def _collect_input_tokens_from_text(text: str) -> Set[str]:
    """
    從 TvSystem 區塊內部文字，找出 inputSource 類的值：
    - 屬性形式：inputSource="DVBT,DVBC"
    - 子節點形式：<inputSource value="DVBT" />、<inputSource>DVBT</inputSource>
    - 其他可能：<input value="DVBT"> 或 key=value 行
    """
    tokens: Set[str] = set()

    # inputSource="..."
    for m in re.finditer(r'\binputSource\s*=\s*"(.*?)"', text, flags=re.IGNORECASE | re.DOTALL):
        for t in _tokenize_values(m.group(1)):
            tokens.add(t.upper())

    # inputSource='...'
    for m in re.finditer(r"\binputSource\s*=\s*'(.*?)'", text, flags=re.IGNORECASE | re.DOTALL):
        for t in _tokenize_values(m.group(1)):
            tokens.add(t.upper())

    # <inputSource value="...">
    for m in re.finditer(r'<\s*inputSource\b[^>]*\bvalue\s*=\s*"(.*?)"[^>]*>', text, flags=re.IGNORECASE | re.DOTALL):
        for t in _tokenize_values(m.group(1)):
            tokens.add(t.upper())
    for m in re.finditer(r"<\s*inputSource\b[^>]*\bvalue\s*=\s*'(.*?)'[^>]*>", text, flags=re.IGNORECASE | re.DOTALL):
        for t in _tokenize_values(m.group(1)):
            tokens.add(t.upper())

    # <inputSource>DVBT</inputSource>
    for m in re.finditer(r'<\s*inputSource\s*>(.*?)</\s*inputSource\s*>', text, flags=re.IGNORECASE | re.DOTALL):
        for t in _tokenize_values(m.group(1)):
            tokens.add(t.upper())

    # 泛用：<input value="DVBT"> / key=value
    for m in re.finditer(r'<\s*input\b[^>]*\bvalue\s*=\s*"(.*?)"[^>]*>', text, flags=re.IGNORECASE | re.DOTALL):
        for t in _tokenize_values(m.group(1)):
            tokens.add(t.upper())
    for m in re.finditer(r"\binput\s*=\s*'(.*?)'", text, flags=re.IGNORECASE | re.DOTALL):
        for t in _tokenize_values(m.group(1)):
            tokens.add(t.upper())

    return tokens

 # ---------- 解析 tvSysMapCfgs.xml → CountryTvSysMapXML → countryTvSysMap.xml ----------
def _try_parse_xml_any(text: str):
    """盡量把文字當 XML parse；不成就加一層 root 再試。"""
    try:
        import xml.etree.ElementTree as ET
        return ET.fromstring(text)
    except Exception:
        pass
    try:
        import xml.etree.ElementTree as ET
        return ET.fromstring("<_root>"+text+"</_root>")
    except Exception:
        return None

def _iter_elems_by_tag(root, tag_name: str):
    tgt = tag_name.lower()
    for el in root.iter():
        if el.tag.split('}')[-1].lower() == tgt:
            yield el

def resolve_country_tvsysmap_paths_from_tvsysmap(tvsysmap_path: Path, root: Path) -> List[Path]:
    """
    讀 tvSysMapCfgs.xml，找 <CountryTvSysMapXML> 的路徑（可一或多個），轉為專案實體路徑。
    """
    paths: List[Path] = []
    txt = _read_text(tvsysmap_path)
    rt = _try_parse_xml_any(txt)
    if rt is not None:
        for el in _iter_elems_by_tag(rt, "CountryTvSysMapXML"):
            v = (el.text or "").strip()
            if not v:
                continue
            p = _resolve_to_project(root, v)
            if p: paths.append(p)
    else:
        # fallback：regex
        for m in re.finditer(r'<\s*CountryTvSysMapXML\s*>(.*?)</\s*CountryTvSysMapXML\s*>',
                             txt, flags=re.IGNORECASE | re.DOTALL):
            v = (m.group(1) or "").strip()
            if not v: continue
            p = _resolve_to_project(root, v)
            if p: paths.append(p)
    return paths

def parse_country_tv_system_file(ctvs_path: Path) -> List[Tuple[str, str, Optional[str]]]:
     """
     讀 countryTvSysMap.xml，回傳清單 [(COUNTRY_NAME, TV_SYSTEM, TV_CONFIG), ...]
     """
     res: List[Tuple[str, str, Optional[str]]] = []
     txt = _read_text(ctvs_path)
     rt = _try_parse_xml_any(txt)
     if rt is not None:
         for block in _iter_elems_by_tag(rt, "COUNTRY_TVCONFIG_MAP"):
             name = ""
             sys = ""
             cfg = None
             for child in block:
                 tag = child.tag.split('}')[-1].upper()
                 val = (child.text or "").strip()
                 if tag == "COUNTRY_NAME":
                     name = val
                 elif tag == "TV_SYSTEM":
                     sys = val
                 elif tag == "TV_CONFIG":
                     cfg = val
             if name:
                 res.append((name, sys, cfg))
         return res
     # fallback：regex 區塊掃描
     for m in re.finditer(r'<\s*COUNTRY_TVCONFIG_MAP\s*>(.*?)</\s*COUNTRY_TVCONFIG_MAP\s*>',
                          txt, flags=re.IGNORECASE | re.DOTALL):
         blk = m.group(1) or ""
         def _tag(tag):
             mm = re.search(rf'<\s*{tag}\s*>(.*?)</\s*{tag}\s*>', blk, flags=re.IGNORECASE | re.DOTALL)
             return (mm.group(1) or "").strip() if mm else ""
         name = _tag("COUNTRY_NAME")
         sys = _tag("TV_SYSTEM")
         cfg = _tag("TV_CONFIG") or None
         if name:
             res.append((name, sys, cfg))
     return res

def _name_to_alpha2(name: str) -> Optional[str]:
    t = name.strip().upper().replace(" ", "_")
    return NAME_TO_ALPHA2.get(t, None)

def _parse_input_mapping_value(v: str) -> dict:
    """
    解析 inputSource 的 value 欄位，例如：
      "TV:NULL,ATV:ATV,DVBT:ANTENNA,DVBC:CABLE,DVBS:SATELLITE,HDMI1:HDMI1,..."
    回傳 dict（key/val 皆大寫、已去空白；忽略沒冒號/空白項）
    """
    if not v:
        return {}
    # 去除換行與多餘空白，避免分割出空 token
    vv = " ".join(v.replace("\n", " ").replace("\r", " ").split())
    mapping = {}
    for token in vv.split(","):
        token = token.strip().strip('"\'')

        if not token or ":" not in token:
            continue
        k, val = token.split(":", 1)
        k = k.strip().upper()
        val = val.strip().upper()
        if not k:
            continue
        mapping[k] = val
    return mapping

def analyze_tvsysmap_for_types_and_inputs(text: str) -> Tuple[Set[str], Set[str]]:
    """
    傳回 (types_found, dvb_inputs_found_non_null)

    - types_found：出現的 TvSystem type（只收 DVB / DVB_CO / DTMB）
    - dvb_inputs_found_non_null：在該等 TvSystem 區塊內，inputSource 的 DVBT/DVBC/DVBS
      其 value 至少一個「非 NULL/NONE/空字串」者，將其 key 納入（例：{"DVBT","DVBC"}）
    """
    types_found: Set[str] = set()
    dvb_non_null: Set[str] = set()

    # ---------- 先嘗試 XML 解析 ----------
    root = _try_parse_xml_any(text)
    if root is not None:
        for el in root.iter():
            tag = el.tag.split("}")[-1].lower()
            if tag != "tvsystem":
                continue

            t = el.attrib.get("type", "").strip().upper()
            if t in TVSYS_ALLOWED_TYPES:
                types_found.add(t)

                for sub in el.iter():
                    stag = sub.tag.split("}")[-1].lower()
                    if stag != "item":
                        continue
                    cat = sub.attrib.get("category", "").strip()
                    key = sub.attrib.get("key", "").strip().upper()
                    if cat == "[LiveTV]" and key == "INPUTSOURCE":
                        mapping = _parse_input_mapping_value(sub.attrib.get("value", ""))
                        for k in ("DVBT", "DVBC", "DVBS"):
                            if k in mapping:
                                val = (mapping[k] or "").strip().upper()
                                if val not in NULL_TOKENS:
                                    dvb_non_null.add(k)
        return types_found, dvb_non_null

    # ---------- XML 失敗 → regex fallback ----------
    for m in re.finditer(r'<\s*TvSystem\b([^>]*)>(.*?)</\s*TvSystem\s*>',
                         text, flags=re.IGNORECASE | re.DOTALL):
        attrs = m.group(1) or ""
        inner = m.group(2) or ""

        mt = re.search(r'\btype\s*=\s*"(.*?)"', attrs, flags=re.IGNORECASE) or \
             re.search(r"\btype\s*=\s*'(.*?)'", attrs, flags=re.IGNORECASE)
        t = (mt.group(1) if mt else "").strip().upper()
        if t not in TVSYS_ALLOWED_TYPES:
            continue
        types_found.add(t)

        for im in re.finditer(r'<\s*Item\b([^>]*)>', inner, flags=re.IGNORECASE | re.DOTALL):
            iattrs = im.group(1) or ""
            mc = re.search(r'\bcategory\s*=\s*"(.*?)"', iattrs, flags=re.IGNORECASE) or \
                 re.search(r"\bcategory\s*=\s*'(.*?)'", iattrs, flags=re.IGNORECASE)
            mk = re.search(r'\bkey\s*=\s*"(.*?)"', iattrs, flags=re.IGNORECASE) or \
                 re.search(r"\bkey\s*=\s*'(.*?)'", iattrs, flags=re.IGNORECASE)
            mv = re.search(r'\bvalue\s*=\s*"(.*?)"', iattrs, flags=re.IGNORECASE | re.DOTALL) or \
                 re.search(r"\bvalue\s*=\s*'(.*?)'", iattrs, flags=re.IGNORECASE | re.DOTALL)

            cat = (mc.group(1) if mc else "").strip()
            key = (mk.group(1) if mk else "").strip().upper()
            val = (mv.group(1) if mv else "")

            if cat == "[LiveTV]" and key == "INPUTSOURCE":
                mapping = _parse_input_mapping_value(val)
                for k in ("DVBT", "DVBC", "DVBS"):
                    if k in mapping:
                        v2 = (mapping[k] or "").strip().upper()
                        if v2 not in NULL_TOKENS:
                            dvb_non_null.add(k)

    return types_found, dvb_non_null

# ---------- PID1 檢核核心 ----------
def check_pid1_rules(
    root: Path,
    *,
    eu_allowed_countries: Optional[Set[str]] = None,
    nondvb_denylist: Optional[Set[str]] = None,
    include_device_sys: bool = False,
) -> List[Dict[str, object]]:
    """
    規則：
      1) Model_1/Board_1 檔案存在（Board_1 可選；若定義必須存在）
      2) 國家集合需屬於歐洲 allow-list；且不得命中非 DVB deny-list
      3) CLTV 與多制式切換僅做資訊列示，不影響通過與否（除非多制式值無法解析為布林）
      4) model.ini → tvSysMap → [VOLUME_CURVE_CFG] 路徑存在
      5) tvSysMap XML：至少存在一個 <TvSystem type="DVB|DVB_CO|DTMB">，
         且其 inputSource 值集合需包含 DVBT/DVBC/DVBS 中任一（且非 NULL）
    """
    allow = set(eu_allowed_countries or EU_ALLOWED_DEFAULT)
    deny = set(nondvb_denylist or NON_DVB_DENYLIST_DEFAULT)

    pairs_sys = find_pid_model_paths(root, pid_index=1)
    pairs_device = find_device_pid_model_paths(root, pid_index=1) if include_device_sys else []

    rows: List[Dict[str, object]] = []
    combined = [("sys", *t) for t in pairs_sys] + [("device", *t) for t in pairs_device]
    for origin, sys_ini, model_path, board_path in combined:
        row: Dict[str, object] = {
            "origin": origin,
            "sys_ini": str(sys_ini.relative_to(root)) if sys_ini.exists() else str(sys_ini),
            "model_ini": str(model_path.relative_to(root)) if (model_path and model_path.exists()) else (str(model_path) if model_path else "(not set)"),
            "board_ini": str(board_path.relative_to(root)) if (board_path and board_path.exists()) else (str(board_path) if board_path else ""),
            "exists_model": bool(model_path and model_path.exists()),
            "exists_board": (board_path is None) or (board_path.exists()),
            "countries": "",
            "vol_curve_files": "",
            "vol_curve_missing": "",
            "tvsys_files": "",
            "tvsys_types": "",
            "tvsys_inputs": "",
            "ctvs_files": "",
            "ctvs_missing": "",
            "ctvs_systems": "",
            "ctvs_eu_non_dvb": "",
            "violations": "",
            "cltv": "",
            "multi_std": "",
            "status": "OK",
        }
        problems: List[str] = []

        # 1) 存在性
        if not model_path:
            problems.append("Model_1 未設定")
        elif not (model_path.exists() and model_path.is_file()):
            problems.append("Model_1 檔案不存在")
        if board_path and not (board_path.exists() and board_path.is_file()):
            problems.append("Board_1 檔案不存在")

        # 2) 國家/制式
        countries: Set[str] = set()
        if model_path and model_path.exists():
            # 先用新的 CountryTvSysMap 流程
            tvmaps = find_tvsysmap_files_from_model(model_path, root)
            ctvs_paths: List[Path] = []
            ctvs_missing: List[Path] = []
            ctvs_map: Dict[str, str] = {}
            for mp in tvmaps:
                cpaths = resolve_country_tvsysmap_paths_from_tvsysmap(mp, root)
                for cp in cpaths:
                    if cp.exists():
                        ctvs_paths.append(cp)
                        for name, tvsys, _cfg in parse_country_tv_system_file(cp):
                            a2 = _name_to_alpha2(name) or ""
                            if a2:
                                countries.add(a2)
                                if tvsys:
                                    ctvs_map[a2] = tvsys.strip().upper()
                    else:
                        ctvs_missing.append(cp)
            # 若無法由 CountryTvSysMap 取得任何國家，再 fallback 舊法
            if not countries:
                countries = gather_pid_countries_from_model(model_path, root)
        row["countries"] = ",".join(sorted(countries)) if countries else "(not found)"
        # 記錄 CountryTvSysMap 解析結果
        if model_path and model_path.exists():
            row["ctvs_files"]   = ", ".join(sorted({str(p.relative_to(root)) for p in ctvs_paths})) if locals().get("ctvs_paths") else ""
            row["ctvs_missing"] = ", ".join(sorted({str(p) for p in ctvs_missing})) if locals().get("ctvs_missing") else ""
            if locals().get("ctvs_missing"):
                problems.append("CountryTvSysMapXML 缺檔: " + ", ".join(sorted(str(p) for p in ctvs_missing)))
            if locals().get("ctvs_map"):
                row["ctvs_systems"] = ", ".join(sorted(f"{k}:{v}" for k, v in ctvs_map.items()))
                # 僅對 EU/EFTA/GB/CH 的國家要求 TV_SYSTEM 必須是 DVB 或 DVB_CO
                bad = []
                for a2 in sorted(countries & allow):
                    tvs = ctvs_map.get(a2, "")
                    if tvs and tvs not in {"DVB", "DVB_CO"}:
                        bad.append(f"{a2}:{tvs}")
                if bad:
                    row["ctvs_eu_non_dvb"] = ", ".join(bad)
                    problems.append("CountryTvSysMap EU 非 DVB 制式: " + ", ".join(bad))
            else:
                # 沒有 ctvs_map → 用舊的 allow/deny 規則維持舊行為
                if countries:
                    extra = countries - allow
                    if extra:
                        problems.append(f"包含非歐洲國家: {','.join(sorted(extra))}")
                    hit = countries & deny
                    if hit:
                        problems.append(f"包含非 DVB 國家: {','.join(sorted(hit))}")

        # 3) CLTV / Multi-Std 資訊
        cltv, multi = (None, None)
        if model_path and model_path.exists():
            cltv, multi = detect_cltv_and_multistd(model_path)
        row["cltv"] = "enabled" if cltv else ("disabled/absent" if cltv is not None else "unknown")
        if multi is None:
            row["multi_std"] = "not_set"
        else:
            row["multi_std"] = "on" if multi else "off"
            if multi not in (True, False):
                problems.append("多制式切換值非布林")

        # 4) tvSysMap → [VOLUME_CURVE_CFG] 檢查 & 5) TvSystem / inputSource 檢查
        tvmaps: List[Path] = []
        types_found_all: Set[str] = set()
        inputs_found_all: Set[str] = set()
        vc_refs_all: List[Path] = []
        vc_missing_all: List[Path] = []

        if model_path and model_path.exists():
            tvmaps = find_tvsysmap_files_from_model(model_path, root)

            # vol curve
            for mp in tvmaps:
                refs, miss = extract_volcurve_paths_from_tvsysmap(mp, root)
                vc_refs_all.extend(refs)
                vc_missing_all.extend(miss)

            # TvSystem + inputSource
            for mp in tvmaps:
                txt = _read_text(mp)
                types_found, inputs_found = analyze_tvsysmap_for_types_and_inputs(txt)
                types_found_all |= types_found
                #print("type =",types_found_all)
                inputs_found_all |= inputs_found
                #print("inputs =",inputs_found_all)

        # 報表欄位整理
        def _rel(p: Path) -> str:
            try:
                return str(p.relative_to(root))
            except Exception:
                return str(p)

        row["tvsys_files"] = ", ".join(sorted({_rel(p) for p in tvmaps})) if tvmaps else ""
        row["tvsys_types"] = ", ".join(sorted(types_found_all)) if types_found_all else ""
        row["tvsys_inputs"] = ", ".join(sorted(inputs_found_all)) if inputs_found_all else ""

        uniq_refs = sorted({_rel(p) for p in vc_refs_all})
        uniq_missing = sorted({_rel(p) for p in vc_missing_all})
        row["vol_curve_files"] = ", ".join(uniq_refs) if uniq_refs else ""
        row["vol_curve_missing"] = ", ".join(uniq_missing) if uniq_missing else ""
        if uniq_missing:
            problems.append("VOLUME_CURVE_CFG 缺檔: " + ", ".join(uniq_missing))

        # 規則：需有 allowed 的 TvSystem type
        if tvmaps:
            if not types_found_all:
                problems.append('TvSysMap 未包含 <TvSystem type="DVB|DVB_CO|DTMB">')
            # 若有 type，需至少有 DVBT/DVBC/DVBS 之一（且非 NULL）
            if types_found_all and not (inputs_found_all & DVB_INPUT_TOKENS):
                problems.append("TvSysMap 的 inputSource 未含 DVBT/DVBC/DVBS")

        # 結果
        if problems:
            row["status"] = "ERROR"
            row["violations"] = " ; ".join(problems)

        rows.append(row)

    return rows


# ---------- 報表輸出 ----------
def beautify_excel(writer, sheet_columns: Dict[str, List[str]]) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.formatting.rule import CellIsRule

    table_style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True, showColumnStripes=False)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    wrap_alignment = Alignment(wrap_text=True, vertical="top")

    for sheet, cols in sheet_columns.items():
        ws = writer.sheets[sheet]
        max_row = ws.max_row
        max_col = ws.max_column
        if max_row < 2 or max_col < 1:
            continue

        ref = f"A1:{chr(64+max_col)}{max_row}"
        try:
            table = Table(displayName=f"{sheet.replace(' ','_')}_Table", ref=ref)
            table.tableStyleInfo = table_style
            ws.add_table(table)
        except Exception:
            ws.auto_filter.ref = ref

        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        preferred = {
            "origin": 10, "sys_ini": 28, "model_ini": 36, "board_ini": 36,
            "exists_model": 12, "exists_board": 12, "countries": 36,
            "vol_curve_files": 40, "vol_curve_missing": 40,
            "tvsys_files": 36, "tvsys_types": 18, "tvsys_inputs": 18,
            "ctvs_files": 36, "ctvs_missing": 36, "ctvs_systems": 60, "ctvs_eu_non_dvb": 44,
            "violations": 60, "cltv": 14, "multi_std": 12, "status": 10,
            "item": 22, "value": 16,
        }
        wrap_cols = {"countries","violations","model_ini","board_ini",
                     "vol_curve_files","vol_curve_missing",
                      "tvsys_files","tvsys_types","tvsys_inputs",
                      "ctvs_files","ctvs_missing","ctvs_systems","ctvs_eu_non_dvb"}
        for j in range(1, max_col+1):
            col_letter = ws.cell(row=1, column=j).column_letter
            header = str(ws.cell(row=1, column=j).value or "")
            width = preferred.get(header, 18)
            ws.column_dimensions[col_letter].width = width
            if header in wrap_cols:
                for i in range(2, max_row+1):
                    ws.cell(row=i, column=j).alignment = wrap_alignment

        # status 色塊
        status_col = None
        for j in range(1, max_col+1):
            if str(ws.cell(row=1, column=j).value).lower() == "status":
                status_col = ws.cell(row=1, column=j).column_letter
                break
        if status_col:
            rng = f"{status_col}2:{status_col}{max_row}"
            green = PatternFill("solid", fgColor="C6EFCE")
            red = PatternFill("solid", fgColor="FFC7CE")
            ws.conditional_formatting.add(rng, CellIsRule(operator='equal', formula=['"OK"'], fill=green))
            ws.conditional_formatting.add(rng, CellIsRule(operator='equal', formula=['"ERROR"'], fill=red))


def write_pid1_reports(
    root: Path,
    rows: List[Dict[str, object]],
    *,
    xlsx_path: Optional[Path],
    csv_dir: Optional[Path],
) -> None:
    summary = [
        {"item": "scan_root", "value": str(root)},
        {"item": "pid1_models_checked", "value": len(rows)},
        {"item": "pid1_errors", "value": sum(1 for r in rows if r.get("status") == "ERROR")},
    ]
    # 統計錯誤原因（可快速歸因）
    counter = Counter()
    for r in rows:
        if r.get("status") == "ERROR":
            for part in str(r.get("violations","")).split(" ; "):
                if part.strip():
                    counter[part.strip()] += 1
    for k, v in counter.most_common():
        summary.append({"item": f"error::{k}", "value": v})

    # Excel
    if xlsx_path:
        try:
            import pandas as pd  # type: ignore
            with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
                cols_summary = ["item","value"]
                cols_pid1 = [
                    "origin","sys_ini","model_ini","board_ini",
                    "exists_model","exists_board","countries",
                    "vol_curve_files","vol_curve_missing",
                    "tvsys_files","tvsys_types","tvsys_inputs",
                    "ctvs_files","ctvs_missing","ctvs_systems","ctvs_eu_non_dvb",
                    "violations","cltv","multi_std","status"
                ]
                pd.DataFrame(summary, columns=cols_summary).to_excel(writer, index=False, sheet_name="Summary")
                pd.DataFrame(rows, columns=cols_pid1).to_excel(writer, index=False, sheet_name="PID1")
                beautify_excel(writer, {"Summary": cols_summary, "PID1": cols_pid1})
            print(f"✔ 已輸出 Excel 報表：{xlsx_path}")
            return
        except Exception as e:
            print(f"[WARN] 產生 Excel 失敗：{e}\n→ 將改輸出 CSV。")

    # CSV 後備
    out_dir = csv_dir or root
    out_dir.mkdir(parents=True, exist_ok=True)

    def write_csv(name: str, headers: List[str], data: List[Dict[str, object]]):
        p = out_dir / name
        import csv
        with p.open("w", encoding="utf-8", newline="") as f:
            w = csv.DictWriter(f, fieldnames=headers)
            w.writeheader()
            for row in data:
                w.writerow({h: row.get(h, "") for h in headers})
        print(f"✔ 已輸出 CSV：{p}")

    write_csv("pid1_summary.csv", ["item","value"], summary)
    write_csv(
        "pid1.csv",
        [
            "origin","sys_ini","model_ini","board_ini",
            "exists_model","exists_board","countries",
            "vol_curve_files","vol_curve_missing",
            "tvsys_files","tvsys_types","tvsys_inputs",
            "ctvs_files","ctvs_missing","ctvs_systems","ctvs_eu_non_dvb",
            "violations","cltv","multi_std","status"
        ],
        rows
    )


# ---------- CLI ----------
def main():
    parser = argparse.ArgumentParser(
        description="PID1 檢核：Model_1/Board_1 存在性 + 歐洲&DVB 規則 + CLTV/多制式資訊 + tvSysMap[VOLUME_CURVE_CFG] 檔案存在 + TvSystem/inputSource 規則"
    )
    parser.add_argument("--root", required=True, type=Path, help="tvconfigs 專案根（如 ~/tvconfigs_home/tv109/kipling/configs）")
    parser.add_argument("--report-xlsx", type=Path, default=None, help="輸出 Excel 的路徑（.xlsx）")
    parser.add_argument("--csv-dir", type=Path, default=None, help="若 Excel 失敗，CSV 的輸出資料夾（預設 root）")
    parser.add_argument("--include-device-sys", action="store_true", help="同時掃描 device/*/*/sys/*/device_sys.ini 覆寫 PID1")
    parser.add_argument("--eu-allowed-countries", type=str, default="",
                        help="歐洲 allow-list（逗號分隔），留空使用預設 EU+EFTA+GB+CH")
    parser.add_argument("--pid1-deny-nondvb", type=str, default=",".join(sorted(NON_DVB_DENYLIST_DEFAULT)),
                        help="非 DVB 國家 deny-list（逗號分隔）")
    parser.add_argument("--fail-warning", action="store_true", help="如有 PID1 錯誤時以非 0 退出（CI 友善）")

    args = parser.parse_args()
    root = args.root.resolve()
    if not root.exists():
        print(f"[ERROR] root 不存在：{root}")
        raise SystemExit(2)

    allow = {c.strip().upper() for c in args.eu_allowed_countries.split(",") if c.strip()} or None
    deny = {c.strip().upper() for c in args.pid1_deny_nondvb.split(",") if c.strip()}

    rows = check_pid1_rules(
        root,
        eu_allowed_countries=allow,
        nondvb_denylist=deny,
        include_device_sys=args.include_device_sys,
    )

    # 終端摘要
    total = len(rows)
    errors = sum(1 for r in rows if r.get("status") == "ERROR")
    print(f"PID1 檢核：共 {total} 組，錯誤 {errors}")
    for r in rows:
        if r.get("status") == "ERROR":
            print(f"- [{r['origin']}] {r['sys_ini']} -> {r['model_ini']} | {r['violations']}")

    # 報表
    write_pid1_reports(root, rows, xlsx_path=args.report_xlsx, csv_dir=args.csv_dir)

    # 退出碼
    if errors and args.fail_warning:
        raise SystemExit(1)
    raise SystemExit(0)


if __name__ == "__main__":
    main()

