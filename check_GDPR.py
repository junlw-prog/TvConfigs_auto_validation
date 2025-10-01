#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
check_is_support_gdpr.py

參考 tv_multi_standard_validation.py 的「/tvconfigs → --root 映射」與報表格式：
- 表頭固定：Rules, Result, condition_1, condition_2, ...
- 依 model.ini 檔名前綴分頁：PID_1、PID_2…；其他 → others
- 若 xlsx 已存在則附加

功能：
從指定的 model.ini 擷取 isSupportGDPR 的值並輸出到報表。
Result 規則：
- 找到 isSupportGDPR（非空）→ PASS
- 未找到或為空 → FAIL

用法：
python3.8 check_is_support_gdpr.py --model-ini model/1_xxx.ini --root . --report
python3.8 check_is_support_gdpr.py --model-ini model/1_xxx.ini --root . --report-xlsx kipling.xlsx
"""

import argparse
import os
import re
from typing import Optional

# -----------------------------
# Utilities for report（比照 tv_multi_standard_validation.py 風格）
# -----------------------------

def _sheet_name_for_model(model_ini_path: str) -> str:
    """
    根據 model.ini 檔名的前綴決定頁簽：
      - 前綴是數字 N → 'PID_N'
      - 其他 → 'others'
    例: '1_EU_XXX.ini' → 'PID_1'
    """
    base = os.path.basename(model_ini_path or "")
    m = re.match(r"^(\d+)_", base)
    if m:
        return f"PID_{int(m.group(1))}"
    return "others"


def _ensure_openpyxl():
    try:
        import openpyxl  # noqa
    except ImportError:
        raise SystemExit(
            "[ERROR] 需要 openpyxl 以支援報表輸出與附加。\n"
            "  安裝： pip install --user openpyxl\n"
        )


def export_report(res: dict, xlsx_path: str = "kipling.xlsx", num_condition_cols: int = 5) -> None:
    """
    欄位無值時以 'N/A' 填入。依 model.ini 檔名前綴分頁（PID_1、PID_2…；非數字→others），既有資料則附加。
    表頭固定為: Rules, Result, condition_1, condition_2, condition_3, ...
    統一：所有欄位同寬、自動換行、垂直置頂（包含表頭）。
    """
    _ensure_openpyxl()
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    COMMON_WIDTH = 80
    COMMON_ALIGN = Alignment(wrap_text=True, vertical="top")
    BOLD = Font(bold=True)

    def _na(s: str) -> str:
        s = (s or "").strip()
        return s if s else "N/A"

    sheet_name = _sheet_name_for_model(res.get("model_ini", ""))

    # 開啟或新建 xlsx
    try:
        wb = load_workbook(xlsx_path)
    except Exception:
        wb = Workbook()

    # 建立或取得 sheet
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)
        headers = ["Rules", "Result"] + [f"condition_{i}" for i in range(1, num_condition_cols + 1)]
        ws.append(headers)

    # 準備資料
    rules   = "Check isSupportGDPR exists"
    result  = "PASS" if res.get("passed", False) else "FAIL"
    value   = _na(res.get("is_support_gdpr", ""))
    model   = _na(res.get("model_ini", ""))

    conds = [
        f"isSupportGDPR = {value}",  # condition_1
        f"model.ini = {model}",      # condition_2
        "N/A",                       # condition_3
        "N/A",                       # condition_4
        "N/A",                       # condition_5
    ][:num_condition_cols]

    # 寫入 row
    row_values = [rules, result] + conds
    ws.append(row_values)
    last_row = ws.max_row

    # 套用樣式：欄寬、換行、垂直靠上
    total_cols = 2 + num_condition_cols
    for col_idx in range(1, total_cols + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = COMMON_WIDTH

    for cell in ws[1]:  # header
        cell.font = BOLD
        cell.alignment = COMMON_ALIGN

    for cell in ws[last_row]:
        cell.alignment = COMMON_ALIGN

    # 移除預設 Sheet
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        try:
            wb.remove(wb["Sheet"])
        except Exception:
            pass

    wb.save(xlsx_path)


# -----------------------------
# Core parsing
# -----------------------------

def _read_text(path: str) -> str:
    # 寬鬆讀取，容忍常見編碼
    for enc in ("utf-8", "latin-1", "utf-16"):
        try:
            with open(path, "r", encoding=enc) as f:
                return f.read()
        except UnicodeDecodeError:
            continue
        except FileNotFoundError:
            raise
    with open(path, "r") as f:
        return f.read()


def _strip_comment(line: str) -> str:
    # 去掉 # 或 ; 之後的註解
    line = line.split("#", 1)[0]
    line = line.split(";", 1)[0]
    return line.strip()


def _resolve_tvconfigs_path(root: str, tvconfigs_like: str) -> str:
    """
    把 "/tvconfigs/xxx/yyy.ini" 映射為 "<root>/xxx/yyy.ini"
    其他相對路徑: 以 root 為基底
    絕對路徑（非 /tvconfigs 開頭）維持不動
    """
    if tvconfigs_like.startswith("/tvconfigs/"):
        rel = tvconfigs_like[len("/tvconfigs/"):]
        return os.path.normpath(os.path.join(root, rel))
    if tvconfigs_like.startswith("./") or tvconfigs_like.startswith("../"):
        return os.path.normpath(os.path.join(root, tvconfigs_like))
    if tvconfigs_like.startswith("/"):
        # 非 /tvconfigs 絕對路徑：原樣返回
        return tvconfigs_like
    # 其他當作相對於 root
    return os.path.normpath(os.path.join(root, tvconfigs_like))


def parse_is_support_gdpr(model_ini_path: str) -> Optional[str]:
    """
    從 model.ini 內解析 isSupportGDPR（忽略大小寫、允許前後引號與空白）。
    例如：
      isSupportGDPR = true
      IsSupportGdpr= "FALSE"
    """
    txt = _read_text(model_ini_path)

    # 單行關鍵字解析（只取第一個命中的值）
    pat = re.compile(r'^\s*isSupportGDPR\s*=\s*"?([^"\r\n#;]+)"?\s*$', re.IGNORECASE)
    for raw in txt.splitlines():
        line = _strip_comment(raw)
        if not line:
            continue
        m = pat.match(line)
        if m:
            return m.group(1).strip()
    return None


def build_result(model_ini: str, value: Optional[str]) -> dict:
    passed = bool(value and value.strip())
    return {
        "passed": passed,
        "model_ini": model_ini,
        "is_support_gdpr": value or "",
    }


# -----------------------------
# Main
# -----------------------------

def main():
    parser = argparse.ArgumentParser(description="Parse isSupportGDPR from model.ini and export to Excel report (kipling.xlsx).")
    parser.add_argument("--model-ini", required=True, help="path to model ini (e.g., model/1_xxx.ini)")
    parser.add_argument("--root", required=True, help="tvconfigs project root (maps /tvconfigs/* to here)")
    parser.add_argument("-v", "--verbose", action="store_true", help="verbose logs")

    # 新增：--report 直接輸出到 kipling.xlsx；--report-xlsx 可指定路徑
    parser.add_argument("--report", action="store_true", help="export report to kipling.xlsx")
    parser.add_argument("--report-xlsx", metavar="FILE", help="export report to specific xlsx file")

    args = parser.parse_args()

    model_ini = args.model_ini
    if not os.path.exists(model_ini):
        raise SystemExit(f"[ERROR] model ini not found: {model_ini}")

    root = os.path.abspath(os.path.normpath(args.root))
    if args.verbose:
        print(f"[INFO] model_ini: {model_ini}")
        print(f"[INFO] root     : {root}")

    # 解析 isSupportGDPR
    value = parse_is_support_gdpr(model_ini)
    if args.verbose:
        print(f"[INFO] isSupportGDPR: {value if value else '(N/A)'}")

    # 輸出到 console
    res = build_result(model_ini, value)
    print(f"Result : {'PASS' if res['passed'] else 'FAIL'}")
    print(f"isSupportGDPR: {res['is_support_gdpr'] or '(N/A)'}")

    # 報表輸出
    if args.report or args.report_xlsx:
        xlsx_path = args.report_xlsx if args.report_xlsx else "kipling.xlsx"
        export_report(res, xlsx_path=xlsx_path)
        sheet = _sheet_name_for_model(model_ini)
        print(f"[INFO] Report appended to: {xlsx_path} (sheet: {sheet})")


if __name__ == "__main__":
    main()
