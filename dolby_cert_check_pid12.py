#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
tvserv_flags_check.py
---------------------
Reference/format aligned with ai_aipq_check.py and tv_multi_standard_validation.py:
- Headers: Rules, Result, condition_1, condition_2, ...
- Uniform column widths, wrap text, vertical top
- Sheet chosen by model.ini basename prefix (PID_X / others)
- Append to existing kipling.xlsx if present (or use --report-xlsx to specify)

Checks in TvServIni (string compare after stripping, case-insensitive for booleans):
  - DEFAULT_PICTURE_MODE           == 9
  - DEFAULT_DOLBY_PICTURE_MODE     == 1
  - ENABLE_AQ                      == false
  - SUPPORT_MAT                    == true
  - SUPPORT_DOLBY_CERT             == true

All must match to PASS, otherwise FAIL.
"""

import argparse
import os
import re
from typing import Dict, List, Optional, Tuple

# -----------------------------
# Report helpers
# -----------------------------

def _sheet_name_for_model(model_ini_path: str) -> str:
    base = os.path.basename(model_ini_path or "")
    m = re.match(r"^(\d+)_", base)
    if m:
        return f"PID_{int(m.group(1))}"
    return "others"

def _ensure_openpyxl():
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        raise SystemExit(
            "[ERROR] 需要 openpyxl 以支援報表輸出與附加。\n"
            "  安裝： pip install --user openpyxl\n"
        )

def export_report(res: dict, xlsx_path: str = "kipling.xlsx", num_condition_cols: int = 7) -> None:
    """
    表頭固定: Rules, Result, condition_1, condition_2, ...
    所有欄同寬、換行、垂直置頂。依 model.ini 前綴分頁。
    """
    _ensure_openpyxl()
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    COMMON_WIDTH = 80
    COMMON_ALIGN = Alignment(wrap_text=True, vertical="top")
    BOLD = Font(bold=True)

    def _na(s: Optional[str]) -> str:
        s = (s or "").strip()
        return s if s else ""

    sheet_name = _sheet_name_for_model(res.get("model_ini", ""))

    # 開啟或新建 xlsx
    try:
        wb = load_workbook(xlsx_path)
    except Exception:
        wb = Workbook()

    # 建立或取得 sheet
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)
        headers = ["Rules", "Result"] + [f"condition_{i}" for i in range(1, num_condition_cols + 1)]
        ws.append(headers)

    # 準備資料
    rules = (
        "2.Dolby cert =true, 針對 AV sync , scaler 有做 hack flow 去處理 AV sync\n" \
        "3.分 DVB/ATSC/ISDB 是因為送測有可能要特別包特定制式的 image\n" \
        "    - TvServIni → DEFAULT_PICTURE_MODE=9\n" \
        "    - TvServIni → DEFAULT_DOLBY_PICTURE_MODE=1\n" \
        "    - TvServIni → ENABLE_AQ=false\n" \
        "    - TvServIni → SUPPORT_MAT=true\n" \
        "    - TvServIni → SUPPORT_DOLBY_CERT=true\n" 
    )
    result = "PASS" if res.get("passed") else "FAIL"

    vals = res.get("vals", {})
    conds = [
        f"TvServIni = {_na(res.get('tvserv_ini_path'))}",
        f"DEFAULT_PICTURE_MODE = {_na(vals.get('DEFAULT_PICTURE_MODE'))}",
        f"DEFAULT_DOLBY_PICTURE_MODE = {_na(vals.get('DEFAULT_DOLBY_PICTURE_MODE'))}",
        f"ENABLE_AQ = {_na(vals.get('ENABLE_AQ'))}",
        f"SUPPORT_MAT = {_na(vals.get('SUPPORT_MAT'))}",
        f"SUPPORT_DOLBY_CERT = {_na(vals.get('SUPPORT_DOLBY_CERT'))}",
        _na("; ".join(res.get("notes", []))),
    ]
    # 對齊欄位數
    #conds = (conds + ["N/A"] * num_condition_cols)[:num_condition_cols]

    ws.append([rules, result] + conds)
    last_row = ws.max_row

    # 給儲存格指派上色
    rules_color = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
    failed_color = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")
    # 上色
    first_cell = ws.cell(row=last_row, column=1)  # 欄位1對應的是 'A' 列
    first_cell.fill = rules_color
    if result == "FAIL":
        ws.cell(row=last_row, column=2).fill = failed_color
    if conds[0] == "TvServIni = N/A":
        ws.cell(row=last_row, column=3).fill = failed_color
    if conds[1] != "DEFAULT_PICTURE_MODE = 9" :
        ws.cell(row=last_row, column=4).fill = failed_color
    if conds[2] != "DEFAULT_DOLBY_PICTURE_MODE = 1":
        ws.cell(row=last_row, column=5).fill = failed_color
    if conds[5] != "SUPPORT_DOLBY_CERT = true":
        ws.cell(row=last_row, column=8).fill = failed_color
    if conds[4] != "SUPPORT_MAT = true":
        ws.cell(row=last_row, column=7).fill = failed_color
    if conds[3] != "ENABLE_AQ = false":
        ws.cell(row=last_row, column=6).fill = failed_color

    # 欄寬/置頂/換行
    total_cols = 2 + num_condition_cols
    for col_idx in range(1, total_cols + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COMMON_WIDTH

    # 首列加粗並置頂
    for cell in ws[1]:
        cell.font = BOLD
        cell.alignment = COMMON_ALIGN
    # 本列置頂
    for cell in ws[last_row]:
        cell.alignment = COMMON_ALIGN

    # 移除預設 Sheet
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        try:
            wb.remove(wb["Sheet"])
        except Exception:
            pass

    wb.save(xlsx_path)

# -----------------------------
# Core parsing helpers (path mapping & file reading)
# -----------------------------

def _read_text(path: str) -> str:
    for enc in ("utf-8", "latin-1", "utf-16"):
        try:
            with open(path, "r", encoding=enc) as f:
                return f.read()
        except UnicodeDecodeError:
            continue
        except FileNotFoundError:
            raise
    with open(path, "r") as f:
        return f.read()

def _strip_comment(line: str) -> str:
    # 去掉 # 或 ; 後面的註解
    line = line.split("#", 1)[0]
    line = line.split(";", 1)[0]
    return line.strip()

def _resolve_tvconfigs_path(root: str, tvconfigs_like: str) -> str:
    """
    把 "/tvconfigs/xxx/yyy.ini" 映射為 "<root>/xxx/yyy.ini"
    其他相對路徑: 以 root 為基底；非 /tvconfigs 絕對路徑維持不動
    """
    if tvconfigs_like.startswith("/tvconfigs/"):
        rel = tvconfigs_like[len("/tvconfigs/"):]
        return os.path.normpath(os.path.join(root, rel))
    if tvconfigs_like.startswith("./") or tvconfigs_like.startswith("../"):
        return os.path.normpath(os.path.join(root, tvconfigs_like))
    if tvconfigs_like.startswith("/"):
        return tvconfigs_like
    return os.path.normpath(os.path.join(root, tvconfigs_like))

# -----------------------------
# Model.ini → TvServIni
# -----------------------------

def parse_model_ini_for_tvserv(model_ini_path: str, root: str) -> Optional[str]:
    """
    從 model.ini 找:
      TvServIni = "<path>"
    回傳對應到檔案系統的路徑（已映射到 --root）
    """
    txt = _read_text(model_ini_path)
    tvserv_path = None
    for raw in txt.splitlines():
        line = _strip_comment(raw)
        if not line:
            continue
        m = re.match(r'^\s*TvServIni\s*=\s*"?([^"]+)"?\s*$', line, re.IGNORECASE)
        if m:
            tvserv_path = _resolve_tvconfigs_path(root, m.group(1).strip())
            break
    return tvserv_path

# -----------------------------
# tvserv_ini parsing
# -----------------------------

TARGET_KEYS = {
    "DEFAULT_PICTURE_MODE",
    "DEFAULT_DOLBY_PICTURE_MODE",
    "ENABLE_AQ",
    "SUPPORT_MAT",
    "SUPPORT_DOLBY_CERT",
}

def parse_tvserv_flags(tvserv_ini_path: str) -> Dict[str, str]:
    """
    解析簡單 key=value，忽略註解與空白。
    關注鍵大小寫不敏感，回存為大寫鍵。
    """
    out: Dict[str, str] = {}
    if not tvserv_ini_path or not os.path.exists(tvserv_ini_path):
        return out

    txt = _read_text(tvserv_ini_path)
    for raw in txt.splitlines():
        line = _strip_comment(raw)
        if not line or "=" not in line:
            continue
        key, val = line.split("=", 1)
        ku = key.strip().upper()
        if ku in TARGET_KEYS:
            out[ku] = val.strip()
    return out

# -----------------------------
# Check logic
# -----------------------------

REQUIRED = {
    "DEFAULT_PICTURE_MODE": "9",
    "DEFAULT_DOLBY_PICTURE_MODE": "1",
    "ENABLE_AQ": "false",
    "SUPPORT_MAT": "true",
    "SUPPORT_DOLBY_CERT": "true",
}

def _normalize_bool_like(s: str) -> str:
    # 僅用於布林值比較：true/false（大小寫不敏感）
    return str(s).strip().lower()

def evaluate(vals: Dict[str, str]) -> Tuple[bool, List[str]]:
    """
    回傳 (passed, notes)
    notes 會列出缺失或不符的項目
    """
    notes: List[str] = []
    passed = True
    for k, expect in REQUIRED.items():
        if k not in vals:
            notes.append(f"缺少 {k}")
            passed = False
        else:
            actual = vals[k]
            if expect in ("true", "false"):
                a = _normalize_bool_like(actual)
                e = expect
                if a != e:
                    notes.append(f"{k} 不符 (expect {expect}, got {actual})")
                    passed = False
            else:
                if str(actual).strip() != str(expect).strip():
                    notes.append(f"{k} 不符 (expect {expect}, got {actual})")
                    passed = False
    return passed, notes

# -----------------------------
# Main
# -----------------------------

def main():
    p = argparse.ArgumentParser(
        description=(
            "Check TvServIni for required flags: "
            "DEFAULT_PICTURE_MODE=9, DEFAULT_DOLBY_PICTURE_MODE=1, "
            "ENABLE_AQ=false, SUPPORT_MAT=true, SUPPORT_DOLBY_CERT=true. "
            "With optional Excel report."
        )
    )
    p.add_argument("--model-ini", required=True, help="path to model ini (e.g., model/1_xxx.ini)")
    p.add_argument("--root", required=True, help="tvconfigs project root (maps /tvconfigs/* to here)")
    p.add_argument("--report", action="store_true", help="export report to xlsx (default: kipling.xlsx)")
    p.add_argument("--report-xlsx", metavar="FILE", help="export report to specific xlsx file")
    p.add_argument("-v", "--verbose", action="store_true", help="verbose logs")
    args = p.parse_args()

    model_ini = args.model_ini
    if not os.path.exists(model_ini):
        raise SystemExit(f"[ERROR] model ini not found: {model_ini}")

    root = os.path.abspath(os.path.normpath(args.root))
    if args.verbose:
        print(f"[INFO] model_ini: {model_ini}")
        print(f"[INFO] root     : {root}")

    # 1) 取 TvServIni
    tvserv_path = parse_model_ini_for_tvserv(model_ini, root)
    if args.verbose:
        print(f"[INFO] TvServIni : {tvserv_path or '(not found in model.ini)'}")

    vals: Dict[str, str] = {}
    notes: List[str] = []

    if not tvserv_path:
        notes.append("model.ini 未找到 TvServIni")
        passed = False
    elif not os.path.exists(tvserv_path):
        notes.append(f"TvServIni 指向檔案不存在: {tvserv_path}")
        passed = False
    else:
        # 2) 解析 tvserv_ini，抽出五項旗標
        vals = parse_tvserv_flags(tvserv_path)
        passed, more = evaluate(vals)
        notes.extend(more)

    # Console 輸出（僅 PASS/FAIL）
    print(f"Result  : {'PASS' if passed else 'FAIL'}")

    # 3) 報表
    if args.report or args.report_xlsx:
        xlsx_path = args.report_xlsx if args.report_xlsx else "kipling.xlsx"
        res = {
            "passed": passed,
            "model_ini": model_ini,
            "tvserv_ini_path": tvserv_path or "",
            "vals": vals,
            "notes": notes,
        }
        export_report(res, xlsx_path=xlsx_path)
        sheet = _sheet_name_for_model(model_ini)
        print(f"[INFO] Report appended to: {xlsx_path} (sheet: {sheet})")

if __name__ == "__main__":
    main()
