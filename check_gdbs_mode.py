#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import argparse, os, re
from typing import Optional

# -----------------------------
# Utilities for report
# -----------------------------

def _sheet_name_for_model(model_ini_path: str) -> str:
    import os, re
    base = os.path.basename(model_ini_path or "")
    m = re.match(r"^(\d+)_", base)
    if m:
        return f"PID_{int(m.group(1))}"
    return "others"

def _ensure_openpyxl():
    try:
        import openpyxl  # noqa
    except ImportError:
        raise SystemExit(
            "[ERROR] 需要 openpyxl 以支援報表輸出與附加。\n"
            "  安裝： pip install --user openpyxl\n"
        )

def export_report(res: dict, xlsx_path: str = "kipling.xlsx", num_condition_cols: int = 5) -> None:
    _ensure_openpyxl()
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    COMMON_WIDTH = 80
    COMMON_ALIGN = Alignment(wrap_text=True, vertical="top")
    BOLD = Font(bold=True)

    def _na(s: str) -> str:
        s = (s or "").strip()
        return s if s else "N/A"

    sheet_name = _sheet_name_for_model(res.get("model_ini", ""))

    # 開啟或新建 xlsx
    try:
        wb = load_workbook(xlsx_path)
    except Exception:
        wb = Workbook()

    # 建立或取得 sheet
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)
        headers = ["Rules", "Result"] + [f"condition_{i}" for i in range(1, num_condition_cols + 1)]
        ws.append(headers)

    # 準備資料
    rules    = "DV_GDBS_DELAY exist?\nGDBS_MODE=1?"
    result   = "PASS" if res.get("passed", False) else "FAIL"
    dv_path  = _na(res.get("dv_gdbs_delay", ""))
    gdbs_mode = _na(res.get("gdbs_mode", ""))

    conds = [
        f"DV_GDBS_DELAY = {dv_path}",   # condition_1
        f"GDBS_MODE = {gdbs_mode}",     # condition_2
    ]
 
    row_values = [rules, result] + conds
    ws.append(row_values)
    last_row = ws.max_row

    total_cols = 2 + num_condition_cols
    for col_idx in range(1, total_cols + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = COMMON_WIDTH

    for cell in ws[1]:  # header
        cell.font = BOLD
        cell.alignment = COMMON_ALIGN

    for cell in ws[last_row]:
        cell.alignment = COMMON_ALIGN

    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        try:
            wb.remove(wb["Sheet"])
        except Exception:
            pass

    wb.save(xlsx_path)

# -----------------------------
# Core parsing / validation
# -----------------------------

def _read_text(path: str) -> str:
    for enc in ("utf-8", "latin-1", "utf-16"):
        try:
            with open(path, "r", encoding=enc) as f:
                return f.read()
        except UnicodeDecodeError:
            continue
        except FileNotFoundError:
            raise
    with open(path, "r") as f:
        return f.read()

def _strip_comment(line: str) -> str:
    line = line.split("#", 1)[0]
    line = line.split(";", 1)[0]
    return line.strip()

def _resolve_tvconfigs_path(root: str, tvconfigs_like: str) -> str:
    if tvconfigs_like.startswith("/tvconfigs/"):
        rel = tvconfigs_like[len("/tvconfigs/"):]
        return os.path.normpath(os.path.join(root, rel))
    if tvconfigs_like.startswith("./") or tvconfigs_like.startswith("../"):
        return os.path.normpath(os.path.join(root, tvconfigs_like))
    if tvconfigs_like.startswith("/"):
        return tvconfigs_like
    return os.path.normpath(os.path.join(root, tvconfigs_like))

def parse_model_ini_for_gdbs(model_ini_path: str, root: str) -> Optional[str]:
    txt = _read_text(model_ini_path)
    for raw in txt.splitlines():
        line = _strip_comment(raw)
        if not line:
            continue
        m = re.match(r'^\s*DV_GDBS_DELAY\s*=\s*"?([^"]+)"?\s*$', line, re.IGNORECASE)
        if m:
            return _resolve_tvconfigs_path(root, m.group(1).strip())
    return None

def parse_gdbs_mode(target_file: str) -> Optional[str]:
    if not target_file or not os.path.exists(target_file):
        return None
    txt = _read_text(target_file)
    for raw in txt.splitlines():
        line = _strip_comment(raw)
        m = re.match(r'^\s*GDBS_MODE\s*=\s*(\d+)\s*$', line, re.IGNORECASE)
        if m:
            return m.group(1)
    return None

def build_result(model_ini: str, target_file: Optional[str], gdbs_mode: Optional[str]) -> dict:
    passed = (target_file and os.path.exists(target_file) and gdbs_mode == "1")
    return {
        "passed": passed,
        "model_ini": model_ini,
        "dv_gdbs_delay": target_file or "",
        "gdbs_mode": gdbs_mode or "N/A",
    }

# -----------------------------
# Main
# -----------------------------

def main():
    parser = argparse.ArgumentParser(description="Check DV_GDBS_DELAY -> GDBS_MODE=1")
    parser.add_argument("--model-ini", required=True, help="path to model ini")
    parser.add_argument("--root", required=True, help="tvconfigs root path")
    parser.add_argument("-v", "--verbose", action="store_true")
    parser.add_argument("--report", action="store_true", help="export to kipling.xlsx")
    parser.add_argument("--report-xlsx", metavar="FILE", help="export to specific xlsx")

    args = parser.parse_args()
    model_ini = args.model_ini
    root = os.path.abspath(args.root)

    target_file = parse_model_ini_for_gdbs(model_ini, root)
    gdbs_mode = parse_gdbs_mode(target_file) if target_file else None
    res = build_result(model_ini, target_file, gdbs_mode)

    print(f"[INFO] model_ini : {model_ini}")
    print(f"[INFO] DV_GDBS_DELAY : {target_file or '(not found)'}")
    print(f"[INFO] GDBS_MODE : {gdbs_mode or '(not found)'}")
    print(f"Result: {'PASS' if res['passed'] else 'FAIL'}")

    if args.report or args.report_xlsx:
        xlsx_path = args.report_xlsx if args.report_xlsx else "kipling.xlsx"
        export_report(res, xlsx_path=xlsx_path, num_condition_cols=5)

if __name__ == "__main__":
    main()
