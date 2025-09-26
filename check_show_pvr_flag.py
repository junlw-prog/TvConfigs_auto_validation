#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
check_show_pvr_flag.py

檢查規則:
- 在指定的 model.ini 內, 搜尋 "SHOW_PVR = false" (忽略大小寫, 並忽略註解行)
- 若找到, 則 PASS
- 反之 FAIL

輸出:
- 終端顯示 PASS/FAIL
- 若指定 --report 或 --report-xlsx, 會輸出到 kipling.xlsx (或指定檔案)
  格式與 tv_multi_standard_validation.py 相同 (Rules, Result, condition_1...)
"""

import argparse
import os
import re
from typing import Dict

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def _sheet_name_for_model(model_ini_path: str) -> str:
    base = os.path.basename(model_ini_path or "")
    m = re.match(r"^(\d+)_", base)
    if m:
        return f"PID_{int(m.group(1))}"
    return "others"


def _read_text(path: str) -> str:
    for enc in ("utf-8", "latin-1", "utf-16"):
        try:
            with open(path, "r", encoding=enc) as f:
                return f.read()
        except UnicodeDecodeError:
            continue
        except FileNotFoundError:
            raise
    with open(path, "r") as f:
        return f.read()


def _strip_comment(line: str) -> str:
    line = line.split("#", 1)[0]
    line = line.split(";", 1)[0]
    return line.strip()


def check_show_pvr(model_ini: str) -> Dict:
    txt = _read_text(model_ini)
    found_line = None
    for raw in txt.splitlines():
        line = _strip_comment(raw)
        if not line:
            continue
        if re.match(r'^\s*SHOW_PVR\s*=\s*false\s*$', line, re.IGNORECASE):
            found_line = raw.strip()
            break
    passed = bool(found_line)
    return {
        "passed": passed,
        "model_ini": model_ini,
        "setting_line": found_line or "N/A",
    }


def export_report(res: Dict, xlsx_path: str = "kipling.xlsx", num_condition_cols: int = 3) -> None:
    COMMON_WIDTH = 80
    COMMON_ALIGN = Alignment(wrap_text=True, vertical="top")
    BOLD = Font(bold=True)

    sheet_name = _sheet_name_for_model(res.get("model_ini", ""))

    try:
        wb = load_workbook(xlsx_path)
    except Exception:
        wb = Workbook()

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)
        headers = ["Rules", "Result"] + [f"condition_{i}" for i in range(1, num_condition_cols + 1)]
        ws.append(headers)

    rules = "1. Disable PVR\n" \
            "  1) model.ini->SHOW_PVR=false"
    result = "PASS" if res.get("passed", False) else "FAIL"
    conds = [
        #f"model.ini = {res.get('model_ini', 'N/A')}",
        f"{res.get('setting_line', 'N/A')}",
    ][:num_condition_cols]

    row = [rules, result] + conds
    ws.append(row)
    last_row = ws.max_row

    # 給儲存格指派上色
    rules_color = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
    failed_color = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")
    # 上色
    first_cell = ws.cell(row=last_row, column=1)  # 欄位1對應的是 'A' 列
    first_cell.fill = rules_color
    if result == "FAIL":
        ws.cell(row=last_row, column=2).fill = failed_color
    if conds[0] == "N/A":
        ws.cell(row=last_row, column=3).fill = failed_color

    total_cols = 2 + num_condition_cols
    for col_idx in range(1, total_cols + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = COMMON_WIDTH

    for cell in ws[1]:
        cell.font = BOLD
        cell.alignment = COMMON_ALIGN

    for cell in ws[last_row]:
        cell.alignment = COMMON_ALIGN

    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        try:
            wb.remove(wb["Sheet"])
        except Exception:
            pass

    wb.save(xlsx_path)


def main():
    parser = argparse.ArgumentParser(description="Check SHOW_PVR flag in model.ini")
    parser.add_argument("--model-ini", required=True, help="path to model ini")
    parser.add_argument("--report", action="store_true", help="export report to kipling.xlsx")
    parser.add_argument("--report-xlsx", metavar="FILE", help="export report to specific xlsx file")
    args = parser.parse_args()

    res = check_show_pvr(args.model_ini)

    print(f"Result : {'PASS' if res['passed'] else 'FAIL'}")
    print(f" - model.ini      : {res['model_ini']}")
    print(f" - setting line   : {res['setting_line']}")

    if args.report or args.report_xlsx:
        xlsx_path = args.report_xlsx if args.report_xlsx else "kipling.xlsx"
        export_report(res, xlsx_path=xlsx_path)
        print(f"[INFO] Report appended to: {xlsx_path} (sheet: {_sheet_name_for_model(res['model_ini'])})")


if __name__ == "__main__":
    main()
