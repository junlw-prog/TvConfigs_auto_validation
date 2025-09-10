#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
check_dias_4k60.py

依照 tv_multi_standard_validation.py 的「路徑解析」與「頁簽命名」規則，
新增 xlsx 輸出（去除 Model INI / Panel File 欄，且不自動補 N/A 欄位）。

檢查內容：
  - DISP_HORIZONTAL_TOTAL >= 3840
  - DISP_VERTICAL_TOTAL   >= 2160
  - DISPLAY_REFRESH_RATE  >= 60
"""

import argparse
import os
import re
import sys
from pathlib import Path
from typing import Dict, List, Tuple, Optional

# -----------------------------
# 共用：與 tv_multi_standard_validation.py 對齊
# -----------------------------

def _strip_comment(line: str) -> str:
    # 去掉 # 或 ; 後面的註解
    line = line.split("#", 1)[0]
    line = line.split(";", 1)[0]
    return line.strip()

def _resolve_tvconfigs_path(root: str, tvconfigs_like: str) -> str:
    """
    把 "/tvconfigs/xxx/yyy.ini" 映射為 "<root>/xxx/yyy.ini"
    其他相對路徑: 以 root 為基底
    絕對路徑（非 /tvconfigs 開頭）維持不動
    """
    if tvconfigs_like.startswith("/tvconfigs/"):
        rel = tvconfigs_like[len("/tvconfigs/"):]
        return os.path.normpath(os.path.join(root, rel))
    if tvconfigs_like.startswith("./") or tvconfigs_like.startswith("../"):
        return os.path.normpath(os.path.join(root, tvconfigs_like))
    if tvconfigs_like.startswith("/"):
        # 非 /tvconfigs 絕對路徑：原樣返回
        return tvconfigs_like
    # 其他當作相對於 root
    return os.path.normpath(os.path.join(root, tvconfigs_like))

def _sheet_name_for_model(model_ini_path: str) -> str:
    """
    根據 model.ini 檔名的前綴決定頁簽：
      - 前綴是數字 N → 'PID_N'
      - 其他 → 'others'
    例: '1_EU_XXX.ini' → 'PID_1'
    """
    base = os.path.basename(model_ini_path or "")
    m = re.match(r"^(\d+)_", base)
    if m:
        return f"PID_{int(m.group(1))}"
    return "others"

def _ensure_openpyxl():
    try:
        import openpyxl  # noqa
    except ImportError:
        raise SystemExit(
            "[ERROR] 需要 openpyxl 以支援報表輸出與附加。\n"
            "  安裝： pip install --user openpyxl\n"
        )

# -----------------------------
# 解析 model.ini 與 panel.ini
# -----------------------------

PANEL_NAME_RE = re.compile(r'^\s*m_pPanelName\s*=\s*"(?P<path>[^"]+)"\s*;', re.I)

def parse_panel_path_from_model_ini(model_ini: Path) -> Optional[str]:
    """讀取 model.ini 中 m_pPanelName 的路徑字串。"""
    try:
        with model_ini.open("r", encoding="utf-8", errors="ignore") as f:
            for raw in f:
                m = PANEL_NAME_RE.match(raw)
                if m:
                    return m.group("path").strip()
    except FileNotFoundError:
        return None
    return None

def read_panel_values(panel_ini_path: Path) -> Dict[str, Optional[float]]:
    """讀取 panel.ini 三個關鍵數值（可為 int/float），缺值回傳 None。"""
    keys = {
        "DISP_HORIZONTAL_TOTAL": None,
        "DISP_VERTICAL_TOTAL": None,
        "DISPLAY_REFRESH_RATE": None,
    }
    if not panel_ini_path.exists():
        return keys
    with panel_ini_path.open("r", encoding="utf-8", errors="ignore") as f:
        for raw in f:
            line = _strip_comment(raw)
            if not line or "=" not in line:
                continue
            k, v = [x.strip() for x in line.split("=", 1)]
            if k in keys:
                # 去掉結尾的分號與單位
                v = v.rstrip(";")
                try:
                    keys[k] = float(v)
                except ValueError:
                    # 不是數字就忽略
                    pass
    return keys

# -----------------------------
# 報表輸出（移除 Model INI、Panel File 欄位；不自動補 N/A）
# -----------------------------

def export_report(res: Dict, xlsx_path: str) -> None:
    """
    表頭僅包含: Rules, Result, condition_1..N（依實際條件數量產生，不補 N/A）
    依 model.ini 檔名前綴分頁（PID_1、PID_2…；非數字→others），既有資料則附加。
    """
    _ensure_openpyxl()
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    COMMON_WIDTH = 80
    COMMON_ALIGN = Alignment(wrap_text=True, vertical="top")
    BOLD = Font(bold=True)

    # 取欄數（Rules + Result + len(conditions)）
    conds: List[str] = res.get("conditions", [])
    rules: str = res.get("rules", "")
    passed: bool = bool(res.get("passed", False))
    sheet_name = _sheet_name_for_model(res.get("model_ini", ""))

    # 開檔或建立
    if os.path.exists(xlsx_path):
        wb = load_workbook(xlsx_path)
    else:
        wb = Workbook()

    # 取得/建立工作表
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)
        # 建立表頭
        headers = ["Rules", "Result"] + [f"condition_{i+1}" for i in range(len(conds) or 1)]
        ws.append(headers)
        # 樣式
        for col_idx in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = COMMON_WIDTH
        for cell in ws[1]:
            cell.font = BOLD
            cell.alignment = COMMON_ALIGN

    # 寫入 row
    row_values = [rules, "PASS" if passed else "FAIL"] + (conds if conds else [""])
    ws.append(row_values)
    last_row = ws.max_row

    # 套用樣式到該列
    total_cols = len(row_values)
    for col_idx in range(1, total_cols + 1):
        ws.cell(row=last_row, column=col_idx).alignment = COMMON_ALIGN

    # 如果存在預設 Sheet 且有其它工作表，移除預設 Sheet 以保持乾淨
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        try:
            wb.remove(wb["Sheet"])
        except Exception:
            pass

    wb.save(xlsx_path)

# -----------------------------
# 主程式
# -----------------------------

def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="Check DIAS 4K60 panel timing and export xlsx report.")
    parser.add_argument("--model-ini", required=True, help="path to model ini (e.g., model/1_xxx.ini)")
    parser.add_argument("--root", required=True, help="tvconfigs project root (maps /tvconfigs/* to here)")
    parser.add_argument("--report", action="store_true", help="export report to xlsx (default: kipling.xlsx)")
    parser.add_argument("--report-xlsx", metavar="FILE", help="export report to specific xlsx file")
    parser.add_argument("-v", "--verbose", action="store_true", help="verbose logs")
    args = parser.parse_args(argv)

    model_ini = Path(args.model_ini)
    if not model_ini.exists():
        print(f"[ERROR] model ini not found: {model_ini}", file=sys.stderr)
        return 2

    # 解析 panel 路徑（從 model.ini 取 m_pPanelName）
    raw_panel_path = parse_panel_path_from_model_ini(model_ini)
    if not raw_panel_path:
        print("[FAIL] 找不到 m_pPanelName 設定於 model.ini。")
        return 1

    panel_resolved = Path(_resolve_tvconfigs_path(args.root, raw_panel_path))
    if args.verbose:
        print(f"[INFO] model.ini : {model_ini}")
        print(f"[INFO] panel.raw : {raw_panel_path}")
        print(f"[INFO] panel.path: {panel_resolved}")

    vals = read_panel_values(panel_resolved)
    # 驗證條件
    conds = [
        f"DISP_HORIZONTAL_TOTAL={vals['DISP_HORIZONTAL_TOTAL']} (>=3840)",
        f"DISP_VERTICAL_TOTAL={vals['DISP_VERTICAL_TOTAL']} (>=2160)",
        f"DISPLAY_REFRESH_RATE={vals['DISPLAY_REFRESH_RATE']} (>=60)",
    ]

    ok = True
    reasons: List[str] = []
    if vals["DISP_HORIZONTAL_TOTAL"] is None or vals["DISP_HORIZONTAL_TOTAL"] < 3840:
        ok = False
        reasons.append("DISP_HORIZONTAL_TOTAL < 3840 或缺少")
    if vals["DISP_VERTICAL_TOTAL"] is None or vals["DISP_VERTICAL_TOTAL"] < 2160:
        ok = False
        reasons.append("DISP_VERTICAL_TOTAL < 2160 或缺少")
    if vals["DISPLAY_REFRESH_RATE"] is None or vals["DISPLAY_REFRESH_RATE"] < 60:
        ok = False
        reasons.append("DISPLAY_REFRESH_RATE < 60 或缺少")

    # Console 輸出
    if ok:
        print("[PASS] 所有條件皆符合（>=3840, >=2160, >=60）。")
    else:
        print("[FAIL] 不符合條件：")
        for r in reasons:
            print(f"  - {r}")

    # 報表
    if args.report or args.report_xlsx:
        xlsx = args.report_xlsx if args.report_xlsx else "kipling.xlsx"
        res = {
            "passed": ok,
            "rules": "DIAS_4K60",
            "conditions": conds,
            "model_ini": str(model_ini),
        }
        export_report(res, xlsx_path=xlsx)
        print(f"[INFO] Report appended to: {xlsx} (sheet: {_sheet_name_for_model(str(model_ini))})")

    return 0 if ok else 1

if __name__ == "__main__":
    sys.exit(main())
