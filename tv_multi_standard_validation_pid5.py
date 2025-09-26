#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import argparse
import os
import re
from typing import Dict, List, Tuple, Optional


# -----------------------------
# Utilities for report
# -----------------------------

def _sheet_name_for_model(model_ini_path: str) -> str:
    """
    根據 model.ini 檔名的前綴決定頁簽：
      - 前綴是數字 N → 'PID_N'
      - 其他 → 'others'
    例: '1_EU_XXX.ini' → 'PID_1'
    """
    base = os.path.basename(model_ini_path or "")
    m = re.match(r"^(\d+)_", base)
    if m:
        return f"PID_{int(m.group(1))}"
    return "others"


def _ensure_openpyxl():
    try:
        import openpyxl  # noqa
    except ImportError:
        raise SystemExit(
            "[ERROR] 需要 openpyxl 以支援報表輸出與附加。\n"
            "  安裝： pip install --user openpyxl\n"
        )

#def export_report(res: dict, xlsx_path: str = "kipling.xlsx", sheet_name: str) -> None:
def export_report(res: dict, xlsx_path: str, sheet_name: str,  num_condition_cols: int = 5) -> None:
    """
    欄位無值時以 'N/A' 填入。依 model.ini 檔名前綴分頁（PID_1、PID_2…；非數字→others），既有資料則附加。
    表頭固定為: Rules, Result, condition_1, condition_2, condition_3, ...
    統一：所有欄位同寬、同為自動換行、垂直置頂（包含表頭）。
    """
    _ensure_openpyxl()
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    COMMON_WIDTH = 80
    COMMON_ALIGN = Alignment(wrap_text=True, vertical="top")
    BOLD = Font(bold=True)

    def _na(s: str) -> str:
        s = (s or "").strip()
        return s if s else "N/A"

    # 開啟或新建 xlsx
    try:
        wb = load_workbook(xlsx_path)
    except Exception:
        wb = Workbook()

    # 建立或取得 sheet
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)
        headers = ["Rules", "Result"] + [f"condition_{i}" for i in range(1, num_condition_cols + 1)]
        ws.append(headers)

    # 給儲存格指派上色
    rules_color = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
    failed_color = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")

    # 準備資料
    rules    = f"6. 要 enable 多制式切換\n" \
               f"    - model.ini->COUNTRY_PATH是否宣告?\n" \
               f"    - model.ini->tvSysMap是否宣告?"
    result   = "PASS" if res.get("passed", False) else "FAIL"
    standard = _na(res.get("standard", ""))
    tvsysmap = _na(res.get("tv_sys_map", ""))
    cpath    = _na(res.get("country_path", ""))
    targets  = _na(", ".join(res.get("customer_target_countries", []) or []))
    missing  = _na(", ".join(res.get("missing", []) or []))

    # condition values
    conds = [
        f"TvSysMap = {tvsysmap}",     # condition_1
        f"Country Path = {cpath}",    # condition_2
        f"Standard = {standard}",     # condition_3
        f"Target Countries = {targets}",  # condition_4
        f"Missing = {missing}",       # condition_5
    ][:num_condition_cols]

    # 寫入 row
    row_values = [rules, result] + conds
    ws.append(row_values)
    last_row = ws.max_row

    # 上色
    first_cell = ws.cell(row=last_row, column=1)  # 欄位1對應的是 'A' 列
    first_cell.fill = rules_color
    if cpath == "N/A":
        ws.cell(row=last_row, column=4).fill = failed_color
    if tvsysmap == "N/A":
        ws.cell(row=last_row, column=3).fill = failed_color
    if result == "FAIL":
        ws.cell(row=last_row, column=2).fill = failed_color

    # 套用樣式：欄寬、換行、垂直靠上
    total_cols = 2 + num_condition_cols
    for col_idx in range(1, total_cols + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = COMMON_WIDTH

    for cell in ws[1]:  # header
        cell.font = BOLD
        cell.alignment = COMMON_ALIGN

    for cell in ws[last_row]:
        cell.alignment = COMMON_ALIGN

    # 移除預設 Sheet
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        try:
            wb.remove(wb["Sheet"])
        except Exception:
            pass

    wb.save(xlsx_path)

# -----------------------------
# Core parsing / validation
# -----------------------------

def _read_text(path: str) -> str:
    for enc in ("utf-8", "latin-1", "utf-16"):
        try:
            with open(path, "r", encoding=enc) as f:
                return f.read()
        except UnicodeDecodeError:
            continue
        except FileNotFoundError:
            raise
    # 若皆失敗，讓原始例外往外拋
    with open(path, "r") as f:
        return f.read()


def _strip_comment(line: str) -> str:
    # 去掉 # 或 ; 之後的註解
    line = line.split("#", 1)[0]
    line = line.split(";", 1)[0]
    return line.strip()


def _resolve_tvconfigs_path(root: str, tvconfigs_like: str) -> str:
    """
    把 "/tvconfigs/xxx/yyy.ini" 映射為 "<root>/xxx/yyy.ini"
    其他相對路徑: 以 root 為基底
    絕對路徑（非 /tvconfigs 開頭）維持不動
    """
    if tvconfigs_like.startswith("/tvconfigs/"):
        rel = tvconfigs_like[len("/tvconfigs/"):]
        return os.path.normpath(os.path.join(root, rel))
    if tvconfigs_like.startswith("./") or tvconfigs_like.startswith("../"):
        return os.path.normpath(os.path.join(root, tvconfigs_like))
    if tvconfigs_like.startswith("/"):
        # 非 /tvconfigs 絕對路徑：原樣返回
        return tvconfigs_like
    # 其他當作相對於 root
    return os.path.normpath(os.path.join(root, tvconfigs_like))


def parse_model_ini_for_paths(model_ini_path: str, root: str) -> Tuple[Optional[str], Optional[str]]:
    """
    從 model.ini 找：
      - tvSysMap = "<path>"
      - COUNTRY_PATH = "<path>"
    回傳對應到檔案系統的絕對/相對實體路徑（已映射到 --root）
    """
    txt = _read_text(model_ini_path)
    tvsysmap = None
    country_path = None

    # 找 tvSysMap 與 COUNTRY_PATH（忽略前置空白，允許有引號）
    for raw in txt.splitlines():
        line = _strip_comment(raw)
        if not line:
            continue
        m1 = re.match(r'^\s*tvSysMap\s*=\s*"?([^"]+)"?\s*$', line, re.IGNORECASE)
        if m1 and tvsysmap is None:
            tvsysmap = _resolve_tvconfigs_path(root, m1.group(1).strip())
            continue
        m2 = re.match(r'^\s*COUNTRY_PATH\s*=\s*"?([^"]+)"?\s*$', line, re.IGNORECASE)
        if m2 and country_path is None:
            country_path = _resolve_tvconfigs_path(root, m2.group(1).strip())
            continue

    return tvsysmap, country_path


def parse_country_list(country_ini_path: str) -> List[str]:
    """
    嘗試從 COUNTRY_PATH 指到的 ini 取出國家清單。
    設計成「寬鬆解析」：過濾註解與空白，將可能的國家 token（A-Z/_，逗號分隔）擷取出來。
    """
    if not country_ini_path or not os.path.exists(country_ini_path):
        return []

    txt = _read_text(country_ini_path)
    tokens: List[str] = []

    for raw in txt.splitlines():
        line = _strip_comment(raw)
        if not line:
            continue
        # 用逗號、空白、等號都拆開試試
        for part in re.split(r"[,\s=]+", line):
            part = part.strip()
            # 篩選看起來像國家名稱的 token（大寫＋底線）
            if part and re.fullmatch(r"[A-Z][A-Z0-9_]*", part):
                tokens.append(part)

    # 去重、維持順序
    seen = set()
    uniq = []
    for t in tokens:
        if t not in seen:
            seen.add(t)
            uniq.append(t)
    return uniq


def build_result(args, model_ini: str, tvsysmap: Optional[str], country_path: Optional[str], targets: List[str]) -> Dict:
    """
    產生報表所需的結果結構（你原本的檢查可改寫這裡填滿更多欄位）
    """
    missing: List[str] = []
    if tvsysmap and not os.path.exists(tvsysmap):
        missing.append(tvsysmap)
    if country_path and not os.path.exists(country_path):
        missing.append(country_path)

    passed = (tvsysmap and os.path.exists(tvsysmap)) and (country_path and os.path.exists(country_path))

    return {
        "standard": args.standard or "",
        "passed": bool(passed),
        #"model_ini": model_ini,
        "tv_sys_map": tvsysmap or "",
        "country_path": country_path or "",
        "customer_target_countries": targets,
        "missing": missing,
    }


# -----------------------------
# Main
# -----------------------------

def main():
    parser = argparse.ArgumentParser(description="TV multi-standard validation with Excel report (kipling.xlsx).")
    parser.add_argument("--model-ini", required=True, help="path to model ini (e.g., model/1_xxx.ini)")
    parser.add_argument("--root", required=True, help="tvconfigs project root (maps /tvconfigs/* to here)")
    parser.add_argument("--standard", choices=["DVB", "ATSC", "ISDB"], default=None, help="target standard")
    parser.add_argument("-v", "--verbose", action="store_true", help="verbose logs")

    # 報表參數：--report 與 --report-xlsx（任一存在即輸出；未指定路徑則用 kipling.xlsx）
    parser.add_argument("--report", action="store_true", help="export report to xlsx (default: kipling.xlsx)")
    parser.add_argument("--report-xlsx", metavar="FILE", help="export report to specific xlsx file")

    args = parser.parse_args()

    model_ini = args.model_ini
    if not os.path.exists(model_ini):
        raise SystemExit(f"[ERROR] model ini not found: {model_ini}")

    root = os.path.abspath(os.path.normpath(args.root))

    if args.verbose:
        print(f"[INFO] model_ini: {model_ini}")
        print(f"[INFO] root     : {root}")
        print(f"[INFO] standard : {args.standard or ''}")

    # 解析 model.ini -> tvSysMap 與 COUNTRY_PATH
    tvsysmap_path, country_path = parse_model_ini_for_paths(model_ini, root)
    if args.verbose:
        print(f"[INFO] tvSysMap     : {tvsysmap_path or '(not found in model.ini)'}")
        print(f"[INFO] COUNTRY_PATH : {country_path or '(not found in model.ini)'}")

    # 擷取國家清單（寬鬆解析）
    targets = parse_country_list(country_path) if country_path else []
    if args.verbose:
        print(f"[INFO] Target Countries: {', '.join(targets) if targets else '(none)'}")

    # 產生結果
    res = build_result(args, model_ini, tvsysmap_path, country_path, targets)

    # 簡單輸出到 console
    print(f"Standard: {res['standard']}")
    print(f"Result  : {'PASS' if res['passed'] else 'FAIL'}")

    # 報表輸出
    if args.report or args.report_xlsx:
        xlsx_path = args.report_xlsx if args.report_xlsx else "kipling.xlsx"
        sheet_name = _sheet_name_for_model(model_ini)
        export_report(res, xlsx_path, sheet_name)
        #sheet = _sheet_name_for_model(model_ini)
        print(f"[INFO] Report appended to: {xlsx_path} (sheet: {sheet_name})")


if __name__ == "__main__":
    main()
