
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
check_tvconfig_and_mheg5.py  (rev: drop model.ini/resolved-file columns & no auto N/A)

變更：
- 報表不再輸出「Model.ini」與「Resolved file」兩欄（兩個測項皆移除）。
- 取消自動以 N/A 補值：條件缺少就留白（空字串）。
- 其餘規格同前一版：兩個測項各佔一列、表頭/樣式/分頁規則沿用。
"""

import argparse
import os
import re
from typing import List, Optional

# -----------------------------
# Report helpers
# -----------------------------

def _ensure_openpyxl():
    try:
        import openpyxl  # noqa
    except ImportError:
        raise SystemExit(
            "[ERROR] 需要 openpyxl 以支援報表輸出與附加。\n"
            "  安裝： pip install --user openpyxl\n"
        )

def _sheet_name_for_model(model_ini_path: str) -> str:
    import re, os
    base = os.path.basename(model_ini_path or "")
    m = re.match(r"^(\d+)_", base)
    if m:
        return f"PID_{int(m.group(1))}"
    return "others"

def export_report_row(model_ini: str, rules: str, passed: bool, conditions: List[str], xlsx_path: str):
    """
    寫入一行報表：Rules, Result, condition_1..condition_10
    - 不再以 N/A 自動補值；沒有就留空白。
    - 既有工作簿附加；頁簽按 PID_x/others；首列粗體、欄寬 80、自動換行、垂直靠上。
    """
    _ensure_openpyxl()
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    COMMON_WIDTH = 80
    COMMON_ALIGN = Alignment(wrap_text=True, vertical="top")
    BOLD = Font(bold=True)

    sheet_name = _sheet_name_for_model(model_ini)

    # 開啟或新建 xlsx
    try:
        wb = load_workbook(xlsx_path)
    except Exception:
        wb = Workbook()

    # 建立或取得 sheet
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)
        headers = ["Rules", "Result"]
        for i in range(1, 11):
            headers.append(f"condition_{i}")
        ws.append(headers)

    # 填資料（不足 10 個條件會自動留白）
    row = [rules or "", "PASS" if passed else "FAIL"]
    row.extend(conditions[:10])
    # 若條件數小於 10，補空字串到 10
    if len(conditions) < 10:
        row.extend([""] * (10 - len(conditions)))

    ws.append(row)
    last_row = ws.max_row

    # 給儲存格指派上色
    rules_color = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
    failed_color = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")
    # 上色
    first_cell = ws.cell(row=last_row, column=1)  # 欄位1對應的是 'A' 列
    first_cell.fill = rules_color
    if row[1] == "FAIL":
        ws.cell(row=last_row, column=2).fill = failed_color

    # 樣式
    total_cols = len(ws[1])
    for col_idx in range(1, total_cols + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = COMMON_WIDTH

    for cell in ws[1]:  # header
        cell.font = BOLD
        cell.alignment = COMMON_ALIGN

    for cell in ws[last_row]:
        cell.alignment = COMMON_ALIGN

    # 移除預設 Sheet
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        try:
            wb.remove(wb["Sheet"])
        except Exception:
            pass

    wb.save(xlsx_path)

# -----------------------------
# Parsing helpers
# -----------------------------

def _read_text(path: str) -> str:
    for enc in ("utf-8", "latin-1", "utf-16"):
        try:
            with open(path, "r", encoding=enc) as f:
                return f.read()
        except UnicodeDecodeError:
            continue
        except FileNotFoundError:
            raise
    with open(path, "r") as f:
        return f.read()

def _strip_comment(line: str) -> str:
    line = line.split("#", 1)[0]
    line = line.split(";", 1)[0]
    return line.strip()

def _resolve_tvconfigs_path(root: str, tvconfigs_like: str) -> str:
    if tvconfigs_like.startswith("/tvconfigs/"):
        rel = tvconfigs_like[len("/tvconfigs/"):]
        return os.path.normpath(os.path.join(root, rel))
    if tvconfigs_like.startswith("./") or tvconfigs_like.startswith("../"):
        return os.path.normpath(os.path.join(root, tvconfigs_like))
    if tvconfigs_like.startswith("/"):
        return tvconfigs_like
    return os.path.normpath(os.path.join(root, tvconfigs_like))

def find_tvconfig_value_case_sensitive(model_ini_path: str) -> Optional[str]:
    txt = _read_text(model_ini_path)
    for raw in txt.splitlines():
        line = _strip_comment(raw)
        if not line:
            continue
        m = re.match(r'^\s*TvConfig\s*=\s*"?([^"]+)"?\s*$', line)  # case sensitive key
        if m:
            return m.group(1).strip()
    return None

# -----------------------------
# Core checks
# -----------------------------

ALLOWED_TVCONFIG_VALUES = {
    "/tvconfigs/tv_config/tv.config.dvb_ntsc",
    "/tvconfigs/tv_config/tv.config.dvbt_ntsc",
}

def check_tvconfig_path(tvconfig_value: Optional[str]) -> bool:
    return tvconfig_value in ALLOWED_TVCONFIG_VALUES

def check_mheg5_flag(tvconfig_fs_path: Optional[str]) -> bool:
    #if not tvconfig_fs_path or not os.path.exists(tvconfig_fs_path):
        #return False
    txt = _read_text(tvconfig_fs_path)
    pattern = r'^\s*persist\.vendor\.rtk\.tv\.enable_mheg5\s*=\s*false\s*$'
    for line in txt.splitlines():
        if re.match(pattern, line):
            return True
    return "persist.vendor.rtk.tv.enable_mheg5=false" in txt

# -----------------------------
# Main
# -----------------------------

def main():
    parser = argparse.ArgumentParser(description="Check TvConfig value and MHEG5 flag; export Excel in tv_multi_standard_validation style.")
    parser.add_argument("--model-ini", required=True, help="path to model ini (e.g., model/1_xxx.ini)")
    parser.add_argument("--root", required=True, help="tvconfigs project root (maps /tvconfigs/* to here)")
    parser.add_argument("-v", "--verbose", action="store_true", help="verbose logs")
    parser.add_argument("--report", action="store_true", help="export report to xlsx (default: kipling.xlsx)")
    parser.add_argument("--report-xlsx", metavar="FILE", help="export report to specific xlsx file")

    args = parser.parse_args()

    model_ini = args.model_ini
    if not os.path.exists(model_ini):
        raise SystemExit(f"[ERROR] model ini not found: {model_ini}")

    root = os.path.abspath(os.path.normpath(args.root))

    if args.verbose:
        print(f"[INFO] model_ini: {model_ini}")
        print(f"[INFO] root     : {root}")

    # 1) 取得 TvConfig 設定
    tvconfig_value = find_tvconfig_value_case_sensitive(model_ini)
    if args.verbose:
        print(f"[INFO] TvConfig: {tvconfig_value or '(not found)'}")

    # 映射到實際檔案路徑
    tvconfig_fs_path = _resolve_tvconfigs_path(root, tvconfig_value) if tvconfig_value else None
    if args.verbose:
        print(f"[INFO] tvconfig file: {tvconfig_fs_path or '(N/A)'}  exists={os.path.exists(tvconfig_fs_path) if tvconfig_fs_path else False}")

    # 測項一：TvConfig 值是否為允許清單
    path_ok = check_tvconfig_path(tvconfig_value)
    rules_1 = (
        "4.Model ini:\n" 
        "  公版 TvConfig = /tvconfigs/tv_config/tv.config.dvb_ntsc\n"
        "  哥倫比亞 TvConfig = /tvconfigs/tv_config/tv.config.dvbt_ntsc\n"
    )
    conditions_1 = [
        f"TvConfig = {tvconfig_value or ''}",
        #f"File exists = {os.path.exists(tvconfig_fs_path) if tvconfig_fs_path else False}",
    ]

    # 測項二：tvconfig 內容是否含 enable_mheg5=false
    mheg5_is_false = check_mheg5_flag(tvconfig_fs_path)
    rules_2 = "3.沒有mheg5,因為columbia and tawian沒有"
    conditions_2 = [
        f"TvConfig = {tvconfig_value or ''}",
        #f"File exists = {os.path.exists(tvconfig_fs_path) if tvconfig_fs_path else False}",
        f"persist.vendor.rtk.tv.enable_mheg5={False if mheg5_is_false else True}",
    ]

    # console output
    print(f"[CHECK-1] TvConfig path allowed? -> {'PASS' if path_ok else 'FAIL'}")
    print(f"[CHECK-2] enable_mheg5=true present? -> {'PASS' if mheg5_is_false else 'FAIL'}")

    # export report
    if args.report or args.report_xlsx:
        xlsx_path = args.report_xlsx if args.report_xlsx else "kipling.xlsx"
        export_report_row(model_ini, rules_2, mheg5_is_false, conditions_2, xlsx_path)
        export_report_row(model_ini, rules_1, path_ok, conditions_1, xlsx_path)
        sheet = _sheet_name_for_model(model_ini)
        print(f"[INFO] Report appended to: {xlsx_path} (sheet: {sheet}, +2 rows)")

if __name__ == "__main__":
    main()
