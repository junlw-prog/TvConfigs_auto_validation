#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
check_tvservini_picture_modes.py

參考 tv_multi_standard_validation.py：
- 檔案路徑解析：將 "/tvconfigs/..." 映射到 --root 下；相對路徑以 --root 為基底；其他絕對路徑原樣保留
- 報表樣式：相同的欄位、同寬、換行、垂直靠上；分頁規則 PID_N / others；已存在 xlsx 則附加

功能：
從 model.ini 找到 TvServIni 指向的 tvserv_ini 檔案，檢查以下三個設定是否符合預期：
  1) DEFAULT_PICTURE_MODE = 4
  2) DEFAULT_DOLBY_PICTURE_MODE = 1
  3) SUPPORT_MAT = true   （大小寫不敏感）
三者皆符合 → PASS；否則 → FAIL

用法範例：
python3 check_tvservini_picture_modes.py --model-ini model/1_xxx.ini --root . --report
python3 check_tvservini_picture_modes.py --model-ini model/1_xxx.ini --root . --report-xlsx result.xlsx -v
"""

import argparse
import os
import re
from typing import Dict, List, Tuple, Optional

# -----------------------------
# Utilities for report (same style as tv_multi_standard_validation.py)
# -----------------------------

def _sheet_name_for_model(model_ini_path: str) -> str:
    """
    根據 model.ini 檔名的前綴決定頁簽：
      - 前綴是數字 N → 'PID_N'
      - 其他 → 'others'
    例: '1_EU_XXX.ini' → 'PID_1'
    """
    base = os.path.basename(model_ini_path or "")
    m = re.match(r"^(\d+)_", base)
    if m:
        return f"PID_{int(m.group(1))}"
    return "others"


def _ensure_openpyxl():
    try:
        import openpyxl  # noqa
    except ImportError:
        raise SystemExit(
            "[ERROR] 需要 openpyxl 以支援報表輸出與附加。\n"
            "  安裝： pip install --user openpyxl\n"
        )


def export_report(res: dict, xlsx_path: str = "kipling.xlsx", num_condition_cols: int = 6) -> None:
    """
    欄位無值時以 'N/A' 填入。依 model.ini 檔名前綴分頁（PID_1、PID_2…；非數字→others），既有資料則附加。
    表頭固定為: Rules, Result, condition_1, condition_2, condition_3, ...
    統一：所有欄位同寬、同為自動換行、垂直置頂（包含表頭）。
    """
    _ensure_openpyxl()
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    COMMON_WIDTH = 80
    COMMON_ALIGN = Alignment(wrap_text=True, vertical="top")
    BOLD = Font(bold=True)

    def _na(s: str) -> str:
        s = (s or "").strip()
        return s if s else "N/A"

    sheet_name = _sheet_name_for_model(res.get("model_ini", ""))

    # 開啟或新建 xlsx
    try:
        wb = load_workbook(xlsx_path)
    except Exception:
        wb = Workbook()

    # 建立或取得 sheet
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)
        headers = ["Rules", "Result"] + [f"condition_{i}" for i in range(1, num_condition_cols + 1)]
        ws.append(headers)

    # 準備資料
    rules = "TvServIni flags must match: DEFAULT_PICTURE_MODE=4, DEFAULT_DOLBY_PICTURE_MODE=1, SUPPORT_MAT=true"
    result = "PASS" if res.get("passed", False) else "FAIL"

    tvservini = _na(res.get("tvservini_path", ""))
    c_pm      = _na(res.get("check_default_picture_mode", ""))
    c_dpm     = _na(res.get("check_default_dolby_picture_mode", ""))
    c_mat     = _na(res.get("check_support_mat", ""))
    missing   = _na(", ".join(res.get("missing", []) or []))
    model_ini = _na(res.get("model_ini", ""))

    # condition values（可依需求調整欄位順序/內容）
    conds = [
        f"TvServIni = {tvservini}",                   # condition_1
        f"DEFAULT_PICTURE_MODE → {c_pm}",             # condition_2
        f"DEFAULT_DOLBY_PICTURE_MODE → {c_dpm}",      # condition_3
        f"SUPPORT_MAT → {c_mat}",                     # condition_4
        f"Missing = {missing}",                       # condition_5
        f"Model.ini = {model_ini}",                   # condition_6
    ][:num_condition_cols]

    # 寫入 row
    row_values = [rules, result] + conds
    ws.append(row_values)
    last_row = ws.max_row

    # 套用樣式：欄寬、換行、垂直靠上（包含表頭）
    total_cols = 2 + num_condition_cols
    for col_idx in range(1, total_cols + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = COMMON_WIDTH

    for cell in ws[1]:  # header
        cell.font = BOLD
        cell.alignment = COMMON_ALIGN

    for cell in ws[last_row]:
        cell.alignment = COMMON_ALIGN

    # 移除預設 Sheet
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        try:
            wb.remove(wb["Sheet"])
        except Exception:
            pass

    wb.save(xlsx_path)

# -----------------------------
# Core parsing / validation
# -----------------------------

def _read_text(path: str) -> str:
    for enc in ("utf-8", "latin-1", "utf-16"):
        try:
            with open(path, "r", encoding=enc) as f:
                return f.read()
        except UnicodeDecodeError:
            continue
        except FileNotFoundError:
            raise
    with open(path, "r") as f:
        return f.read()


def _strip_comment(line: str) -> str:
    # 去掉 # 或 ; 之後的註解
    line = line.split("#", 1)[0]
    line = line.split(";", 1)[0]
    return line.strip()


def _resolve_tvconfigs_path(root: str, tvconfigs_like: str) -> str:
    """
    把 "/tvconfigs/xxx/yyy.ini" 映射為 "<root>/xxx/yyy.ini"
    其他相對路徑: 以 root 為基底
    絕對路徑（非 /tvconfigs 開頭）維持不動
    """
    if tvconfigs_like.startswith("/tvconfigs/"):
        rel = tvconfigs_like[len("/tvconfigs/"):]
        return os.path.normpath(os.path.join(root, rel))
    if tvconfigs_like.startswith("./") or tvconfigs_like.startswith("../"):
        return os.path.normpath(os.path.join(root, tvconfigs_like))
    if tvconfigs_like.startswith("/"):
        # 非 /tvconfigs 絕對路徑：原樣返回
        return tvconfigs_like
    # 其他當作相對於 root
    return os.path.normpath(os.path.join(root, tvconfigs_like))


def parse_model_ini_for_tvservini(model_ini_path: str, root: str) -> Optional[str]:
    """
    從 model.ini 找：
      - TvServIni = "<path>"
    回傳對應到檔案系統的實體路徑（已映射到 --root）
    """
    txt = _read_text(model_ini_path)
    tvservini = None

    # 找 TvServIni（忽略前置空白，允許有引號；忽略大小寫）
    for raw in txt.splitlines():
        line = _strip_comment(raw)
        if not line:
            continue
        m = re.match(r'^\s*TvServIni\s*=\s*"?([^"]+)"?\s*$', line, re.IGNORECASE)
        if m and tvservini is None:
            tvservini = _resolve_tvconfigs_path(root, m.group(1).strip())
            break

    return tvservini


def _parse_key_values_ini_like(path: str) -> Dict[str, str]:
    """
    很寬鬆的 ini 解析：
      - 忽略空白與註解（# 或 ; 後面）
      - 接受 "KEY = VALUE" / "KEY=VALUE"
      - key：大小寫不敏感；value：保留原樣（左右空白會 strip）
    """
    kv: Dict[str, str] = {}
    if not path or not os.path.exists(path):
        return kv

    txt = _read_text(path)
    for raw in txt.splitlines():
        line = _strip_comment(raw)
        if not line or "=" not in line:
            continue
        key, val = line.split("=", 1)
        key = key.strip().lower()
        val = val.strip()
        if key:
            kv[key] = val
    return kv


def check_flags(tvservini_path: Optional[str]) -> Dict[str, str]:
    """
    檢查三個旗標，回傳字串描述供報表填寫：
      - DEFAULT_PICTURE_MODE == 4
      - DEFAULT_DOLBY_PICTURE_MODE == 1
      - SUPPORT_MAT.lower() == "true"
    任何條件缺值 → 視為不符合（並在 missing 中標記路徑）
    """
    desc_pm = "N/A"
    desc_dpm = "N/A"
    desc_mat = "N/A"
    ok_pm = ok_dpm = ok_mat = False

    if tvservini_path and os.path.exists(tvservini_path):
        kv = _parse_key_values_ini_like(tvservini_path)

        # DEFAULT_PICTURE_MODE
        v_pm = kv.get("default_picture_mode")
        if v_pm is not None:
            ok_pm = (v_pm.strip() == "4")
            desc_pm = f"expected 4, found {v_pm}"
        else:
            desc_pm = "expected 4, found (missing)"

        # DEFAULT_DOLBY_PICTURE_MODE
        v_dpm = kv.get("default_dolby_picture_mode")
        if v_dpm is not None:
            ok_dpm = (v_dpm.strip() == "1")
            desc_dpm = f"expected 1, found {v_dpm}"
        else:
            desc_dpm = "expected 1, found (missing)"

        # SUPPORT_MAT
        v_mat = kv.get("support_mat")
        if v_mat is not None:
            ok_mat = (v_mat.strip().lower() == "true")
            desc_mat = f"expected true, found {v_mat}"
        else:
            desc_mat = "expected true, found (missing)"

    return {
        "desc_pm": desc_pm,
        "desc_dpm": desc_dpm,
        "desc_mat": desc_mat,
        "ok_pm": ok_pm,
        "ok_dpm": ok_dpm,
        "ok_mat": ok_mat,
    }


def build_result(args, model_ini: str, tvservini_path: Optional[str], flag_chk: Dict[str, str]) -> Dict:
    missing: List[str] = []
    if tvservini_path and not os.path.exists(tvservini_path):
        missing.append(tvservini_path)
    if not tvservini_path:
        missing.append("(TvServIni not found in model.ini)")

    passed = bool(flag_chk.get("ok_pm") and flag_chk.get("ok_dpm") and flag_chk.get("ok_mat"))

    return {
        "passed": passed,
        "model_ini": model_ini,
        "tvservini_path": tvservini_path or "",
        "check_default_picture_mode": flag_chk.get("desc_pm", ""),
        "check_default_dolby_picture_mode": flag_chk.get("desc_dpm", ""),
        "check_support_mat": flag_chk.get("desc_mat", ""),
        "missing": missing,
    }

# -----------------------------
# Main
# -----------------------------

def main():
    parser = argparse.ArgumentParser(description="Check TvServIni flags (DEFAULT_PICTURE_MODE, DEFAULT_DOLBY_PICTURE_MODE, SUPPORT_MAT) with Excel report.")
    parser.add_argument("--model-ini", required=True, help="path to model ini (e.g., model/1_xxx.ini)")
    parser.add_argument("--root", required=True, help="tvconfigs project root (maps /tvconfigs/* to here)")
    parser.add_argument("-v", "--verbose", action="store_true", help="verbose logs")

    # 報表參數：--report 與 --report-xlsx（任一存在即輸出；未指定路徑則用 kipling.xlsx）
    parser.add_argument("--report", action="store_true", help="export report to xlsx (default: kipling.xlsx)")
    parser.add_argument("--report-xlsx", metavar="FILE", help="export report to specific xlsx file")

    args = parser.parse_args()

    model_ini = args.model_ini
    if not os.path.exists(model_ini):
        raise SystemExit(f"[ERROR] model ini not found: {model_ini}")

    root = os.path.abspath(os.path.normpath(args.root))

    if args.verbose:
        print(f"[INFO] model_ini: {model_ini}")
        print(f"[INFO] root     : {root}")

    # 解析 model.ini -> TvServIni
    tvservini_path = parse_model_ini_for_tvservini(model_ini, root)
    if args.verbose:
        print(f"[INFO] TvServIni : {tvservini_path or '(not found in model.ini)'}")

    # 檢查旗標
    flag_chk = check_flags(tvservini_path)

    # 產生結果
    res = build_result(args, model_ini, tvservini_path, flag_chk)

    # 簡單輸出到 console
    print(f"Result  : {'PASS' if res['passed'] else 'FAIL'}")
    if args.verbose:
        print(f"  - DEFAULT_PICTURE_MODE       → {res['check_default_picture_mode']}")
        print(f"  - DEFAULT_DOLBY_PICTURE_MODE → {res['check_default_dolby_picture_mode']}")
        print(f"  - SUPPORT_MAT                → {res['check_support_mat']}")
        if res['missing']:
            print(f"  - Missing: {', '.join(res['missing'])}")

    # 報表輸出
    if args.report or args.report_xlsx:
        xlsx_path = args.report_xlsx if args.report_xlsx else "kipling.xlsx"
        export_report(res, xlsx_path=xlsx_path)
        sheet = _sheet_name_for_model(model_ini)
        print(f"[INFO] Report appended to: {xlsx_path} (sheet: {sheet})")


if __name__ == "__main__":
    main()
