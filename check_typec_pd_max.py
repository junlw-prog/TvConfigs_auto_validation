
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
check_typec_pd_max.py
---------------------
參考 tv_multi_standard_validation.py 的路徑解析、xlsx 輸出格式與 sheet 命名規則。
功能：
1) 解析 model.ini，取得 PcbPath 的值；將 /tvconfigs/* 映射到 --root 之下。
2) 打開 PcbPath 指向的檔案，搜尋參數 TYPEC_PD_MAX 的值（大小寫不敏感，忽略註解）；
   若找到則打印其值；未找到則視為 N/A。
3) 產生 xlsx 報表（預設 kipling.xlsx、可用 --report-xlsx 指定）；
   表頭固定為：Rules, Result, condition_1..N；Result 欄固定為 "N/A"；
   sheet 依 model.ini 檔名前綴 PID_x 命名（非數字前綴則 "others"）。
"""
import argparse
import os
import re
from typing import Optional

# -----------------------------
# Utilities
# -----------------------------

def _sheet_name_for_model(model_ini_path: str) -> str:
    base = os.path.basename(model_ini_path or "")
    m = re.match(r"^(\d+)_", base)
    if m:
        return f"PID_{int(m.group(1))}"
    return "others"


def _ensure_openpyxl():
    try:
        import openpyxl  # noqa
    except ImportError:
        raise SystemExit(
            "[ERROR] 需要 openpyxl 以支援報表輸出與附加。\n"
            "  安裝： pip install --user openpyxl\n"
        )


def _read_text(path: str) -> str:
    for enc in ("utf-8", "latin-1", "utf-16"):
        try:
            with open(path, "r", encoding=enc) as f:
                return f.read()
        except UnicodeDecodeError:
            continue
        except FileNotFoundError:
            raise
    with open(path, "r") as f:
        return f.read()


def _strip_comment(line: str) -> str:
    line = line.split("#", 1)[0]
    line = line.split(";", 1)[0]
    return line.strip()


def _resolve_tvconfigs_path(root: str, tvconfigs_like: str) -> str:
    if not tvconfigs_like:
        return ""
    tvconfigs_like = tvconfigs_like.strip().strip('"').strip("'")
    if tvconfigs_like.startswith("/tvconfigs/"):
        rel = tvconfigs_like[len("/tvconfigs/"):]
        return os.path.normpath(os.path.join(root, rel))
    if tvconfigs_like.startswith("./") or tvconfigs_like.startswith("../"):
        return os.path.normpath(os.path.join(root, tvconfigs_like))
    if tvconfigs_like.startswith("/"):
        return tvconfigs_like
    return os.path.normpath(os.path.join(root, tvconfigs_like))


# -----------------------------
# Parsing
# -----------------------------

def parse_model_ini_for_pcbpath(model_ini_path: str, root: str) -> Optional[str]:
    """
    從 model.ini 讀取 PcbPath 的值，並映射到檔案系統路徑。
    支援：大小寫不敏感、可含引號；忽略註解。
    """
    txt = _read_text(model_ini_path)
    pcb_rel = None
    for raw in txt.splitlines():
        line = _strip_comment(raw)
        if not line:
            continue
        m = re.match(r'^\s*PcbPath\s*=\s*"?([^"]+)"?\s*$', line, re.IGNORECASE)
        if m:
            pcb_rel = m.group(1).strip()
            break
    if not pcb_rel:
        return None
    return _resolve_tvconfigs_path(root, pcb_rel)


def parse_typec_pd_max(pcb_file_path: str) -> Optional[str]:
    """
    從 pcb 檔案內搜尋 TYPEC_PD_MAX 的設定值。
    允許格式：
      TYPEC_PD_MAX = 123
      TYPEC_PD_MAX=TRUE
      TYPEC_PD_MAX : 45
    取 '=' 或 ':' 後第一段非空白字串。
    """
    if not pcb_file_path or not os.path.exists(pcb_file_path):
        return None
    txt = _read_text(pcb_file_path)
    pattern = re.compile(r'^\s*TYPEC_PD_MAX\s*[:=]\s*([^\s#;]+)', re.IGNORECASE)
    for raw in txt.splitlines():
        line = _strip_comment(raw)
        m = pattern.match(line)
        if m:
            return m.group(1).strip()
    return None


# -----------------------------
# Report
# -----------------------------

def export_report(model_ini: str,
                  pcb_path_cfg: Optional[str],
                  pcb_path_resolved: Optional[str],
                  typec_pd_max: Optional[str],
                  xlsx_path: str = "kipling.xlsx",
                  num_condition_cols: int = 5) -> None:
    """
    表頭固定: Rules, Result, condition_1..N；其中 Result 固定填 "N/A"
    """
    _ensure_openpyxl()
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    COMMON_WIDTH = 80
    COMMON_ALIGN = Alignment(wrap_text=True, vertical="top")
    BOLD = Font(bold=True)

    def _na(v: Optional[str]) -> str:
        s = (v or "").strip() if isinstance(v, str) else ("" if v is None else str(v))
        return s if s else "N/A"

    sheet_name = _sheet_name_for_model(model_ini)

    # 開啟或新建 xlsx
    try:
        wb = load_workbook(xlsx_path)
    except Exception:
        wb = Workbook()

    # 建立或取得 sheet
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)
        headers = ["Rules", "Result"] + [f"condition_{i}" for i in range(1, num_condition_cols + 1)]
        ws.append(headers)

    rules = "Parse model.ini PcbPath → open pcb file → read TYPEC_PD_MAX"
    result = "N/A"  # 依需求固定為 N/A

    conds = [
        f"model.ini: PcbPath = {_na(pcb_path_cfg)}",
        f"PCB file = {_na(pcb_path_resolved)} (exists={os.path.exists(pcb_path_resolved) if pcb_path_resolved else False})",
        f"TYPEC_PD_MAX = {_na(typec_pd_max)}",
        "Notes = N/A",
        "Extra = N/A",
    ][:num_condition_cols]

    row_values = [rules, result] + conds
    ws.append(row_values)
    last_row = ws.max_row

    total_cols = 2 + num_condition_cols
    for col_idx in range(1, total_cols + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = COMMON_WIDTH

    for cell in ws[1]:
        cell.font = BOLD
        cell.alignment = COMMON_ALIGN

    for cell in ws[last_row]:
        cell.alignment = COMMON_ALIGN

    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        try:
            wb.remove(wb["Sheet"])
        except Exception:
            pass

    wb.save(xlsx_path)


# -----------------------------
# Main
# -----------------------------

def main():
    parser = argparse.ArgumentParser(description="Check TYPEC_PD_MAX from PcbPath with Excel report (kipling.xlsx)")
    parser.add_argument("--model-ini", required=True, help="path to model ini (e.g., model/1_xxx.ini)")
    parser.add_argument("--root", required=True, help="tvconfigs project root (maps /tvconfigs/* to here)")
    parser.add_argument("-v", "--verbose", action="store_true", help="verbose logs")

    # 報表參數
    parser.add_argument("--report", action="store_true", help="export report to xlsx (default: kipling.xlsx)")
    parser.add_argument("--report-xlsx", metavar="FILE", help="export report to specific xlsx file")

    args = parser.parse_args()

    model_ini = args.model_ini
    if not os.path.exists(model_ini):
        raise SystemExit(f"[ERROR] model ini not found: {model_ini}")

    root = os.path.abspath(os.path.normpath(args.root))

    if args.verbose:
        print(f"[INFO] model_ini: {model_ini}")
        print(f"[INFO] root     : {root}")

    pcb_cfg = None
    pcb_path = None
    try:
        pcb_path = parse_model_ini_for_pcbpath(model_ini, root)
        if pcb_path:
            # 反推原始 PcbPath 設定（相對或 /tvconfigs 開頭），僅供輸出顯示
            # 直接再讀一次原文抓字串，以利表格呈現
            txt = _read_text(model_ini)
            m = re.search(r'^\s*PcbPath\s*=\s*"?([^"]+)"?\s*$', txt, re.IGNORECASE | re.MULTILINE)
            if m:
                pcb_cfg = m.group(1).strip()
    except FileNotFoundError:
        pcb_path = None

    if args.verbose:
        print(f"[INFO] PcbPath (cfg) : {pcb_cfg or '(not found)'}")
        print(f"[INFO] PcbPath (file): {pcb_path or '(resolve failed)'}")

    typec_val = parse_typec_pd_max(pcb_path) if pcb_path else None

    # Console 簡述
    print("TYPEC_PD_MAX :", typec_val if typec_val else "N/A")

    # Excel 報表
    if args.report or args.report_xlsx:
        xlsx_path = args.report_xlsx if args.report_xlsx else "kipling.xlsx"
        export_report(
            model_ini=model_ini,
            pcb_path_cfg=pcb_cfg,
            pcb_path_resolved=pcb_path,
            typec_pd_max=typec_val,
            xlsx_path=xlsx_path
        )
        sheet = _sheet_name_for_model(model_ini)
        print(f"[INFO] Report appended to: {xlsx_path} (sheet: {sheet})")


if __name__ == "__main__":
    main()
