
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
check_dias_5k.py (with XLSX report, aligned to check_dias_4k60.py)
------------------------------------------------------------------
用途：
1) 從 model.ini 解析 m_pPanelName 指向的面板檔案路徑；
2) 依照 /tvconfigs/... 轉為專案根目錄下的相對路徑；
3) 檢查是否符合（>=）：
     - DISP_HORIZONTAL_TOTAL >= 5120
     - DISP_VERTICAL_TOTAL   >= 2880
     - DISPLAY_REFRESH_RATE  >= 60
4) 輸出 XLSX 報表（與 check_dias_4k60.py 風格一致）：
   - 分頁：依 model.ini 檔名數字前綴決定 PID_<N>，否則 others
   - 表頭：Rules, Result, condition_1..condition_N（動態欄位，不自動補 N/A）
   - 不輸出「Model INI / Panel File」欄位
   - 既有活頁簿則附加；欄寬一致、換行、垂直靠上

相容：Python 3.8+
"""

import argparse
import re
import sys
import os
from pathlib import Path

# -----------------------------
# Report helpers
# -----------------------------

def _ensure_openpyxl():
    try:
        import openpyxl  # noqa
    except ImportError:
        raise SystemExit(
            "[ERROR] 需要 openpyxl 以支援報表輸出。\n"
            "安裝： pip install --user openpyxl\n"
        )

def _sheet_name_for_model(model_ini_path: str) -> str:
    """
    根據 model.ini 檔名的前綴決定頁簽：
      - 前綴是數字 N → 'PID_N'
      - 其他 → 'others'
    例: '1_EU_XXX.ini' → 'PID_1'
    """
    base = os.path.basename(model_ini_path or "")
    m = re.match(r"^(\d+)_", base)
    if m:
        return f"PID_{int(m.group(1))}"
    return "others"

def export_report(res: dict, xlsx_path: str = "kipling.xlsx", conditions: list = None) -> None:
    """
    動態輸出欄位，不做 N/A padding：
    表頭固定前兩欄: Rules, Result；
    後續依據實際條件數生成 condition_1..condition_N。
    若工作表已存在且表頭的 condition 欄不足本次需要，會自動擴增。
    """
    _ensure_openpyxl()
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    COMMON_WIDTH = 80
    COMMON_ALIGN = Alignment(wrap_text=True, vertical="top")
    BOLD = Font(bold=True)

    sheet_name = _sheet_name_for_model(res.get("model_ini", ""))
    conds = list(conditions or [])

    # 開啟或新建 xlsx
    try:
        wb = load_workbook(xlsx_path)
    except Exception:
        wb = Workbook()

    # 建立或取得 sheet
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # 若是空表，補 header
        if ws.max_row < 1:
            ws.append(["Rules", "Result"])
    else:
        ws = wb.create_sheet(title=sheet_name)
        ws.append(["Rules", "Result"])  # 先放兩欄，稍後依需要擴增

    # 依需要擴增 header 的 condition 欄位
    header = [c.value for c in ws[1]]
    current_cond_cols = max(0, len(header) - 2)
    needed_cond_cols = len(conds)
    if needed_cond_cols > current_cond_cols:
        new_header = ["Rules", "Result"] + [f"condition_{i}" for i in range(1, needed_cond_cols + 1)]
        ws.delete_rows(1)
        ws.append(new_header)
        header = new_header

    # 準備資料
    rules  = res.get("rules", "DIAS panel check: H_TOTAL>=5120, V_TOTAL>=2880, REFRESH_RATE>=60")
    result = "PASS" if res.get("passed", False) else "FAIL"
    row_values = [rules, result] + conds

    # 寫入 row
    ws.append(row_values)
    last_row = ws.max_row

    # 套用樣式：欄寬、換行、垂直靠上（含表頭）
    total_cols = ws.max_column
    for col_idx in range(1, total_cols + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = COMMON_WIDTH

    for cell in ws[1]:
        cell.font = BOLD
        cell.alignment = COMMON_ALIGN

    for cell in ws[last_row]:
        cell.alignment = COMMON_ALIGN

    # 移除預設 Sheet
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        try:
            wb.remove(wb["Sheet"])
        except Exception:
            pass

    wb.save(xlsx_path)

# -----------------------------
# Core logic
# -----------------------------

KEY_PATTERNS = {
    "DISP_HORIZONTAL_TOTAL": re.compile(r"^\s*DISP_HORIZONTAL_TOTAL\s*=\s*([0-9]+)", re.IGNORECASE),
    "DISP_VERTICAL_TOTAL":   re.compile(r"^\s*DISP_VERTICAL_TOTAL\s*=\s*([0-9]+)", re.IGNORECASE),
    "DISPLAY_REFRESH_RATE":  re.compile(r"^\s*DISPLAY_REFRESH_RATE\s*=\s*([0-9]+)", re.IGNORECASE),
}

THRESHOLDS = {
    "DISP_HORIZONTAL_TOTAL": 5120,
    "DISP_VERTICAL_TOTAL":   2880,
    "DISPLAY_REFRESH_RATE":  60,
}

PANEL_NAME_RE = re.compile(
    r'^\s*m_pPanelName\s*=\s*"(.*?)"\s*;?.*$',  # 擷取雙引號中的路徑
    re.IGNORECASE
)

def resolve_panel_path(raw_path: str, root: Path) -> Path:
    """將 model.ini 中的 m_pPanelName 路徑轉為實際檔案路徑。
    - "/tvconfigs/..." → 以 root 為基底（去掉前綴）
    - "/panel/..."     → root/panel/...
    - 相對路徑         → root/相對路徑
    - 其他絕對路徑     → 原樣使用
    """
    raw_path = raw_path.strip()
    if raw_path.startswith("/tvconfigs/"):
        sub = raw_path[len("/tvconfigs/"):]
        return root / sub
    if raw_path.startswith("/panel/"):
        return root / raw_path.lstrip("/")
    p = Path(raw_path)
    if p.is_absolute():
        return p
    return root / p

def parse_panel_name(model_ini: Path) -> str:
    """從 model.ini 讀取 m_pPanelName 內容（雙引號內字串）。"""
    try:
        with model_ini.open("r", encoding="utf-8", errors="ignore") as f:
            for line in f:
                m = PANEL_NAME_RE.match(line)
                if m:
                    return m.group(1)
    except FileNotFoundError:
        raise FileNotFoundError(f"Model ini not found: {model_ini}")
    raise ValueError('找不到 m_pPanelName = "..." 行，請確認 model.ini。')

def extract_values(panel_ini: Path) -> dict:
    """從 panel 檔案擷取三個目標鍵的整數值。若缺少則不放入 dict。"""
    values = {}
    try:
        with panel_ini.open("r", encoding="utf-8", errors="ignore") as f:
            for line in f:
                for key, pat in KEY_PATTERNS.items():
                    if key in values:
                        continue
                    m = pat.match(line)
                    if m:
                        try:
                            values[key] = int(m.group(1))
                        except ValueError:
                            pass
    except FileNotFoundError:
        raise FileNotFoundError(f"Panel ini not found: {panel_ini}")
    return values

def check(values: dict) -> (bool, list):
    """檢查數值是否都「>=」門檻。回傳 (ok, errors)。"""
    errors = []
    for key, th in THRESHOLDS.items():
        if key not in values:
            errors.append(f"缺少 {key} 欄位")
            continue
        v = values[key]
        if v < th:
            errors.append(f"{key} = {v} 不滿足 >= {th}")
    return (len(errors) == 0, errors)

# -----------------------------
# Main
# -----------------------------

def main():
    ap = argparse.ArgumentParser(
        description="檢查 DIAS 5K panel 是否符合：H_TOTAL>=5120, V_TOTAL>=2880, REFRESH_RATE>=60；可輸出 XLSX 報表。"
    )
    ap.add_argument("--model-ini", required=True, help="model/*.ini 路徑")
    ap.add_argument("--root", default=".", help="專案根目錄（含 panel/ 子資料夾），預設為目前目錄")
    ap.add_argument("--report", action="store_true", help="輸出報表到 kipling.xlsx（若未提供 --report-xlsx）")
    ap.add_argument("--report-xlsx", metavar="FILE", help="自訂輸出報表路徑（.xlsx）")
    args = ap.parse_args()

    model_ini = Path(args.model_ini).resolve()
    root = Path(args.root).resolve()

    # 解析 panel 原始路徑
    try:
        raw_panel = parse_panel_name(model_ini)
    except Exception as e:
        print(f"[FAIL] 解析 model.ini 失敗：{e}")
        if args.report or args.report_xlsx:
            res = {
                "passed": False,
                "model_ini": str(model_ini),  # 用於決定分頁名稱
                "errors": [str(e)],
                "rules": "DIAS panel check: H_TOTAL>=5120, V_TOTAL>=2880, REFRESH_RATE>=60",
            }
            conditions = [
                "Panel Raw Path = <parse error>",
                "Thresholds: H>=5120, V>=2880, R>=60",
                "Errors: " + "; ".join(res["errors"]),
            ]
            export_report(res, xlsx_path=(args.report_xlsx or "kipling.xlsx"), conditions=conditions)
        return 2

    panel_path = resolve_panel_path(raw_panel, root).resolve()

    # 擷取 panel 參數
    try:
        values = extract_values(panel_path)
    except Exception as e:
        print(f"[FAIL] 開啟 panel 檔失敗：{e}")
        if args.report or args.report_xlsx:
            res = {
                "passed": False,
                "model_ini": str(model_ini),
                "errors": [str(e)],
                "rules": "DIAS panel check: H_TOTAL>=5120, V>=2880, R>=60",
            }
            conditions = [
                f"Panel Raw Path = {raw_panel}",
                "Thresholds: H>=5120, V>=2880, R>=60",
                "Errors: " + "; ".join(res["errors"]),
            ]
            export_report(res, xlsx_path=(args.report_xlsx or "kipling.xlsx"), conditions=conditions)
        return 3

    ok, errors = check(values)

    print("=== DIAS 5K Panel 檢查報告 ===")
    print(f"Model INI     : {model_ini}")
    print(f"Root          : {root}")
    print(f"Panel Raw Path: {raw_panel}")
    print(f"Panel File    : {panel_path}")
    print("擷取值：")
    for k in ["DISP_HORIZONTAL_TOTAL", "DISP_VERTICAL_TOTAL", "DISPLAY_REFRESH_RATE"]:
        v = values.get(k, "<缺少>")
        print(f"  - {k}: {v}")

    if ok:
        print("[PASS] 所有條件皆符合（>=5120, >=2880, >=60）。")
    else:
        print("[FAIL] 不符合條件：")
        for err in errors:
            print(f"  - {err}")

    # 報表輸出（無 Model INI/Panel File 欄位、無 N/A padding）
    if args.report or args.report_xlsx:
        res = {
            "passed": bool(ok),
            "model_ini": str(model_ini),  # 僅用於分頁命名
            "rules": "DIAS panel check: H_TOTAL>=5120, V_TOTAL>=2880, REFRESH_RATE>=60",
        }
        conds = [
            f"Panel Raw Path = {raw_panel}",
            f"Extracted: H={values.get('DISP_HORIZONTAL_TOTAL','<缺>')}, V={values.get('DISP_VERTICAL_TOTAL','<缺>')}, R={values.get('DISPLAY_REFRESH_RATE','<缺>')}",
            "Thresholds: H>=5120, V>=2880, R>=60",
        ]
        if errors:
            conds.append("Errors: " + "; ".join(errors))

        export_report(res, xlsx_path=(args.report_xlsx or "kipling.xlsx"), conditions=conds)

    return 0 if ok else 1

if __name__ == "__main__":
    sys.exit(main())
