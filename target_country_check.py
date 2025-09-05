#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
target_country_check.py

1) Read COUNTRY_PATH from model.ini and load countries inside as target_country.
2) From --standard (DVB/ATSC/ISDB), open root/country/<STANDARD>.ini as global_country.
3) Check every target_country is included in global_country; fail if any missing.
4) Print the comparison steps and append the result to an Excel report.

報表格式（對齊 tv_multi_standard_validation.py 的風格）：
  - 表頭固定順序：Rules, Result, condition_1, condition_2, condition_3, …
  - Result 僅輸出 PASS / FAIL（不含 "Result = " 前綴）
  - 不輸出 Model.ini 欄位（但仍用於決定分頁 PID_* / others）
  - 表頭粗體、所有欄位同寬、換行、垂直置頂（含表頭與資料列）
"""
import argparse
import os
import re
from typing import List, Optional

# -----------------------------
# Utilities for report (compatible style with tv_multi_standard_validation.py)
# -----------------------------

def _sheet_name_for_model(model_ini_path: str) -> str:
    base = os.path.basename(model_ini_path or "")
    m = re.match(r"^(\d+)_", base)
    if m:
        return f"PID_{int(m.group(1))}"
    return "others"


def _ensure_openpyxl():
    try:
        import openpyxl  # noqa
    except ImportError:
        raise SystemExit(
            "[ERROR] 需要 openpyxl 以支援報表輸出與附加。\n"
            "  安裝： pip install --user openpyxl\n"
        )


def export_report(res: dict, xlsx_path: str = "kipling.xlsx", num_condition_cols: int = 5) -> None:
    """
    表頭固定為: Rules, Result, condition_1, condition_2, condition_3, ...
    欄位無值時以 'N/A' 填入。依 model.ini 檔名前綴分頁（PID_1、PID_2…；非數字→others），既有資料則附加。
    全欄統一樣式：同寬、換行、垂直置頂（包含表頭）。
    """
    _ensure_openpyxl()
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    COMMON_WIDTH = 80
    COMMON_ALIGN = Alignment(wrap_text=True, vertical="top")
    BOLD = Font(bold=True)

    def _na(s: str) -> str:
        s = (s or "").strip()
        return s if s else "N/A"

    sheet_name = _sheet_name_for_model(res.get("model_ini", ""))

    # 開啟或新建 xlsx
    try:
        wb = load_workbook(xlsx_path)
    except Exception:
        wb = Workbook()

    # 建立或取得 sheet（表頭固定順序）
    header = ["Rules", "Result"] + [f"condition_{i}" for i in range(1, num_condition_cols + 1)]
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row < 1:
            ws.append(header)
    else:
        ws = wb.create_sheet(title=sheet_name)
        ws.append(header)

    # 取值
    rules      = _na(res.get("rules", ""))
    result     = _na(res.get("result", ""))
    conditions = [ _na(x) for x in (res.get("conditions", []) or []) ]

    # 補足 condition_* 欄位數
    if len(conditions) < num_condition_cols:
        conditions += ["N/A"] * (num_condition_cols - len(conditions))
    else:
        conditions = conditions[:num_condition_cols]

    # 寫入一列
    row_values = [rules, result] + conditions
    ws.append(row_values)
    last_row = ws.max_row

    # ── 統一樣式：所有欄位同寬 & 換行 & 垂直置頂（含表頭） ──
    total_cols = 2 + num_condition_cols
    for col_idx in range(1, total_cols + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = COMMON_WIDTH

    # 表頭樣式
    for cell in ws[1]:
        cell.font = BOLD
        cell.alignment = COMMON_ALIGN

    # 資料列樣式（最新一列）
    for col_idx in range(1, total_cols + 1):
        ws.cell(row=last_row, column=col_idx).alignment = COMMON_ALIGN

    # 移除預設空白 Sheet（若存在且非唯一）
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        try:
            wb.remove(wb["Sheet"])
        except Exception:
            pass

    wb.save(xlsx_path)


# -----------------------------
# Core parsing / validation
# -----------------------------

def _read_text(path: str) -> str:
    for enc in ("utf-8", "latin-1", "utf-16"):
        try:
            with open(path, "r", encoding=enc) as f:
                return f.read()
        except UnicodeDecodeError:
            continue
        except FileNotFoundError:
            raise
    # 若皆失敗，讓原始例外往外拋
    with open(path, "r") as f:
        return f.read()


def _strip_comment(line: str) -> str:
    # 去掉 # 或 ; 之後的註解
    line = line.split("#", 1)[0]
    line = line.split(";", 1)[0]
    return line.strip()


def _resolve_tvconfigs_path(root: str, tvconfigs_like: str) -> str:
    """
    把 "/tvconfigs/xxx/yyy.ini" 映射為 "<root>/xxx/yyy.ini"
    其他相對路徑: 以 root 為基底
    絕對路徑（非 /tvconfigs 開頭）維持不動
    """
    tvconfigs_like = tvconfigs_like.strip()
    if tvconfigs_like.startswith("/tvconfigs/"):
        rel = tvconfigs_like[len("/tvconfigs/"):]
        return os.path.normpath(os.path.join(root, rel))
    if tvconfigs_like.startswith("./") or tvconfigs_like.startswith("../"):
        return os.path.normpath(os.path.join(root, tvconfigs_like))
    if tvconfigs_like.startswith("/"):
        # 非 /tvconfigs 絕對路徑：原樣返回
        return tvconfigs_like
    # 其他當作相對於 root
    return os.path.normpath(os.path.join(root, tvconfigs_like))


def parse_model_ini_for_country_path(model_ini_path: str, root: str) -> Optional[str]:
    """
    從 model.ini 找：
      - COUNTRY_PATH = "<path>"（不分大小寫）
    回傳對應到檔案系統的實際路徑（已映射到 --root）
    """
    txt = _read_text(model_ini_path)
    country_path = None

    for raw in txt.splitlines():
        line = _strip_comment(raw)
        if not line:
            continue
        m = re.match(r'^\s*COUNTRY_PATH\s*=\s*"?([^"]+)"?\s*$', line, re.IGNORECASE)
        if not m:
            m = re.match(r'^\s*country_path\s*=\s*"?([^"]+)"?\s*$', line, re.IGNORECASE)
        if m and country_path is None:
            country_path = _resolve_tvconfigs_path(root, m.group(1).strip())
            continue

    return country_path


def parse_country_list(country_ini_path: str) -> List[str]:
    """
    寬鬆解析國家清單：
      - 過濾註解與空白
      - 以逗號、等號、空白拆字
      - 抽取看起來像國家 token（大寫、數字、底線）
    """
    if not country_ini_path or not os.path.exists(country_ini_path):
        return []

    txt = _read_text(country_ini_path)
    tokens: List[str] = []

    for raw in txt.splitlines():
        line = _strip_comment(raw)
        if not line:
            continue
        for part in re.split(r"[,\s=]+", line):
            part = part.strip()
            if part and re.fullmatch(r"[A-Z][A-Z0-9_]*", part):
                tokens.append(part)

    # 去重、維持順序
    seen = set()
    uniq = []
    for t in tokens:
        if t not in seen:
            seen.add(t)
            uniq.append(t)
    return uniq


def main():
    parser = argparse.ArgumentParser(description="Compare target COUNTRY_PATH countries vs global <STANDARD>.ini and export to Excel.")
    parser.add_argument("--model-ini", required=True, help="path to model ini (e.g., model/1_xxx.ini)")
    parser.add_argument("--root", required=True, help="tvconfigs project root (maps /tvconfigs/* to here)")
    parser.add_argument("--standard", required=True, choices=["DVB", "ATSC", "ISDB"], help="target standard to open country/<STANDARD>.ini")
    parser.add_argument("--report", action="store_true", help="export report to xlsx (default: kipling.xlsx)")
    parser.add_argument("--report-xlsx", metavar="FILE", help="export report to specific xlsx file")
    parser.add_argument("--conditions", type=int, default=5, help="condition_* 欄位數（預設 5）")
    parser.add_argument("-v", "--verbose", action="store_true", help="verbose logs")
    args = parser.parse_args()

    model_ini = args.model_ini
    if not os.path.exists(model_ini):
        raise SystemExit(f"[ERROR] model ini not found: {model_ini}")

    root = os.path.abspath(os.path.normpath(args.root))

    if args.verbose:
        print(f"[INFO] model_ini: {model_ini}")
        print(f"[INFO] root     : {root}")
        print(f"[INFO] standard : {args.standard}")

    # 1) 解析 model.ini → COUNTRY_PATH
    country_path = parse_model_ini_for_country_path(model_ini, root)
    if args.verbose:
        print(f"[INFO] COUNTRY_PATH : {country_path or '(not found in model.ini)'}")

    # 2) target_country from COUNTRY_PATH
    target_countries = parse_country_list(country_path) if country_path else []
    if args.verbose:
        print(f"[INFO] Target Countries ({len(target_countries)}): {', '.join(target_countries) if target_countries else '(none)'}")

    # 3) global_country from root/country/<STANDARD>.ini
    global_ini = os.path.join(root, "country", f"{args.standard}.ini")
    global_countries = parse_country_list(global_ini)
    if args.verbose:
        print(f"[INFO] Global Countries [{args.standard}] ({len(global_countries)}): {', '.join(global_countries) if global_countries else '(none)'}")
        if not os.path.exists(global_ini):
            print(f"[WARN] Global ini not found: {global_ini}")

    # 4) 比對：target 是否全部包含於 global
    global_set = set(global_countries)
    missing = [c for c in target_countries if c not in global_set]
    passed = (len(missing) == 0) and bool(target_countries or global_countries)

    # 印出比對過程
    print("=== Comparison ===")
    print(f"COUNTRY_PATH: {country_path or 'N/A'}")
    print(f"Global ini  : {global_ini if os.path.exists(global_ini) else (global_ini + ' (NOT FOUND)')}")
    print(f"Target -> {len(target_countries)} country tokens")
    print(f"Global -> {len(global_countries)} country tokens")
    if missing:
        print(f"[FAIL] Missing (target not in global): {', '.join(missing)}")
    else:
        print("[PASS] All target countries are present in global list.")

    # 準備報表資料（對齊 tv_multi_standard_validation.py 的格式；不含 Model.ini 欄位）
    result_text = "PASS" if passed else "FAIL"
    rules = f"COUNTRY_PATH countries are included in {args.standard}.ini ?"

    conditions = [
        f"Standard = {args.standard}",                                             # condition_1
        f"Country Path = {country_path if country_path else 'N/A'}",               # condition_2
        f"Target Countries = {', '.join(target_countries) if target_countries else 'N/A'}",  # condition_3
        f"Global Countries = {', '.join(global_countries) if global_countries else 'N/A'}",  # condition_4
        f"Missing = {', '.join(missing) if missing else 'N/A'}",                   # condition_5
    ]

    res = {
        "result": result_text,      # PASS / FAIL
        "rules": rules,             # Rules 欄位內容
        "model_ini": model_ini,     # 用於決定分頁（PID_1..others）
        "conditions": conditions,
    }

    # 螢幕輸出結尾摘要
    print("------------------")
    print(f"Standard: {args.standard}")
    print(f"Result  : {result_text}")

    # 報表輸出
    if args.report or args.report_xlsx:
        xlsx_path = args.report_xlsx if args.report_xlsx else "kipling.xlsx"
        export_report(res, xlsx_path=xlsx_path, num_condition_cols=args.conditions)
        sheet = _sheet_name_for_model(model_ini)
        print(f"[INFO] Report appended to: {xlsx_path} (sheet: {sheet})")


if __name__ == "__main__":
    main()
