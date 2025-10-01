#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
check_isBassTrebleCustomValue.py

需求：
- 參考 check_CI.py 的「/tvconfigs → --root」路徑映射、xlsx 輸出格式、sheet 命名規則
- 在 xlsx 內「移除 model.ini 欄位」（僅用於決定分頁名，不輸出）
- 直接在 model.ini 內搜尋設定：isBassTrebleCustomValue
- 將「找到的值」打印到終端，並寫入報表 Result 欄位
- 若找不到/解析失敗 → Result = "N/A"，Notes 說明

Python 3.8+
（報表需 openpyxl）
"""
import argparse
import os
import re
from typing import Optional, List, Dict


# -----------------------------
# Excel 報表（專案風格）
# -----------------------------

def _sheet_name_for_model(model_ini_path: str) -> str:
    """
    以 model.ini 檔名的數字前綴決定 sheet 名：'PID_<N>'；無數字則 'others'
    例：'1_EU_xxx.ini' → 'PID_1'
    """
    base = os.path.basename(model_ini_path or "")
    m = re.match(r"^(\d+)_", base)
    if m:
        return f"PID_{int(m.group(1))}"
    return "others"


def _ensure_openpyxl():
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        raise SystemExit(
            "[ERROR] 需要 openpyxl 以輸出/附加報表。\n"
            "  安裝： pip install --user openpyxl\n"
        )


def export_report(res: Dict, xlsx_path: str = "kipling.xlsx", num_condition_cols: int = 5) -> None:
    """
    表頭固定：Rules | Result | condition_1..N
    - 不輸出 model.ini 欄位
    - 依 PID_N/others 分頁，若 xlsx 存在則附加一列
    - 欄位等寬、換行、垂直置頂
    """
    _ensure_openpyxl()
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    COMMON_WIDTH = 80
    COMMON_ALIGN = Alignment(wrap_text=True, vertical="top")
    BOLD = Font(bold=True)

    def _na(s: Optional[str]) -> str:
        s = (s or "").strip()
        return s if s else "N/A"

    sheet_name = _sheet_name_for_model(res.get("model_ini_path", ""))

    # 開啟或建立
    try:
        wb = load_workbook(xlsx_path)
    except Exception:
        wb = Workbook()

    # 取得/建立分頁
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)
        headers = ["Rules", "Result"] + [f"condition_{i}" for i in range(1, num_condition_cols + 1)]
        ws.append(headers)

    # 準備資料
    rules = "1) 讀取 model.ini → 2) 解析 isBassTrebleCustomValue"
    result = res.get("result_text") or "N/A"

    def j(items: Optional[List[str]]) -> str:
        return ", ".join(items or []) if items else ""

    conds = [
        f"isBassTrebleCustomValue(raw) = {_na(res.get('raw_value'))}",      # c1
        f"isBassTrebleCustomValue(norm) = {_na(res.get('normalized'))}",    # c2
        f"Found = {_na('YES' if res.get('found') else 'NO')}",              # c3
        f"Notes = {_na(res.get('notes'))}",                                 # c4
        f"Missing = {_na(j(res.get('missing')))}",                          # c5（預留）
    ][:num_condition_cols]

    ws.append([rules, result] + conds)
    last_row = ws.max_row

    # 樣式
    total_cols = 2 + num_condition_cols
    for c in range(1, total_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = COMMON_WIDTH
    for cell in ws[1]:  # header
        cell.alignment = COMMON_ALIGN
        cell.font = BOLD
    for cell in ws[last_row]:
        cell.alignment = COMMON_ALIGN

    # 移除預設 Sheet
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        try:
            wb.remove(wb["Sheet"])
        except Exception:
            pass

    wb.save(xlsx_path)


# -----------------------------
# 基礎解析
# -----------------------------

def _read_text(path: str) -> str:
    for enc in ("utf-8", "latin-1", "utf-16"):
        try:
            with open(path, "r", encoding=enc) as f:
                return f.read()
        except UnicodeDecodeError:
            continue
        except FileNotFoundError:
            raise
    with open(path, "r") as f:
        return f.read()


def _strip_comment(line: str) -> str:
    # 支援 '#' 或 ';' 註解
    line = line.split("#", 1)[0]
    line = line.split(";", 1)[0]
    return line.strip()


def _resolve_tvconfigs_path(root: str, tvconfigs_like: str) -> str:
    """
    /tvconfigs/... → <root>/...
    ./ 或 ../ 相對路徑 → 以 root 做基底
    其他絕對路徑（非 /tvconfigs）維持原樣
    其他純相對路徑 → root/相對
    （本腳本不一定用到，保留與專案一致的路徑規則）
    """
    if tvconfigs_like.startswith("/tvconfigs/"):
        rel = tvconfigs_like[len("/tvconfigs/"):]
        return os.path.normpath(os.path.join(root, rel))
    if tvconfigs_like.startswith("./") or tvconfigs_like.startswith("../"):
        return os.path.normpath(os.path.join(root, tvconfigs_like))
    if tvconfigs_like.startswith("/"):
        return tvconfigs_like
    return os.path.normpath(os.path.join(root, tvconfigs_like))


def _find_key_value_in_ini_text(text: str, key: str) -> Optional[str]:
    """
    搜尋 key = value（忽略註解與空白、大小寫不敏感、允許引號），回傳原始值字串（未去引號）。
    """
    key_re = re.compile(r'^\s*' + re.escape(key) + r'\s*=\s*"?([^"\n\r]+)"?\s*$', re.IGNORECASE)
    for raw in text.splitlines():
        line = _strip_comment(raw)
        if not line:
            continue
        m = key_re.match(line)
        if m:
            return m.group(1).strip()
    return None


def parse_model_ini_value(model_ini_path: str, key: str) -> Optional[str]:
    """
    從 model.ini 讀取 key 的值，回傳原字串；找不到回傳 None。
    """
    txt = _read_text(model_ini_path)
    return _find_key_value_in_ini_text(txt, key)


def normalize_value(val: Optional[str]) -> Optional[str]:
    """
    針對常見布林/數字做簡單正規化：true/false/1/0；其餘原樣回傳。
    """
    if val is None:
        return None
    v = val.strip().lower()
    if v in ("1", "true", "on", "yes"):
        return "true"
    if v in ("0", "false", "off", "no"):
        return "false"
    return val.strip()


def build_result(model_ini: str, raw_val: Optional[str]) -> Dict:
    notes: List[str] = []
    missing: List[str] = []

    found = raw_val is not None
    normalized = normalize_value(raw_val)

    result_text = normalized if normalized else "N/A"
    if not found:
        notes.append("model.ini 未找到 isBassTrebleCustomValue")

    return {
        "model_ini_path": model_ini,    # 僅用於 sheet 命名，不輸出
        "raw_value": raw_val or "",
        "normalized": normalized or "",
        "found": found,
        "result_text": result_text,
        "notes": "; ".join(notes),
        "missing": missing,
    }


# -----------------------------
# Main
# -----------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Read isBassTrebleCustomValue from model.ini and export to Excel."
    )
    parser.add_argument("--model-ini", required=True, help="path to model ini (e.g., model/1_xxx.ini)")
    parser.add_argument("--root", required=True, help="tvconfigs project root (for path mapping consistency)")
    parser.add_argument("-v", "--verbose", action="store_true", help="verbose logs")

    # 報表輸出
    parser.add_argument("--report", action="store_true", help="export to xlsx (default: kipling.xlsx)")
    parser.add_argument("--report-xlsx", metavar="FILE", help="export to specific xlsx file")

    args = parser.parse_args()

    model_ini = args.model_ini
    if not os.path.exists(model_ini):
        raise SystemExit(f"[ERROR] model ini not found: {model_ini}")
    root = os.path.abspath(os.path.normpath(args.root))  # 保留與專案一致（雖本腳本不必依賴）

    if args.verbose:
        print(f"[INFO] model_ini: {model_ini}")
        print(f"[INFO] root     : {root}")

    # 讀 isBassTrebleCustomValue
    raw_val = parse_model_ini_value(model_ini, "isBassTrebleCustomValue")
    if args.verbose:
        print(f"[INFO] isBassTrebleCustomValue = {raw_val if raw_val is not None else '(not found)'}")

    # 組裝結果 + 輸出 console
    res = build_result(model_ini, raw_val)
    print(f"Result(isBassTrebleCustomValue): {res['result_text']}")
    if res.get("notes"):
        print(f"Notes   : {res['notes']}")

    # Excel
    if args.report or args.report_xlsx:
        xlsx_path = args.report_xlsx if args.report_xlsx else "kipling.xlsx"
        export_report(res, xlsx_path=xlsx_path)
        sheet = _sheet_name_for_model(model_ini)
        print(f"[INFO] Report appended to: {xlsx_path} (sheet: {sheet})")


if __name__ == "__main__":
    main()

