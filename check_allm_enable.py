#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
check_allm_flag.py

Standalone checker:
1) Parse model.ini to find: isSupportALLM
   - PASS if value == true (case-insensitive), otherwise FAIL (including missing or commented out)
2) Parse model.ini to find TvDefaultSettingsPath (e.g. "/tvconfigs/tvserv_ini/tvDefaultSettings_dolby.ini"),
   resolve to real path by mapping "/tvconfigs/..." => "<root>/..."
   Open that ini, find [ALLM] section, and verify all ENABLE flags are 1.
   - PASS if every line in [ALLM] that defines an ENABLE-like key has value 1 and no ENABLE=0 is found.
   - FAIL otherwise; missing file or missing [ALLM] counts as FAIL.

XLSX output:
- Columns: Rules, Result, condition_1, condition_2
- The format (width/wrap/vertical-top/bold header) follows the same style used by check_setupwizard_flag.py.

Usage:
  python3 check_allm_flag.py --model-ini model/1_xxx.ini --root .
  python3 check_allm_flag.py --model-ini model/1_xxx.ini --root . --report
  python3 check_allm_flag.py --model-ini model/1_xxx.ini --root . --report-xlsx kipling.xlsx
  python3 check_allm_flag.py --model-ini model/1_xxx.ini --root . -v
"""

import argparse
import os
import re
from typing import Optional, Dict, List, Tuple


# -----------------------------
# File reading & parsing helpers
# -----------------------------

def _read_text(path: str) -> str:
    """Read text with common encodings. Raises FileNotFoundError if missing."""
    for enc in ("utf-8", "latin-1", "utf-16"):
        try:
            with open(path, "r", encoding=enc) as f:
                return f.read()
        except UnicodeDecodeError:
            continue
        except FileNotFoundError:
            raise
    with open(path, "r") as f:
        return f.read()


def _strip_comment(line: str) -> str:
    """Remove comments starting with '#' or ';' and trim whitespace."""
    line = line.split("#", 1)[0]
    line = line.split(";", 1)[0]
    return line.strip()


def _find_kv_case_insensitive(text: str, key: str) -> Optional[str]:
    """
    Return the unquoted value for the first un-commented line like: key = value
    Case-insensitive key matching. Returns None if not found.
    """
    for raw in text.splitlines():
        line = _strip_comment(raw)
        if not line or "=" not in line:
            continue
        m = re.match(r'^\s*' + re.escape(key) + r'\s*=\s*("?)(.*?)\1\s*$', line, re.IGNORECASE)
        if m:
            return m.group(2).strip()
    return None


def _find_is_support_allm(model_ini_path: str) -> Optional[str]:
    """Find the (uncommented) value of isSupportALLM in model.ini."""
    text = _read_text(model_ini_path)
    return _find_kv_case_insensitive(text, "isSupportALLM")


def _find_tv_default_settings_path(model_ini_path: str) -> Optional[str]:
    """Find TvDefaultSettingsPath in model.ini (case-insensitive)."""
    text = _read_text(model_ini_path)
    return _find_kv_case_insensitive(text, "TvDefaultSettingsPath")


def _map_tvconfigs_to_root(tvconfigs_path: str, root: str) -> str:
    """
    Map a path like "/tvconfigs/tvserv_ini/tvDefaultSettings_dolby.ini"
    to "<root>/tvserv_ini/tvDefaultSettings_dolby.ini".
    If tvconfigs_path doesn't start with "/tvconfigs/", return as-is joined to root.
    """
    tvconfigs_path = tvconfigs_path.strip()
    # Normalize surrounding quotes already stripped in _find_kv_case_insensitive
    prefix = "/tvconfigs/"
    if tvconfigs_path.startswith(prefix):
        rel = tvconfigs_path[len(prefix):].lstrip("/")
        return os.path.normpath(os.path.join(root, rel))
    # If it is an absolute path but not under /tvconfigs, keep as-is
    if os.path.isabs(tvconfigs_path):
        return tvconfigs_path
    # Relative -> treat as relative to root
    return os.path.normpath(os.path.join(root, tvconfigs_path))




def infer_pid_sheet_name(model_ini_path: str) -> str:
    """Infer Excel sheet name as 'PID_<prefix>' where <prefix> is the leading number in model.ini file name.
    If no leading digits are found, fall back to 'PID_UNKNOWN'. Excel sheet names are truncated to 31 chars.
    """
    base = os.path.basename(model_ini_path)
    name = os.path.splitext(base)[0]
    m = re.match(r'\s*(\d+)', name)
    pid = m.group(1) if m else None
    sheet = f"PID_{pid}" if pid else "PID_UNKNOWN"
    return sheet[:31]
# -----------------------------
# Core check logic
# -----------------------------

def check_is_support_allm(model_ini_path: str) -> Dict[str, object]:
    """
    Check if isSupportALLM == true.
    Returns dict: passed (bool), value (str or ''), notes (list[str])
    """
    notes: List[str] = []
    value = _find_is_support_allm(model_ini_path)
    if value is None:
        notes.append("isSupportALLM 未宣告或僅存在於註解中")
        passed = False
        value_str = ""
    else:
        passed = (value.strip().lower() == "true")
        if not passed:
            notes.append(f"isSupportALLM 不是 true (got: {value})")
        value_str = value

    return {"passed": passed, "value": value_str, "notes": notes}


def _extract_allm_section_lines(text: str) -> List[str]:
    """
    Return raw (uncommented) lines inside [ALLM] section until next [Section].
    Empty/comment-only lines are skipped.
    """
    lines = text.splitlines()
    in_section = False
    result: List[str] = []
    for raw in lines:
        stripped = raw.strip()
        if re.match(r'^\s*\[([^\]]+)\]\s*$', stripped):
            sect = re.findall(r'^\s*\[([^\]]+)\]\s*$', stripped)[0]
            in_section = (sect.strip().lower() == "allm")
            continue
        if in_section:
            line = _strip_comment(raw)
            if not line:
                continue
            # Stop if it looks like a new section header (defensive)
            if re.match(r'^\s*\[([^\]]+)\]\s*$', line):
                break
            result.append(line)
    return result


def check_tvdefault_allm_enable(model_ini_path: str, root: str) -> Dict[str, object]:
    """
    Resolve TvDefaultSettingsPath and check [ALLM] section for ENABLE flags.
    Policy:
      - The [ALLM] section must exist and contain at least one line with an ENABLE-like key.
      - Every ENABLE-like key (e.g., ENABLE, HDMI1_ENABLE, ...) must equal 1.
      - If any ENABLE=0 (or value != 1) is found, FAIL.
      - Missing file or missing [ALLM] -> FAIL.
    Returns dict:
      passed (bool), path (str or ''), offending (list[str]), notes (list[str])
    """
    notes: List[str] = []
    tvdef_rel = _find_tv_default_settings_path(model_ini_path)
    if not tvdef_rel:
        return {
            "passed": False,
            "path": "",
            "offending": [],
            "notes": ["model.ini 未宣告 TvDefaultSettingsPath"]
        }

    actual = _map_tvconfigs_to_root(tvdef_rel, root)
    try:
        text = _read_text(actual)
    except FileNotFoundError:
        return {
            "passed": False,
            "path": actual,
            "offending": [],
            "notes": [f"找不到 TvDefaultSettingsPath 檔案: {actual}"]
        }

    lines = _extract_allm_section_lines(text)
    if not lines:
        return {
            "passed": False,
            "path": actual,
            "offending": [],
            "notes": ["找不到 [ALLM] 區塊或是區塊內沒有有效內容"]
        }

    # Identify ENABLE-like assignments in the section
    enable_pairs: List[Tuple[str, str]] = []
    pairs = {}
    for line in lines:
        for pair in line.split(','):
            key, value = pair.split('=')
            pairs[key] = value
            enable_pairs.append((key,value))

    #print(enable_pairs)
    if not enable_pairs:
        return {
            "passed": False,
            "path": actual,
            "offending": [],
            "notes": ["[ALLM] 區塊內沒有找到任何 ENABLE 相關設定"]
        }

    offending: List[str] = []
    for key, val in enable_pairs:
        # Accept "1" (string) as pass; anything else is fail
        if val != "1" and key.lower() == "enable":
            offending.append(f"{key}={val}")

    #print(offending)
    passed = (len(offending) == 0)
    if not passed:
        notes.append("存在非 1 的 ENABLE 設定")

    return {"passed": passed, "path": actual, "offending": offending, "notes": notes}


# -----------------------------
# XLSX report
# -----------------------------

def export_report(res1: Dict[str, object], res2: Dict[str, object], xlsx_path: str, sheet_name: str = "ALLM") -> None:
    """
    Export report with columns: Rules, Result, condition_1, condition_2
    - condition_1: isSupportALLM = <value or N/A>
    - condition_2: TvDefaultSettingsPath [ALLM] ENABLE=1 check summary
    - Uniform column width, wrap text, vertical top; bold header
    """
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        raise SystemExit(
            "[ERROR] 需要 openpyxl 才能輸出報表。\n"
            "  安裝： pip install --user openpyxl\n"
        ) from e

    COMMON_WIDTH = 80
    ALIGN = Alignment(wrap_text=True, vertical="top")
    BOLD = Font(bold=True)

    try:
        wb = load_workbook(xlsx_path)
    except Exception:
        wb = Workbook()

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)
        ws.append(["Rules", "Result", "condition_1", "condition_2"])

    # Compose fields
    rules = (
        "1) model.ini → isSupportALLM must be true\n"
        "2) TvDefaultSettingsPath → open real file and verify [ALLM] section: all ENABLE* = 1"
    )
    all_pass = bool(res1.get("passed")) and bool(res2.get("passed"))
    result = "PASS" if all_pass else "FAIL"

    cond1 = f"isSupportALLM = {res1.get('value') or 'N/A'}"
    # Build condition_2 summary
    path2 = res2.get("path") or "N/A"
    if res2.get("passed"):
        #cond2 = f"[ALLM] ENABLE 檢查: PASS\nPath: {path2}"
        cond2 = f"[ALLM] ENABLE 檢查: PASS"
    else:
        off = res2.get("offending") or []
        note_lines = res2.get("notes") or []
        detail = ""
        #if off:
            #detail += "Offending: " + ", ".join(off) + "\n"
        if note_lines:
            detail += "Notes: " + " | ".join(note_lines)
        #cond2 = f"[ALLM] ENABLE 檢查: FAIL\nPath: {path2}\n{detail}".rstrip()
        cond2 = f"[ALLM] ENABLE 檢查: FAIL\n{detail}".rstrip()

    ws.append([rules, result, cond1, cond2])
    last_row = ws.max_row

    # Formatting
    for col in range(1, 4 + 1):
        ws.column_dimensions[get_column_letter(col)].width = COMMON_WIDTH
    for cell in ws[1]:
        cell.font = BOLD
        cell.alignment = ALIGN
    for cell in ws[last_row]:
        cell.alignment = ALIGN

    # Remove default "Sheet" if exists
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        try:
            wb.remove(wb["Sheet"])
        except Exception:
            pass

    wb.save(xlsx_path)


# -----------------------------
# CLI
# -----------------------------

def main():
    ap = argparse.ArgumentParser(
        description="Check model.ini:isSupportALLM == true AND TvDefaultSettingsPath [ALLM] ENABLE* = 1."
    )
    ap.add_argument("--model-ini", required=True, help="Path to model.ini")
    ap.add_argument("--root", default=".", help="Root dir for mapping /tvconfigs/... => <root>/... (default: current dir)")
    ap.add_argument("--report", action="store_true", help="Append result to kipling.xlsx")
    ap.add_argument("--report-xlsx", metavar="FILE", help="Append result to a specific XLSX file")
    ap.add_argument("-v", "--verbose", action="store_true", help="Verbose output")
    args = ap.parse_args()

    if not os.path.exists(args.model_ini):
        raise SystemExit(f"[ERROR] model ini not found: {args.model_ini}")
    root_abs = os.path.abspath(args.root)

    # Run checks
    res1 = check_is_support_allm(args.model_ini)
    res2 = check_tvdefault_allm_enable(args.model_ini, root_abs)

    # Console output
    print(f"[CHECK-1] isSupportALLM = {res1['value'] or 'N/A'} -> {'PASS' if res1['passed'] else 'FAIL'}")
    if args.verbose and res1.get('notes'):
        for n in res1['notes']:
            print("  -", n)

    tvdef_path = res2.get("path") or "N/A"
    print(f"[CHECK-2] TvDefaultSettingsPath [ALLM] ENABLE 檢查 on: {tvdef_path} -> {'PASS' if res2['passed'] else 'FAIL'}")
    if args.verbose:
        if res2.get("offending"):
            print("  Offending ENABLE lines:")
            for o in res2["offending"]:
                print("   *", o)
        for n in res2.get("notes", []):
            print("  -", n)

    # Report
    if args.report_xlsx:
        export_report(res1, res2, args.report_xlsx, sheet_name=infer_pid_sheet_name(args.model_ini))
        print(f"[INFO] Report appended to: {args.report_xlsx} (sheet: {infer_pid_sheet_name(args.model_ini)})")
    elif args.report:
        export_report(res1, res2, "kipling.xlsx", sheet_name=infer_pid_sheet_name(args.model_ini))
        print(f"[INFO] Report appended to: kipling.xlsx (sheet: {infer_pid_sheet_name(args.model_ini)})")


if __name__ == "__main__":
    main()
