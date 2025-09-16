#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
check_TvDefaultSettings_PCMODE_sources.py

需求：
- 參考 check_CI.py 的「/tvconfigs → --root」路徑映射、xlsx 輸出格式、sheet 命名規則
- 在 xlsx 內移除 model.ini 欄位（僅用於決定分頁名，不輸出）
- 從 model.ini 讀取 TvDefaultSettingsPath，打開該檔案
- 逐行尋找包含「PCMODE=AUTO」的行（可能不止一行）
- 對每一行，解析並打印同一行的「source=」值（大小寫不敏感）
- 終端列印與 xlsx 報表都要呈現結果；若未找到則 Result = "N/A"

Python 3.8+
（報表需 openpyxl）
"""
import argparse
import os
import re
from typing import Optional, List, Dict, Tuple


# -----------------------------
# Excel 報表（沿用專案風格）
# -----------------------------

def _sheet_name_for_model(model_ini_path: str) -> str:
    """以 model.ini 檔名的數字前綴決定 sheet 名：'PID_<N>'；無數字則 'others'。"""
    base = os.path.basename(model_ini_path or "")
    m = re.match(r"^(\d+)_", base)
    if m:
        return f"PID_{int(m.group(1))}"
    return "others"


def _ensure_openpyxl():
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        raise SystemExit(
            "[ERROR] 需要 openpyxl 以輸出/附加報表。\n"
            "  安裝： pip install --user openpyxl\n"
        )


def export_report(res: Dict, xlsx_path: str = "kipling.xlsx", num_condition_cols: int = 5) -> None:
    """
    表頭固定：Rules | Result | condition_1..N
    - 不輸出 model.ini 欄位
    - 依 PID_N/others 分頁，若 xlsx 存在則附加一列
    """
    _ensure_openpyxl()
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    COMMON_WIDTH = 80
    COMMON_ALIGN = Alignment(wrap_text=True, vertical="top")
    BOLD = Font(bold=True)

    def _na(s: Optional[str]) -> str:
        s = (s or "").strip()
        return s if s else "N/A"

    sheet_name = _sheet_name_for_model(res.get("model_ini_path", ""))

    try:
        wb = load_workbook(xlsx_path)
    except Exception:
        wb = Workbook()

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)
        headers = ["Rules", "Result"] + [f"condition_{i}" for i in range(1, num_condition_cols + 1)]
        ws.append(headers)

    rules = "1) 解析 TvDefaultSettingsPath → 2) 開檔 → 3) 尋找行含 PCMODE=AUTO → 4) 取同一行 source"
    result = res.get("result_text") or "N/A"

    def j(items: Optional[List[str]]) -> str:
        return ", ".join(items or []) if items else ""

    conds = [
        f"TvDefaultSettingsPath = {_na(res.get('default_settings_path_resolved'))}",  # c1
        f"Matches = {_na(str(res.get('match_count')))}",                               # c2
        f"Sources = {_na(j(res.get('sources_unique')))}",                              # c3
        f"Notes = {_na(res.get('notes'))}",                                           # c4
        f"Missing = {_na(j(res.get('missing')))}",                                    # c5
    ][:num_condition_cols]

    ws.append([rules, result] + conds)
    last_row = ws.max_row

    total_cols = 2 + num_condition_cols
    for c in range(1, total_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = COMMON_WIDTH
    for cell in ws[1]:  # header
        cell.alignment = COMMON_ALIGN
        cell.font = BOLD
    for cell in ws[last_row]:
        cell.alignment = COMMON_ALIGN

    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        try:
            wb.remove(wb["Sheet"])
        except Exception:
            pass

    wb.save(xlsx_path)


# -----------------------------
# 基礎解析工具
# -----------------------------

def _read_text(path: str) -> str:
    for enc in ("utf-8", "latin-1", "utf-16"):
        try:
            with open(path, "r", encoding=enc) as f:
                return f.read()
        except UnicodeDecodeError:
            continue
        except FileNotFoundError:
            raise
    with open(path, "r") as f:
        return f.read()


def _strip_comment(line: str) -> str:
    # 支援 '#' 或 ';' 註解
    line = line.split("#", 1)[0]
    line = line.split(";", 1)[0]
    return line.strip()


def _resolve_tvconfigs_path(root: str, tvconfigs_like: str) -> str:
    """
    /tvconfigs/... → <root>/...
    ./ 或 ../ 相對路徑 → 以 root 為基底
    其他絕對路徑（非 /tvconfigs）維持原樣
    其他純相對路徑 → root/相對
    """
    if tvconfigs_like.startswith("/tvconfigs/"):
        rel = tvconfigs_like[len("/tvconfigs/"):]
        return os.path.normpath(os.path.join(root, rel))
    if tvconfigs_like.startswith("./") or tvconfigs_like.startswith("../"):
        return os.path.normpath(os.path.join(root, tvconfigs_like))
    if tvconfigs_like.startswith("/"):
        return tvconfigs_like
    return os.path.normpath(os.path.join(root, tvconfigs_like))


def _find_key_value_in_ini_text(text: str, key: str) -> Optional[str]:
    """在整份 INI 文字中找單一 key=value（大小寫不敏感），回傳 value（未去引號）。"""
    key_re = re.compile(r'^\s*' + re.escape(key) + r'\s*=\s*"?([^"\n\r]+)"?\s*$', re.IGNORECASE)
    for raw in text.splitlines():
        line = _strip_comment(raw)
        if not line:
            continue
        m = key_re.match(line)
        if m:
            return m.group(1).strip()
    return None


def parse_model_ini_for_default_settings(model_ini_path: str, root: str) -> Optional[str]:
    """從 model.ini 找 TvDefaultSettingsPath，並解析為實體路徑。"""
    txt = _read_text(model_ini_path)
    val = _find_key_value_in_ini_text(txt, "TvDefaultSettingsPath")
    if val is None:
        return None
    return _resolve_tvconfigs_path(root, val)


# -----------------------------
# 專案需求：找 PCMODE=AUTO 並取同行 source
# -----------------------------

PCMODE_TOKEN_RE = re.compile(r'(?i)\bPCMODE\s*=\s*AUTO\b')

def _kv_pairs_from_line(line_no_comment: str) -> dict:
    """
    從單行抽取所有 key=value 配對（大小寫不敏感）。
    值允許引號或無引號；以空白、逗號、Tab 分隔的多組配對皆可抓取。
    修正：避免 re.findall 產生 3 群組導致 unpack 失敗。
    """
    pairs = {}
    # key=「"值"」或 key=值（直到逗號/空白/#/;，避免吃到後面內容）
    pattern = re.compile(
        r'(?i)\b([A-Z0-9_]+)\s*=\s*(?:"([^"]+)"|([^,\s\t#;]+))'
    )
    for m in pattern.finditer(line_no_comment):
        key = m.group(1).lower()
        # 兩個互斥群組其一會是 None
        val = m.group(2) if m.group(2) is not None else m.group(3)
        pairs[key] = (val or "").strip()
    return pairs

def extract_pcmode_auto_sources(default_settings_path: str) -> Tuple[List[str], int, List[str]]:
    """
    讀 default settings 檔逐行：
    - 找到包含「PCMODE=AUTO」的行
    - 從同一行擷取 source= 的值（大小寫不敏感）
    回傳：sources（可重複，保持出現順序）、match_count、raw_matched_lines（去註解）
    """
    sources: List[str] = []
    raw_lines: List[str] = []
    count = 0

    if not default_settings_path or not os.path.exists(default_settings_path):
        return sources, count, raw_lines

    with open(default_settings_path, "r", encoding="utf-8", errors="ignore") as f:
        for raw in f:
            no_comment = _strip_comment(raw)
            if not no_comment:
                continue
            if PCMODE_TOKEN_RE.search(no_comment):
                count += 1
                raw_lines.append(no_comment)
                kv = _kv_pairs_from_line(no_comment)
                # 主要抓 source=，若無再嘗試 src=
                src = kv.get("source") or kv.get("src")
                if src:
                    sources.append(src)

    return sources, count, raw_lines


def build_result(model_ini: str,
                 default_path: Optional[str],
                 sources: List[str],
                 match_count: int,
                 raw_lines: List[str]) -> Dict:
    missing: List[str] = []
    notes: List[str] = []

    if default_path is None:
        notes.append("model.ini 未找到 TvDefaultSettingsPath")
    elif not os.path.exists(default_path):
        missing.append(default_path)

    # 去重但保留順序
    seen = set()
    uniq_sources: List[str] = []
    for s in sources:
        if s not in seen:
            seen.add(s)
            uniq_sources.append(s)

    result_text = ", ".join(uniq_sources) if uniq_sources else "N/A"

    # 補充：若有命中但沒有任何 source= 值
    if match_count > 0 and not uniq_sources:
        notes.append("找到 PCMODE=AUTO 行，但同行未見 source= 鍵值")

    return {
        "model_ini_path": model_ini,                        # 僅用於決定 sheet 名，不輸出
        "default_settings_path_resolved": default_path or "",
        "match_count": match_count,
        "sources_unique": uniq_sources,
        "raw_matched_lines": raw_lines,
        "result_text": result_text,
        "notes": "; ".join(notes),
        "missing": missing,
    }


# -----------------------------
# Main
# -----------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Read TvDefaultSettingsPath → find lines with PCMODE=AUTO and print same-line source= values."
    )
    parser.add_argument("--model-ini", required=True, help="path to model ini (e.g., model/1_xxx.ini)")
    parser.add_argument("--root", required=True, help="tvconfigs project root (maps /tvconfigs/* to here)")
    parser.add_argument("-v", "--verbose", action="store_true", help="verbose logs")

    # 報表輸出
    parser.add_argument("--report", action="store_true", help="export to xlsx (default: kipling.xlsx)")
    parser.add_argument("--report-xlsx", metavar="FILE", help="export to specific xlsx file")

    args = parser.parse_args()

    model_ini = args.model_ini
    if not os.path.exists(model_ini):
        raise SystemExit(f"[ERROR] model ini not found: {model_ini}")
    root = os.path.abspath(os.path.normpath(args.root))

    if args.verbose:
        print(f"[INFO] model_ini: {model_ini}")
        print(f"[INFO] root     : {root}")

    # 解析 TvDefaultSettingsPath
    default_path = parse_model_ini_for_default_settings(model_ini, root)
    if args.verbose:
        print(f"[INFO] TvDefaultSettingsPath → {default_path if default_path else '(not found)'}")

    # 抓 PCMODE=AUTO 行與 source=
    sources, match_count, raw_lines = extract_pcmode_auto_sources(default_path) if default_path else ([], 0, [])
    if args.verbose:
        print(f"[INFO] PCMODE=AUTO matches = {match_count}")
        if raw_lines:
            for i, ln in enumerate(raw_lines[:10], 1):  # 最多預覽 10 行
                print(f"[DBG] line#{i}: {ln}")

    # 組裝結果 + 終端輸出
    res = build_result(model_ini, default_path, sources, match_count, raw_lines)

    if res["result_text"] != "N/A":
        print("Result(source values):", res["result_text"])
    else:
        print("Result(source values): N/A")
    print("Matches:", res["match_count"])
    if res.get("notes"):
        print("Notes  :", res["notes"])
    if res.get("missing"):
        print("Missing:", ", ".join(res["missing"]))

    # Excel
    if args.report or args.report_xlsx:
        xlsx_path = args.report_xlsx if args.report_xlsx else "kipling.xlsx"
        export_report(res, xlsx_path=xlsx_path)
        sheet = _sheet_name_for_model(model_ini)
        print(f"[INFO] Report appended to: {xlsx_path} (sheet: {sheet})")


if __name__ == "__main__":
    main()

