#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
check_logo_path.py
- Parse a model.ini and verify the files pointed by PowerLogoPath and BrandLogoPath exist.
- Report in console and (optionally) append/produce an Excel report with standardized headers.

Report columns (aligned with tv_multi_standard_validation format):
  Rules, Result, condition_1, condition_2, ...

Notes:
- Paths in model.ini usually look like "/tvconfigs/logo/bootfile.image".
  We resolve them against --root by stripping the leading "/" so it maps to <root>/tvconfigs/logo/bootfile.image.
- Result is strictly "PASS" or "FAIL".
- We DO NOT include model_ini nor section columns per user's reporting convention.
"""

import argparse
import re
from pathlib import Path
from typing import Optional, Dict, List, Tuple
from datetime import datetime
import sys

# Excel dependencies are optional. Only required if --report-xlsx is provided.
try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
except Exception:
    openpyxl = None

ENCODINGS = ("utf-8", "utf-16", "utf-16-le", "utf-16-be", "latin-1")

POWER_KEY = "PowerLogoPath"
BRAND_KEY = "BrandLogoPath"

HEADER = ["Rules", "Result", "condition_1", "condition_2"]


def smart_read_text(path: Path) -> str:
    last_err = None
    for enc in ENCODINGS:
        try:
            return path.read_text(encoding=enc, errors="strict")
        except Exception as e:
            last_err = e
            continue
    # Last fallback: permissive read to avoid total failure
    try:
        return path.read_text(encoding="utf-8", errors="ignore")
    except Exception:
        if last_err:
            raise last_err
        raise


def extract_ini_path(content: str, key: str) -> Optional[str]:
    """
    Extract a double-quoted or non-quoted path value for a given key.
    Supports lines like: Key="..."; Key = "/path/file.ext" ; comments after value are ignored.
    """
    # Try quoted first
    pattern_q = re.compile(rf'^\s*{re.escape(key)}\s*=\s*"([^"]+)"', re.IGNORECASE | re.MULTILINE)
    m = pattern_q.search(content)
    if m:
        return m.group(1).strip()

    # Then unquoted until comment or EOL
    pattern_u = re.compile(rf'^\s*{re.escape(key)}\s*=\s*([^\s#;]+)', re.IGNORECASE | re.MULTILINE)
    m = pattern_u.search(content)
    if m:
        return m.group(1).strip()

    return None


def _map_to_tvconfigs_rel(ini_path: Optional[str]) -> Optional[str]:
    if not ini_path:
        return None
    s = ini_path.strip().strip('"').strip("'").replace('\\', '/')
    if s.startswith("/mnt/vendor/tvconfigs/"):
        return s[len("/mnt/vendor/tvconfigs/"):]
    if s.startswith("/tvconfigs/"):
        return s[len("/tvconfigs/"):]
    if s.startswith("tvconfigs/"):
        return s[len("tvconfigs/"):]
    # Fallback: trim leading slash to create a relative-looking path
    if s.startswith("/"):
        return s.lstrip("/")
    return s

def resolve_tvconfigs_path(root: Path, ini_path: Optional[str]) -> Tuple[Optional[Path], Optional[str]]:
    """
    Return (absolute_path, relative_display_path).
    Relative path is the path after stripping the tvconfigs anchor, e.g.:
      /tvconfigs/logo/bootfile.image  ->  logo/bootfile.image
      tvconfigs/logo/bootfile.image   ->  logo/bootfile.image
      /mnt/vendor/tvconfigs/logo/...  ->  logo/...
    """
    rel = _map_to_tvconfigs_rel(ini_path)
    if not rel:
        return None, None
    abs_path = (root / rel).resolve()
    return abs_path, rel
def check_exists(p: Optional[Path]) -> bool:
    return bool(p and p.exists() and p.is_file())


def detect_sheet_name_from_model(model_ini: Path) -> str:
    """
    Sheet naming rule (aligned with tv_multi_standard_validation.py):
    - If filename starts with N_ and 1 <= N <= 20 then "PID_N"
    - Otherwise, "others"
    """
    m = re.match(r'^(\d{1,3})_', model_ini.name)
    if m:
        n = int(m.group(1))
        #if 1 <= n <= 20:
        return f"PID_{n}"
    return "others"
def format_console_row(rules: str, result: str, c1: str, c2: str) -> str:
    return f"{rules:16} | {result:4} | {c1} | {c2}"


def ensure_openpyxl():
    if openpyxl is None:
        raise RuntimeError("openpyxl is required for --report-xlsx but is not installed in this environment.")


def write_to_excel(xlsx_path: Path, sheet_name: str, rows: List[List[str]]):
    ensure_openpyxl()
    if xlsx_path.exists():
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
    else:
        wb = openpyxl.Workbook()
        # Use active if empty, but rename to sheet_name if it's the default empty
        ws = wb.active
        ws.title = sheet_name

    # If the sheet appears empty, write header
    if ws.max_row == 1 and ws.max_column == 1 and ws["A1"].value is None:
        ws.append(HEADER)

    for r in rows:
        ws.append(r)

    # Styling: set column widths, wrap, vertical top, bold header
    col_widths = {1: 18, 2: 8, 3: 60, 4: 60}
    for idx, width in col_widths.items():
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width

    align_wrap_top = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = align_wrap_top

    # Bold header
    for cell in ws[1]:
        cell.font = Font(bold=True)

    wb.save(xlsx_path)


def main():
    ap = argparse.ArgumentParser(description="Check existence of PowerLogoPath and BrandLogoPath from a model.ini")
    ap.add_argument("--model-ini", required=True, help="Path to model.ini to inspect")
    ap.add_argument("--root", default=".", help="Project root where tvconfigs folder resides (default: current dir)")
    ap.add_argument("--report", default=None, help="Append results to this Excel file (creates if missing)")
    ap.add_argument("--report-xlsx", default=None, help=argparse.SUPPRESS)
    ap.add_argument("-v", "--verbose", action="store_true", help="Verbose logs")

    args = ap.parse_args()

    model_ini = Path(args.model_ini).resolve()
    root = Path(args.root).resolve()

    if not model_ini.exists():
        print(f"[ERROR] model.ini not found: {model_ini}", file=sys.stderr)
        sys.exit(2)

    try:
        content = smart_read_text(model_ini)
    except Exception as e:
        print(f"[ERROR] Failed to read model.ini with known encodings: {e}", file=sys.stderr)
        sys.exit(3)

    power_path_raw = extract_ini_path(content, POWER_KEY)
    brand_path_raw = extract_ini_path(content, BRAND_KEY)

    power_abs, power_rel = resolve_tvconfigs_path(root, power_path_raw)
    brand_abs, brand_rel = resolve_tvconfigs_path(root, brand_path_raw)

    power_ok = check_exists(power_abs)
    brand_ok = check_exists(brand_abs)

    # Build conditions strings
    c1 = f'{POWER_KEY}={power_path_raw or "N/A"} -> {power_rel or "N/A"} ' + ("✅" if power_ok else "❌")
    c2 = f'{BRAND_KEY}={brand_path_raw or "N/A"} -> {brand_rel or "N/A"} ' + ("✅" if brand_ok else "❌")

    result = "PASS" if (power_ok and brand_ok) else "FAIL"
    rules = "Logo file check"
    row = [rules, result, c1, c2]

    # Console output
    print("Rules           | Res. | condition_1 | condition_2")
    print("-" * 120)
    print(format_console_row(rules, result, c1, c2))

    # Excel output if requested
    report_target = args.report or args.report_xlsx
    if report_target:
        xlsx_path = Path(report_target).resolve()
        sheet = detect_sheet_name_from_model(model_ini)
        try:
            write_to_excel(xlsx_path, sheet, [row])
            if args.verbose:
                print(f"[INFO] Report appended to {xlsx_path} (sheet: {sheet})")
        except Exception as e:
            print(f"[ERROR] Failed to write Excel report: {e}", file=sys.stderr)
            sys.exit(4)


if __name__ == "__main__":
    main()
