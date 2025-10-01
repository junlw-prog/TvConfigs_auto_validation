#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
check_ci_version_from_default_settings.py

需求：
- 參考 tv_multi_standard_validation.py 的「/tvconfigs → --root」路徑映射、xlsx 輸出格式、sheet 命名規則
- 在 xlsx 內「移除 model.ini 欄位」（僅用於決定分頁名，不輸出）
- 從 model.ini 取得 TvDefaultSettingsPath，打開該檔案找參數 SUPPORTCI20：
    * SUPPORTCI20 = true  → Result 欄位為 "CI 2.0"
    * SUPPORTCI20 = false → Result 欄位為 "CI 1.4.4"
- 若找不到鍵值/檔案不存在/解析失敗 → Result = "N/A"，Notes/Missing 說明

Python 3.8+
（報表需 openpyxl）
"""
import argparse
import os
import re
from typing import Optional, List, Dict


# -----------------------------
# Excel 報表（沿用專案風格）
# -----------------------------

def _sheet_name_for_model(model_ini_path: str) -> str:
    """
    以 model.ini 檔名的數字前綴決定 sheet 名：'PID_<N>'；無數字則 'others'
    例：'1_EU_xxx.ini' → 'PID_1'
    """
    base = os.path.basename(model_ini_path or "")
    m = re.match(r"^(\d+)_", base)
    if m:
        return f"PID_{int(m.group(1))}"
    return "others"


def _ensure_openpyxl():
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        raise SystemExit(
            "[ERROR] 需要 openpyxl 以輸出/附加報表。\n"
            "  安裝： pip install --user openpyxl\n"
        )


def export_report(res: Dict, xlsx_path: str = "kipling.xlsx", num_condition_cols: int = 5) -> None:
    """
    表頭固定：Rules | Result | condition_1..N
    - 不輸出 model.ini 欄位
    - 依 PID_N/others 分頁，若 xlsx 存在則附加一列
    - 欄位等寬、換行、垂直置頂
    """
    _ensure_openpyxl()
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    COMMON_WIDTH = 80
    COMMON_ALIGN = Alignment(wrap_text=True, vertical="top")
    BOLD = Font(bold=True)

    def _na(s: Optional[str]) -> str:
        s = (s or "").strip()
        return s if s else "N/A"

    sheet_name = _sheet_name_for_model(res.get("model_ini_path", ""))

    # 開啟或建立
    try:
        wb = load_workbook(xlsx_path)
    except Exception:
        wb = Workbook()

    # 取得/建立分頁
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)
        headers = ["Rules", "Result"] + [f"condition_{i}" for i in range(1, num_condition_cols + 1)]
        ws.append(headers)

    # 準備資料
    rules = "1) 解析 TvDefaultSettingsPath → 2) 開檔 → 3) 讀 SUPPORTCI20 → 決定 CI 版本"
    result = res.get("result_text") or "N/A"

    def j(items: Optional[List[str]]) -> str:
        return ", ".join(items or []) if items else ""

    conds = [
        f"TvDefaultSettingsPath = {_na(res.get('default_settings_path_resolved'))}",  # c1
        f"SUPPORTCI20 = {_na(res.get('supportci20_text'))}",                          # c2
        f"Decision = {_na(res.get('decision'))}",                                     # c3
        f"Notes = {_na(res.get('notes'))}",                                           # c4
        f"Missing = {_na(j(res.get('missing')))}",                                    # c5
    ][:num_condition_cols]

    ws.append([rules, result] + conds)
    last_row = ws.max_row

    # 樣式
    total_cols = 2 + num_condition_cols
    for c in range(1, total_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = COMMON_WIDTH
    for cell in ws[1]:  # header
        cell.font = BOLD
        cell.alignment = COMMON_ALIGN
    for cell in ws[last_row]:
        cell.alignment = COMMON_ALIGN

    # 移除預設 Sheet
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        try:
            wb.remove(wb["Sheet"])
        except Exception:
            pass

    wb.save(xlsx_path)


# -----------------------------
# 基礎解析
# -----------------------------

def _read_text(path: str) -> str:
    for enc in ("utf-8", "latin-1", "utf-16"):
        try:
            with open(path, "r", encoding=enc) as f:
                return f.read()
        except UnicodeDecodeError:
            continue
        except FileNotFoundError:
            raise
    with open(path, "r") as f:
        return f.read()


def _strip_comment(line: str) -> str:
    line = line.split("#", 1)[0]
    line = line.split(";", 1)[0]
    return line.strip()


def _resolve_tvconfigs_path(root: str, tvconfigs_like: str) -> str:
    """
    /tvconfigs/... → <root>/...
    ./ 或 ../ 相對路徑 → 以 root 做基底
    其他絕對路徑（非 /tvconfigs）維持原樣
    其他純相對路徑 → root/相對
    """
    if tvconfigs_like.startswith("/tvconfigs/"):
        rel = tvconfigs_like[len("/tvconfigs/"):]
        return os.path.normpath(os.path.join(root, rel))
    if tvconfigs_like.startswith("./") or tvconfigs_like.startswith("../"):
        return os.path.normpath(os.path.join(root, tvconfigs_like))
    if tvconfigs_like.startswith("/"):
        return tvconfigs_like
    return os.path.normpath(os.path.join(root, tvconfigs_like))


def _find_key_value_in_ini_text(text: str, key: str) -> Optional[str]:
    """
    搜尋 key = value（忽略註解與空白、大小寫不敏感、允許引號），回傳原始值字串（未去引號）。
    """
    key_re = re.compile(r'^\s*' + re.escape(key) + r'\s*=\s*"?([^"\n\r]+)"?\s*$', re.IGNORECASE)
    for raw in text.splitlines():
        line = _strip_comment(raw)
        if not line:
            continue
        m = key_re.match(line)
        if m:
            return m.group(1).strip()
    return None


def parse_model_ini_for_default_settings(model_ini_path: str, root: str) -> Optional[str]:
    """
    從 model.ini 找 TvDefaultSettingsPath
    """
    txt = _read_text(model_ini_path)
    val = _find_key_value_in_ini_text(txt, "TvDefaultSettingsPath")
    if val is None:
        return None
    return _resolve_tvconfigs_path(root, val)


def parse_supportci20(default_settings_path: str) -> Optional[bool]:
    """
    讀取 default settings 檔案的 SUPPORTCI20，回傳 True/False；找不到或格式異常回傳 None。
    允許的 true 值：1/true/on/yes（大小寫不敏感）
           false 值：0/false/off/no
    """
    if not default_settings_path or not os.path.exists(default_settings_path):
        return None
    txt = _read_text(default_settings_path)
    val = _find_key_value_in_ini_text(txt, "SUPPORTCI20")
    if val is None:
        return None

    v = val.strip().lower()
    if v in ("1", "true", "on", "yes"):
        return True
    if v in ("0", "false", "off", "no"):
        return False
    # 嘗試一般整數解析
    try:
        return bool(int(v))
    except Exception:
        return None


def build_result(model_ini: str, default_path: Optional[str], supportci20: Optional[bool]) -> Dict:
    missing: List[str] = []
    notes: List[str] = []

    if default_path is None:
        notes.append("model.ini 未找到 TvDefaultSettingsPath")
    elif not os.path.exists(default_path):
        missing.append(default_path)

    decision = ""
    result_text = "N/A"
    supportci20_text = "N/A"

    if supportci20 is None:
        decision = "SUPPORTCI20 未找到或格式異常"
        result_text = "N/A"
    else:
        supportci20_text = "true" if supportci20 else "false"
        if supportci20:
            decision = "SUPPORTCI20 = true → CI 2.0"
            result_text = "CI 2.0"
        else:
            decision = "SUPPORTCI20 = false → CI 1.4.4"
            result_text = "CI 1.4.4"

    return {
        "model_ini_path": model_ini,                        # 僅用來決定 sheet 名，不輸出
        "default_settings_path_resolved": default_path or "",
        "supportci20_text": supportci20_text,
        "decision": decision,
        "result_text": result_text,
        "notes": "; ".join(notes),
        "missing": missing,
    }


# -----------------------------
# Main
# -----------------------------

def main():
    parser = argparse.ArgumentParser(description="Read TvDefaultSettingsPath → SUPPORTCI20, and report CI version (CI 2.0 vs CI 1.4.4) in Excel.")
    parser.add_argument("--model-ini", required=True, help="path to model ini (e.g., model/1_xxx.ini)")
    parser.add_argument("--root", required=True, help="tvconfigs project root (maps /tvconfigs/* to here)")
    parser.add_argument("-v", "--verbose", action="store_true", help="verbose logs")

    # 報表輸出
    parser.add_argument("--report", action="store_true", help="export to xlsx (default: kipling.xlsx)")
    parser.add_argument("--report-xlsx", metavar="FILE", help="export to specific xlsx file")

    args = parser.parse_args()

    model_ini = args.model_ini
    if not os.path.exists(model_ini):
        raise SystemExit(f"[ERROR] model ini not found: {model_ini}")
    root = os.path.abspath(os.path.normpath(args.root))

    if args.verbose:
        print(f"[INFO] model_ini: {model_ini}")
        print(f"[INFO] root     : {root}")

    # 解析 TvDefaultSettingsPath
    default_path = parse_model_ini_for_default_settings(model_ini, root)
    if args.verbose:
        print(f"[INFO] TvDefaultSettingsPath → {default_path if default_path else '(not found)'}")

    # 讀 SUPPORTCI20
    sup = parse_supportci20(default_path) if default_path else None
    if args.verbose:
        print(f"[INFO] SUPPORTCI20 = {('true' if sup else 'false') if sup is not None else '(not found)'}")

    # 組裝結果 + 輸出 console
    res = build_result(model_ini, default_path, sup)
    print(f"Result: {res['result_text']}")
    print(f"Decision: {res['decision']}")
    if res.get("notes"):
        print(f"Notes   : {res['notes']}")
    if res.get("missing"):
        print(f"Missing : {', '.join(res['missing'])}")

    # Excel
    if args.report or args.report_xlsx:
        xlsx_path = args.report_xlsx if args.report_xlsx else "kipling.xlsx"
        export_report(res, xlsx_path=xlsx_path)
        sheet = _sheet_name_for_model(model_ini)
        print(f"[INFO] Report appended to: {xlsx_path} (sheet: {sheet})")


if __name__ == "__main__":
    main()
