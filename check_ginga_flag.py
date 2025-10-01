import argparse
import os
import re
import xml.etree.ElementTree as ET
from collections import Counter
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# 檢查的 key 改成 Ginga
KEY = "persist.vendor.rtk.tv.enable_ginga"


def _sheet_name_for_model(model_ini_path: str) -> str:
    base = os.path.basename(model_ini_path or "")
    m = re.match(r"^(\d+)_", base)
    if m:
        return f"PID_{int(m.group(1))}"
    return "others"


def _read_text(path: str) -> str:
    for enc in ("utf-8", "latin-1", "utf-16"):
        try:
            with open(path, "r", encoding=enc) as f:
                return f.read()
        except UnicodeDecodeError:
            continue
        except FileNotFoundError:
            raise
    with open(path, "r") as f:
        return f.read()


def _strip_comment(line: str) -> str:
    line = line.split("#", 1)[0]
    line = line.split(";", 1)[0]
    return line.strip()


def _resolve_tvconfigs_path(root: str, tvconfigs_like: str) -> str:
    if tvconfigs_like.startswith("/tvconfigs/"):
        rel = tvconfigs_like[len("/tvconfigs/"):]
        return os.path.normpath(os.path.join(root, rel))
    if tvconfigs_like.startswith("./") or tvconfigs_like.startswith("../"):
        return os.path.normpath(os.path.join(root, tvconfigs_like))
    if tvconfigs_like.startswith("/"):
        return tvconfigs_like
    return os.path.normpath(os.path.join(root, tvconfigs_like))


def _file_flag_state(path: str, key: str) -> Tuple[bool, Optional[str]]:
    try:
        for enc in ("utf-8", "latin-1", "utf-16"):
            try:
                with open(path, "r", encoding=enc) as f:
                    for raw in f:
                        line = _strip_comment(raw)
                        if not line:
                            continue
                        if f"{key}=0" in line.replace(" ", ""):
                            return True, None
                return False, "NO_FLAG"
            except UnicodeDecodeError:
                continue
    except FileNotFoundError:
        return False, "MISSING"
    return False, "NO_FLAG"


def _load_country_blocks(xml_path: str, notes: List[str]) -> List[ET.Element]:
    txt = _read_text(xml_path)
    txt_no_decl = re.sub(r'<\?xml[^>]*\?>', '', txt)

    try:
        root = ET.fromstring(txt_no_decl)
        blocks = list(root.findall(".//COUNTRY_TVCONFIG_MAP")) or ([root] if root.tag == "COUNTRY_TVCONFIG_MAP" else [])
        if blocks:
            return blocks
    except Exception as e1:
        notes.append(f"XML normal parse failed: {e1}")

    try:
        wrapped = f"<ROOT>{txt_no_decl}</ROOT>"
        root = ET.fromstring(wrapped)
        blocks = list(root.findall(".//COUNTRY_TVCONFIG_MAP"))
        if blocks:
            return blocks
    except Exception as e2:
        notes.append(f"XML wrapped parse failed: {e2}")

    blocks = []
    for m in re.finditer(r"<COUNTRY_TVCONFIG_MAP>.*?</COUNTRY_TVCONFIG_MAP>", txt_no_decl, flags=re.DOTALL | re.IGNORECASE):
        frag = m.group(0)
        try:
            elem = ET.fromstring(frag)
            if elem.tag.upper() == "COUNTRY_TVCONFIG_MAP":
                blocks.append(elem)
        except Exception:
            pass

    if not blocks:
        notes.append("No COUNTRY_TVCONFIG_MAP blocks found")
    return blocks


def check_ginga_flag(model_ini: str, root: str, verbose: bool = False, dedup: bool = True) -> Dict:
    xml_path = os.path.join(root, "TvSysMap", "countryTvSysMap.xml")
    if verbose:
        print(f"[INFO] XML path: {xml_path}")

    checked_count = 0
    failed_detail: List[Tuple[str, str]] = []
    notes: List[str] = []

    if not os.path.exists(xml_path):
        notes.append("countryTvSysMap.xml not found")
    else:
        blocks = _load_country_blocks(xml_path, notes)
        if verbose:
            print(f"[INFO] Parsed COUNTRY_TVCONFIG_MAP blocks: {len(blocks)}")

        for block in blocks:
            tv_system_text = (block.findtext("TV_SYSTEM") or "").strip().upper()
            if not (tv_system_text == "ISDB" or tv_system_text.startswith("ISDB_")):
                continue

            tv_config_text = (block.findtext("TV_CONFIG") or "").strip()
            if not tv_config_text:
                failed_detail.append(("(missing TV_CONFIG)", "NO_FLAG"))
                checked_count += 1
                continue

            cfg_path = _resolve_tvconfigs_path(root, tv_config_text)
            ok, why = _file_flag_state(cfg_path, KEY)
            if not ok:
                failed_detail.append((cfg_path, why or "NO_FLAG"))
            checked_count += 1

    if dedup:
        ctr = Counter(f"{p} ({r})" for p, r in failed_detail)
        failed_summary = [f"{k} x{v}" for k, v in sorted(ctr.items())]
    else:
        failed_summary = [f"{p} ({r})" for p, r in failed_detail]

    overall_pass = (len(failed_detail) == 0) and (not notes)

    return {
        "passed": overall_pass,
        "xml_path": xml_path,
        "checked_count": checked_count,
        "failed_files": failed_summary,
        "failed_count": len(failed_detail),
        "notes": "; ".join(notes) if notes else "",
    }


def export_report(res: Dict, xlsx_path: str = "kipling.xlsx", sheet_name: str = "others", num_condition_cols: int = 6) -> None:
    COMMON_WIDTH = 80
    COMMON_ALIGN = Alignment(wrap_text=True, vertical="top")
    BOLD = Font(bold=True)

    #sheet_name = _sheet_name_for_model("dummy")

    try:
        wb = load_workbook(xlsx_path)
    except Exception:
        wb = Workbook()

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)
        headers = ["Rules", "Result"] + [f"condition_{i}" for i in range(1, num_condition_cols + 1)]
        ws.append(headers)

    rules = "For each ISDB/ISDB_* in countryTvSysMap.xml → tv.config has 'persist.vendor.rtk.tv.enable_ginga=0'?"
    result = "PASS" if res.get("passed", False) else "FAIL"

    conds = [
        f"TvSysMap XML = {res.get('xml_path') or 'N/A'}",
        f"ISDB tv.config checked = {res.get('checked_count', 0)}",
        f"Failures (count) = {res.get('failed_count', 0)}",
        f"Failed tv.config = {', '.join(res.get('failed_files', []) or []) or 'N/A'}",
        f"Notes = {res.get('notes') or 'N/A'}",
    ][:num_condition_cols]

    ws.append([rules, result] + conds)
    last_row = ws.max_row

    total_cols = 2 + num_condition_cols
    for col_idx in range(1, total_cols + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = COMMON_WIDTH

    for cell in ws[1]:
        cell.font = BOLD
        cell.alignment = COMMON_ALIGN

    for cell in ws[last_row]:
        cell.alignment = COMMON_ALIGN

    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        try:
            wb.remove(wb["Sheet"])
        except Exception:
            pass

    wb.save(xlsx_path)


def main():
    parser = argparse.ArgumentParser(description="Check ISDB tv.config ginga flag (persist.vendor.rtk.tv.enable_ginga=0)")
    parser.add_argument("--model-ini", required=True, help="path to model ini (e.g., model/1_xxx.ini)")
    parser.add_argument("--root", required=True, help="tvconfigs project root (maps /tvconfigs/* to here)")
    parser.add_argument("-v", "--verbose", action="store_true", help="verbose logs")
    parser.add_argument("--no-dedup", action="store_true", help="do not deduplicate failed files in outputs")
    parser.add_argument("--report", action="store_true", help="export report to xlsx (default: kipling.xlsx)")
    parser.add_argument("--report-xlsx", metavar="FILE", help="export report to specific xlsx file")
    args = parser.parse_args()

    res = check_ginga_flag(
        args.model_ini,
        os.path.abspath(os.path.normpath(args.root)),
        verbose=args.verbose,
        dedup=not args.no_dedup
    )

    print(f"Result : {'PASS' if res['passed'] else 'FAIL'}")
    print(f" - XML                  : {res['xml_path'] or '(not found)'}")
    print(f" - ISDB tv.config checked: {res['checked_count']}")
    print(f" - Failures (count)     : {res['failed_count']}")
    if res['failed_files']:
        print(f" - Failed tv.config:")
        for p in res['failed_files']:
            print(f"    {p}")
    if res['notes']:
        print(f" - Notes                : {res['notes']}")

    if args.report or args.report_xlsx:
        xlsx_path = args.report_xlsx if args.report_xlsx else "kipling.xlsx"
        sheet_name = _sheet_name_for_model(args.model_ini) 
        export_report(res, xlsx_path=xlsx_path, sheet_name=sheet_name)
        print(f"[INFO] Report appended to: {xlsx_path}")


if __name__ == "__main__":
    main()

