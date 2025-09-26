#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
check_setupwizard_flag.py

Standalone checker:
- Parse model.ini to find: isShowSetupwizard
- PASS if value == true (case-insensitive), otherwise FAIL (including missing or commented out)
- Optional XLSX report with only: Rules, Result, condition_1
- No model.ini/path fields are written into the report

Usage:
  python3 check_setupwizard_flag.py --model-ini model/1_xxx.ini
  python3 check_setupwizard_flag.py --model-ini model/1_xxx.ini --report-xlsx kipling.xlsx
"""

import argparse
import os
import re
from typing import Optional, Dict, Tuple, List

def _sheet_name_for_model(model_ini_path: str) -> str:
    """
    以 model.ini 檔名的數字前綴決定 sheet 名：'PID_<N>'；無數字則 'others'
    例：'1_EU_xxx.ini' → 'PID_1'
    """
    base = os.path.basename(model_ini_path or "")
    m = re.match(r"^(\d+)_", base)
    if m:
        return f"PID_{int(m.group(1))}"
    return "others"

# -----------------------------
# File reading & parsing helpers
# -----------------------------

def _read_text(path: str) -> str:
    """
    Read text with common encodings. Raises FileNotFoundError if missing.
    """
    for enc in ("utf-8", "latin-1", "utf-16"):
        try:
            with open(path, "r", encoding=enc) as f:
                return f.read()
        except UnicodeDecodeError:
            continue
        except FileNotFoundError:
            raise
    # Final fallback (no encoding specified)
    with open(path, "r") as f:
        return f.read()


def _strip_comment(line: str) -> str:
    """
    Remove comments starting with '#' or ';' and trim whitespace.
    """
    line = line.split("#", 1)[0]
    line = line.split(";", 1)[0]
    return line.strip()


def _find_is_show_setupwizard(model_ini_path: str) -> Optional[str]:
    """
    Find the (uncommented) value of isSupportDarkDetail in model.ini.
    Returns the raw value string (without quotes), or None if not found.
    """
    text = _read_text(model_ini_path)
    for raw in text.splitlines():
        line = _strip_comment(raw)
        if not line or "=" not in line:
            continue
        # Accept quoted or unquoted values; capture minimal any-char
        m = re.match(r'^\s*isSupportDarkDetail\s*=\s*("?)(.*?)\1\s*$', line, re.IGNORECASE)
        if m:
            return m.group(2).strip()
    return None


# -----------------------------
# Core check logic
# -----------------------------

def check_is_show_setupwizard(model_ini_path: str) -> Dict[str, object]:
    """
    Check if isSupportDarkDetail == true.
    Returns a result dict with keys: passed (bool), value (str or ''), notes (list[str])
    """
    notes: List[str] = []
    value = _find_is_show_setupwizard(model_ini_path)
    if value is None:
        notes.append("isSupportDarkDetail 未宣告或僅存在於註解中")
        passed = False
        value_str = ""
    else:
        passed = (value.strip().lower() == "true")
        if not passed:
            notes.append(f"isSupportDarkDetail 不是 true (got: {value})")
        value_str = value

    return {
        "passed": passed,
        "value": value_str,
        "notes": notes,
    }


# -----------------------------
# XLSX report (simple, no paths)
# -----------------------------

def export_simple_report(res: Dict[str, object], xlsx_path: str, sheet_name: str = "SupportDarkDetail") -> None:
    """
    Export a compact report with columns: Rules, Result, condition_1
    - No model.ini/path columns
    - Uniform column width, wrap text, vertical top; bold header
    """
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        raise SystemExit(
            "[ERROR] 需要 openpyxl 才能輸出報表。\n"
            "  安裝： pip install --user openpyxl\n"
        ) from e

    COMMON_WIDTH = 80
    ALIGN = Alignment(wrap_text=True, vertical="top")
    BOLD = Font(bold=True)

    # Open or create workbook
    try:
        wb = load_workbook(xlsx_path)
    except Exception:
        wb = Workbook()

    # Get or create sheet
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)
        ws.append(["Rules", "Result", "condition_1"])

    # Row content
    rules = "4. Dolyb Dark UI 要開\n" \
            "    - isSupportDarkDetail=true"
    result = "PASS" if bool(res.get("passed")) else "FAIL"
    cond1 = f"isSupportDarkDetail = {res.get('value') or 'N/A'}"

    ws.append([rules, result, cond1])
    last_row = ws.max_row

    # 給儲存格指派上色
    rules_color = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
    failed_color = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")
    # 上色
    first_cell = ws.cell(row=last_row, column=1)  # 欄位1對應的是 'A' 列
    first_cell.fill = rules_color
    if result == "FAIL":
        ws.cell(row=last_row, column=2).fill = failed_color
    if cond1 == "isSupportDarkDetail = N/A":
        ws.cell(row=last_row, column=3).fill = failed_color

    # Styling
    for col in range(1, 3 + 1):
        ws.column_dimensions[get_column_letter(col)].width = COMMON_WIDTH

    for cell in ws[1]:  # header row
        cell.font = BOLD
        cell.alignment = ALIGN

    for cell in ws[last_row]:
        cell.alignment = ALIGN

    # Remove default "Sheet" if others exist
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        try:
            wb.remove(wb["Sheet"])
        except Exception:
            pass

    wb.save(xlsx_path)


# -----------------------------
# CLI
# -----------------------------

def main():
    ap = argparse.ArgumentParser(
        description="Check model.ini:isSupportDarkDetail == true (PASS), else FAIL."
    )
    ap.add_argument("--model-ini", required=True, help="Path to model.ini")
    ap.add_argument("--report", action="store_true",
                    help="Write result to kipling.xlsx by default")
    ap.add_argument("--report-xlsx", metavar="FILE",
                    help="Write result to a specific XLSX file")
    ap.add_argument("-v", "--verbose", action="store_true", help="Verbose output")
    args = ap.parse_args()

    if not os.path.exists(args.model_ini):
        raise SystemExit(f"[ERROR] model ini not found: {args.model_ini}")

    res = check_is_show_setupwizard(args.model_ini)

    print(f"[CHECK] isSupportDarkDetail = {res['value'] or 'N/A'}")
    print(f"Result : {'PASS' if res['passed'] else 'FAIL'}")
    sheet = _sheet_name_for_model(args.model_ini)

    # Handle report
    if args.report_xlsx:
        export_simple_report(res, args.report_xlsx, sheet)
        print(f"[INFO] Report appended to: {args.report_xlsx}")
    elif args.report:
        export_simple_report(res, "kipling.xlsx", sheet)
        print(f"[INFO] Report appended to: kipling.xlsx")

if __name__ == "__main__":
    main()
