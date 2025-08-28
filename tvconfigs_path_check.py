#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
tvconfigs_path_check.py
掃描 configs/ 下的 .ini，檢查 ini 內指向的 /tvconfigs/... 檔案是否存在，
可輸出「格式化的 Excel」報表（Summary / Missing / All），
並可選擇將缺檔的那一行就地註解（在該行縮排後插入 '# '）。
⚠ 預設不會註解 sys/ 目錄內的 ini（仍會報告）；若要包含 sys/，請加 --include-sys-when-commenting。
"""

import argparse
import sys
from pathlib import Path
import re
from typing import Dict, Iterable, List, NamedTuple, Optional, Set, Tuple
from collections import Counter, defaultdict
import shutil
import os
import tempfile

DEFAULT_EXTS = {"ini", "bin", "dat", "img", "xml", "json", "png", "jpg", "webp"}
TV_PATH_RE = re.compile(r'(/tvconfigs/[^\s";]+)')
KEYVAL_RE = re.compile(r'^\s*([^;#=\s][^=]{0,120}?)\s*=\s*(.*)$')

class Ref(NamedTuple):
    ini_file: Path
    line_no: int
    key: str
    value_preview: str
    raw_tv_path: str
    resolved_path: Path
    exists: bool

def iter_ini_files(root: Path) -> Iterable[Path]:
    for p in root.rglob("*.ini"):
        if p.is_file():
            yield p

def sanitize_tv_path(p: str) -> str:
    return p.rstrip('",; \t\r\n)')

def looks_like_file_of_interest(tv_path: str, allowed_exts: Set[str]) -> bool:
    base = tv_path.split("?")[0]
    ext = Path(base).suffix.lower().lstrip(".")
    if not ext:
        return False
    return ext in allowed_exts

def resolve_to_project(root: Path, tv_path: str, prefix_map: Optional[Dict[str, Path]] = None) -> Optional[Path]:
    tv_path = sanitize_tv_path(tv_path)
    prefix_map = prefix_map or {"/tvconfigs/": root}
    for prefix, base in prefix_map.items():
        if tv_path.startswith(prefix):
            rel = tv_path[len(prefix):]
            return (base / rel).resolve()
    return None

def parse_key_value_preview(line: str) -> Tuple[str, str]:
    m = KEYVAL_RE.match(line.strip())
    if not m:
        return ("", line.strip()[:200])
    key = m.group(1).strip()
    val = m.group(2).strip()
    return (key, val[:200])

def scan_ini_for_tv_paths(ini_file: Path, root: Path, allowed_exts: Set[str], prefix_map: Optional[Dict[str, Path]]) -> List[Ref]:
    refs: List[Ref] = []
    with ini_file.open("r", encoding="utf-8", errors="ignore") as f:
        for i, raw_line in enumerate(f, start=1):
            line = raw_line.rstrip("\n")
            striped = line.strip()
            if not striped or striped.startswith(("#", ";")):
                continue
            matches = TV_PATH_RE.findall(line)
            if not matches:
                continue
            key, preview = parse_key_value_preview(line)
            for tv_path in matches:
                if not looks_like_file_of_interest(tv_path, allowed_exts):
                    continue
                resolved = resolve_to_project(root, tv_path, prefix_map)
                if resolved is None:
                    continue
                refs.append(
                    Ref(
                        ini_file=ini_file,
                        line_no=i,
                        key=key,
                        value_preview=preview,
                        raw_tv_path=tv_path,
                        resolved_path=resolved,
                        exists=resolved.exists(),
                    )
                )
    return refs

# --------- helpers for path checks ---------
def is_under(child: Path, parent: Path) -> bool:
    """Python 3.8 相容：判斷 child 是否在 parent 目錄底下"""
    try:
        child.resolve().relative_to(parent.resolve())
        return True
    except Exception:
        return False

# ---------- 就地註解缺檔行 ----------
def _comment_line_text(original: str) -> str:
    leading_len = len(original) - len(original.lstrip(" \t"))
    leading = original[:leading_len]
    rest = original[leading_len:]
    if not rest:
        return original
    if rest.lstrip().startswith("#") or rest.lstrip().startswith(";"):
        return original
    return f"{leading}# {rest}"

def comment_out_missing_lines(grouped_missing: Dict[Path, Set[int]], *, root: Path,
                              skip_dirs: Tuple[str, ...] = ("sys",),
                              backup_suffix: Optional[str] = ".bak") -> Dict[Path, List[int]]:
    """
    將 grouped_missing[ini_file] 的行號全部就地註解。
    預設跳過位於 root/sys/ 下的檔案（skip_dirs 可調整）。
    會建立備份（.bak；可用 backup_suffix=None 關閉）。
    """
    modified: Dict[Path, List[int]] = {}
    skip_parents = tuple((root / d).resolve() for d in skip_dirs if d)

    for ini_file, lines_to_edit in grouped_missing.items():
        # 跳過 skip_dirs（預設 sys/）
        if any(is_under(ini_file, sp) for sp in skip_parents):
            continue

        if not ini_file.exists() or not lines_to_edit:
            continue

        with ini_file.open("r", encoding="utf-8", errors="ignore") as f:
            lines = f.readlines()

        changed_any = False
        actually_changed: List[int] = []
        for ln in sorted(lines_to_edit):
            idx = ln - 1
            if idx < 0 or idx >= len(lines):
                continue
            original = lines[idx]
            stripped = original.lstrip(" \t")
            if stripped.startswith("#") or stripped.startswith(";"):
                continue
            new_line = _comment_line_text(original.rstrip("\n")) + ("\n" if original.endswith("\n") else "")
            if new_line != original:
                lines[idx] = new_line
                changed_any = True
                actually_changed.append(ln)

        if not changed_any:
            continue

        if backup_suffix:
            bak = ini_file.with_suffix(ini_file.suffix + backup_suffix)
            try:
                shutil.copy2(ini_file, bak)
            except Exception as e:
                print(f"[WARN] 無法建立備份 {bak}: {e}")

        tmp_fd, tmp_path = tempfile.mkstemp(prefix=ini_file.name + ".", dir=str(ini_file.parent))
        os.close(tmp_fd)
        tmp_p = Path(tmp_path)
        try:
            with tmp_p.open("w", encoding="utf-8", newline="") as fw:
                fw.writelines(lines)
            try:
                st = ini_file.stat()
                os.chmod(tmp_p, st.st_mode)
            except Exception:
                pass
            tmp_p.replace(ini_file)
        finally:
            if tmp_p.exists():
                try:
                    tmp_p.unlink()
                except Exception:
                    pass

        modified[ini_file] = actually_changed
    return modified

# ---------- 報表（Excel / CSV） ----------
def beautify_excel(writer, sheet_names: Dict[str, List[str]]) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.formatting.rule import CellIsRule

    table_style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True, showColumnStripes=False)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    wrap_alignment = Alignment(wrap_text=True, vertical="top")

    for sheet, cols in sheet_names.items():
        ws = writer.sheets[sheet]
        max_row = ws.max_row
        max_col = ws.max_column
        if max_row < 2 or max_col < 1:
            continue

        ref = f"A1:{chr(64+max_col)}{max_row}"
        try:
            table = Table(displayName=f"{sheet.replace(' ','_')}_Table", ref=ref)
            table.tableStyleInfo = table_style
            ws.add_table(table)
        except Exception:
            ws.auto_filter.ref = ref

        ws.freeze_panes = "A2"

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        preferred = {"ini_file": 28, "line_no": 8, "key": 26, "value_preview": 60,
                     "tv_path": 48, "resolved_path": 72, "exists": 10, "ext": 8,
                     "item": 22, "value": 16, "modified_lines": 16}
        for j in range(1, max_col+1):
            col_letter = ws.cell(row=1, column=j).column_letter
            header = ws.cell(row=1, column=j).value or ""
            max_len = len(str(header))
            for i in range(2, min(max_row, 3000)+1):
                val = ws.cell(row=i, column=j).value
                if val is None:
                    continue
                l = len(str(val))
                if l > max_len:
                    max_len = l
            width = min(max(preferred.get(str(header), 12), int(max_len*0.95)), 80)
            ws.column_dimensions[col_letter].width = width
            if str(header) in ("value_preview", "tv_path", "resolved_path"):
                for i in range(2, max_row+1):
                    ws.cell(row=i, column=j).alignment = wrap_alignment

        # exists 色塊
        exists_col_idx = None
        for j in range(1, max_col+1):
            if str(ws.cell(row=1, column=j).value).lower() == "exists":
                exists_col_idx = j
                break
        if exists_col_idx:
            col_letter = ws.cell(row=1, column=exists_col_idx).column_letter
            rng = f"{col_letter}2:{col_letter}{max_row}"
            from openpyxl.styles import PatternFill
            green = PatternFill("solid", fgColor="C6EFCE")
            red = PatternFill("solid", fgColor="FFC7CE")
            ws.conditional_formatting.add(rng, CellIsRule(operator='equal', formula=['TRUE'], fill=green))
            ws.conditional_formatting.add(rng, CellIsRule(operator='equal', formula=['FALSE'], fill=red))

def write_reports(root: Path, all_refs: List[Ref], xlsx_path: Optional[Path], csv_dir: Optional[Path],
                  modified_map: Optional[Dict[Path, List[int]]] = None) -> None:
    def row_dict(r: Ref) -> Dict[str, object]:
        return {
            "ini_file": str(r.ini_file.relative_to(root)),
            "line_no": r.line_no,
            "key": r.key,
            "value_preview": r.value_preview,
            "tv_path": r.raw_tv_path,
            "resolved_path": str(r.resolved_path),
            "exists": r.exists,
            "ext": Path(r.resolved_path).suffix.lstrip(".").lower(),
        }

    rows_all = [row_dict(r) for r in all_refs]
    rows_missing = [row for row in rows_all if not row["exists"]]

    ext_counts = Counter([row["ext"] for row in rows_missing])
    summary = [
        {"item": "scan_root", "value": str(root)},
        {"item": "ini_files_scanned", "value": len({r.ini_file for r in all_refs})},
        {"item": "total_references", "value": len(rows_all)},
        {"item": "missing_count", "value": len(rows_missing)},
    ]
    for ext, cnt in sorted(ext_counts.items(), key=lambda x: (-x[1], x[0])):
        summary.append({"item": f"missing_by_ext::{ext or '(none)'}", "value": cnt})

    if modified_map:
        total_files = len(modified_map)
        total_lines = sum(len(v) for v in modified_map.values())
        summary.append({"item": "modified_files", "value": total_files})
        summary.append({"item": "modified_lines", "value": total_lines})

    if xlsx_path:
        try:
            import pandas as pd  # type: ignore
            with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
                cols_missing = ["ini_file", "line_no", "key", "tv_path", "resolved_path", "ext"]
                cols_all = ["ini_file", "line_no", "key", "value_preview", "tv_path", "resolved_path", "exists", "ext"]
                pd.DataFrame(summary).to_excel(writer, index=False, sheet_name="Summary")
                pd.DataFrame(rows_missing, columns=cols_missing).to_excel(writer, index=False, sheet_name="Missing")
                pd.DataFrame(rows_all, columns=cols_all).to_excel(writer, index=False, sheet_name="All")
                if modified_map:
                    mod_rows = [{"ini_file": str(p.relative_to(root)),
                                 "modified_lines": ", ".join(map(str, sorted(v)))}
                                for p, v in modified_map.items()]
                    import pandas as pd  # reuse
                    pd.DataFrame(mod_rows, columns=["ini_file", "modified_lines"]).to_excel(
                        writer, index=False, sheet_name="Modified"
                    )
                beautify_excel(writer, {
                    "Summary": ["item", "value"],
                    "Missing": cols_missing,
                    "All": cols_all,
                    **({"Modified": ["ini_file", "modified_lines"]} if modified_map else {}),
                })
            print(f"✔ 已輸出 Excel 報表（格式化）：{xlsx_path}")
            return
        except Exception as e:
            print(f"[WARN] 產生 Excel 失敗：{e}\n→ 將改輸出 CSV。")

    out_dir = csv_dir or root
    out_dir.mkdir(parents=True, exist_ok=True)
    def write_csv(name: str, headers: List[str], rows: List[Dict[str, object]]):
        p = out_dir / name
        import csv
        with p.open("w", encoding="utf-8", newline="") as f:
            w = csv.DictWriter(f, fieldnames=headers)
            w.writeheader()
            for row in rows:
                w.writerow({h: row.get(h, "") for h in headers})
        print(f"✔ 已輸出 CSV：{p}")

    write_csv("summary.csv", ["item","value"], summary)
    write_csv("missing.csv", ["ini_file","line_no","key","tv_path","resolved_path","ext"], rows_missing)
    write_csv("all.csv", ["ini_file","line_no","key","value_preview","tv_path","resolved_path","exists","ext"], rows_all)
    if modified_map:
        mod_rows = [{"ini_file": str(p.relative_to(root)),
                     "modified_lines": ", ".join(map(str, sorted(v)))}
                    for p, v in modified_map.items()]
        write_csv("modified.csv", ["ini_file","modified_lines"], mod_rows)

# ---------- 主流程 ----------
def main():
    parser = argparse.ArgumentParser(
        description="檢查 ini 內指向的 /tvconfigs/... 檔案是否存在，輸出格式化報表，並可就地註解缺檔行（預設跳過 sys/）"
    )
    parser.add_argument("--root", required=True, type=Path, help="tvconfigs 專案根（如 ~/tvconfigs_home/tv109/kipling/configs）")
    parser.add_argument("--exts", type=str, default=",".join(sorted(DEFAULT_EXTS)),
                        help=f"要檢查的副檔名（逗號分隔），預設：{','.join(sorted(DEFAULT_EXTS))}")
    parser.add_argument("--report-xlsx", type=Path, default=None, help="輸出 Excel 的路徑（.xlsx）")
    parser.add_argument("--csv-dir", type=Path, default=None, help="若 Excel 失敗，CSV 的輸出資料夾（預設 root）")
    parser.add_argument("--fail-warning", action="store_true", help="有缺檔則以非 0 退出碼（CI 友善）")
    # 註解選項
    parser.add_argument("--comment-missing", action="store_true",
                        help="將缺檔的那一行就地註解（在縮排後插入 '# '）。預設僅檢查不修改。")
    parser.add_argument("--include-sys-when-commenting", action="store_true",
                        help="包含 sys/ 目錄文件也進行註解（預設不包含）。")
    parser.add_argument("--backup-suffix", type=str, default="",
                        help="備份副檔名（預設不備份；例如 --backup-suffix .bak 啟用）")
    args = parser.parse_args()

    root: Path = args.root.resolve()
    if not root.exists():
        print(f"[ERROR] root 不存在：{root}", file=sys.stderr)
        sys.exit(2)

    allowed_exts = {e.strip().lower() for e in args.exts.split(",") if e.strip()}
    prefix_map = {"/tvconfigs/": root}

    all_refs: List[Ref] = []
    for ini in iter_ini_files(root):
        all_refs.extend(scan_ini_for_tv_paths(ini, root, allowed_exts, prefix_map))

    missing_refs = [r for r in all_refs if not r.exists]

    print(f"掃描根目錄：{root}")
    print(f"發現參照總數：{len(all_refs)}，缺檔：{len(missing_refs)}")
    if missing_refs:
        print("\n=== 缺檔清單（節選） ===")
        for r in missing_refs[:50]:
            print(f"- {r.ini_file.relative_to(root)}:{r.line_no}  {r.raw_tv_path}  ->  {r.resolved_path}")

    modified_map: Optional[Dict[Path, List[int]]] = None
    if args.comment_missing and missing_refs:
        grouped: Dict[Path, Set[int]] = defaultdict(set)
        for r in missing_refs:
            grouped[r.ini_file].add(r.line_no)
        # 預設跳過 sys/；如需包含，使用 --include-sys-when-commenting
        skip_dirs = () if args.include_sys_when_commenting else ("sys",)
        backup_suffix = args.backup_suffix if args.backup_suffix != "" else None
        modified_map = comment_out_missing_lines(grouped, root=root, skip_dirs=skip_dirs, backup_suffix=backup_suffix)
        tot_files = len(modified_map)
        tot_lines = sum(len(v) for v in modified_map.values())
        print(f"\n✔ 已就地註解缺檔行（不含 sys/）：" + ("" if skip_dirs else "（含 sys/）"))
        print(f"  修改檔案數：{tot_files}，修改行數：{tot_lines}")
        if backup_suffix:
            print(f"（已建立備份：*{backup_suffix}）")

    write_reports(root, all_refs, args.report_xlsx, args.csv_dir, modified_map=modified_map)

    if missing_refs:
        sys.exit(1 if args.fail_warning else 0)
    sys.exit(0)

if __name__ == "__main__":
    main()

