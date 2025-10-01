#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
check_supportedDeviceGooglePairer.py

需求：
- 參考 check_CI.py 的「/tvconfigs → --root」路徑映射、xlsx 輸出格式、sheet 命名規則
- 在 xlsx 內「移除 model.ini 欄位」（僅用於決定分頁名，不輸出）
- 讀取 model.ini 的 supportedDeviceGooglePairer 並打印/寫入報表

Python 3.8+（報表需 openpyxl）
"""

import argparse
import os
import re
from typing import Optional, Dict, List

# -----------------------------
# 路徑/解析工具
# -----------------------------

def _read_text(path: str) -> str:
    for enc in ("utf-8", "latin-1", "utf-16"):
        try:
            with open(path, "r", encoding=enc) as f:
                return f.read()
        except UnicodeDecodeError:
            continue
        except FileNotFoundError:
            raise
    with open(path, "r") as f:
        return f.read()

def _strip_comment(line: str) -> str:
    # 去掉 # 或 ; 後的註解
    s = line.strip()
    if not s:
        return ""
    for mark in ("#", ";"):
        pos = s.find(mark)
        if pos != -1:
            s = s[:pos]
    return s.strip()

def _map_tvconfigs_to_root(path_like: str, root: str) -> str:
    """
    /tvconfigs/... → <root>/...
    ./ 或 ../ → 以 root 為基底
    / 開頭（非 /tvconfigs）維持原樣
    其他相對路徑 → <root>/<相對>
    """
    path_like = (path_like or "").strip()
    if path_like.startswith("/tvconfigs/"):
        rel = path_like[len("/tvconfigs/"):]
        return os.path.normpath(os.path.join(root, rel))
    if path_like.startswith("./") or path_like.startswith("../"):
        return os.path.normpath(os.path.join(root, path_like))
    if path_like.startswith("/"):
        return path_like
    return os.path.normpath(os.path.join(root, path_like))

def _find_key_value_in_ini_text(text: str, key: str) -> Optional[str]:
    """
    在整份 INI 文字中找 key=value（大小寫不敏感），回傳 value（允許引號/不引號；忽略註解與前後空白）。
    """
    key_re = re.compile(r'^\s*' + re.escape(key) + r'\s*=\s*"?([^"\n\r]+)"?\s*$', re.IGNORECASE)
    for raw in text.splitlines():
        line = _strip_comment(raw)
        if not line:
            continue
        m = key_re.match(line)
        if m:
            return m.group(1).strip()
    return None

# -----------------------------
# Excel 報表（沿用專案風格）
# -----------------------------

def _sheet_name_for_model(model_ini_path: str) -> str:
    """
    以 model.ini 檔名的數字前綴決定 sheet 名：'PID_<N>'；無數字則 'others'
    例：'1_EU_xxx.ini' → 'PID_1'
    """
    base = os.path.basename(model_ini_path or "")
    m = re.match(r"^(\d+)_", base)
    if m:
        return f"PID_{int(m.group(1))}"
    return "others"

def _ensure_openpyxl():
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        raise SystemExit(
            "[ERROR] 需要 openpyxl 以輸出/附加報表。\n"
            "  安裝： pip install --user openpyxl\n"
        )

def _na(s: Optional[str]) -> str:
    s = (s or "").strip()
    return s if s else "N/A"

def export_report(res: Dict, xlsx_path: str = "kipling.xlsx", num_condition_cols: int = 5) -> None:
    """
    表頭固定：Rules | Result | condition_1..N
    - 不輸出 model.ini 欄位
    - 依 PID_N/others 分頁，若 xlsx 存在則附加一列
    - 欄位等寬、換行、垂直置頂
    """
    _ensure_openpyxl()
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    COMMON_WIDTH = 80
    COMMON_ALIGN = Alignment(wrap_text=True, vertical="top")
    BOLD = Font(bold=True)

    sheet_name = _sheet_name_for_model(res.get("model_ini_path", ""))

    # 開啟或建立
    try:
        wb = load_workbook(xlsx_path)
    except Exception:
        wb = Workbook()

    # 取得/建立分頁
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)
        headers = ["Rules", "Result"] + [f"condition_{i}" for i in range(1, num_condition_cols + 1)]
        ws.append(headers)
        # 設 header 樣式與欄寬
        for col in range(1, 2 + num_condition_cols + 1):
            ws.cell(row=1, column=col).font = BOLD
            ws.column_dimensions[get_column_letter(col)].width = COMMON_WIDTH

    # 準備資料
    rules = "從 model.ini 讀取 supportedDeviceGooglePairer"
    result = res.get("result_text") or "N/A"

    conds = [
        f"model.ini = {_na(res.get('model_ini_resolved'))}",          # c1
        f"supportedDeviceGooglePairer = {_na(res.get('raw_value'))}", # c2
        f"Notes = {_na(res.get('notes'))}",                           # c3
        f"Missing = {_na(', '.join(res.get('missing') or []))}",      # c4
        "",                                                           # c5(保留)
    ][:num_condition_cols]

    ws.append([rules, result] + conds)
    last_row = ws.max_row
    # 設定該列對齊
    for col in range(1, 2 + num_condition_cols + 1):
        ws.cell(row=last_row, column=col).alignment = COMMON_ALIGN

    # 移除預設 Sheet（若存在且有其他分頁）
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        try:
            wb.remove(wb["Sheet"])
        except Exception:
            pass

    wb.save(xlsx_path)

# -----------------------------
# Core
# -----------------------------

def parse_supported_pairer(model_ini: str, root: str, verbose: bool = False) -> Dict:
    missing: List[str] = []
    notes: List[str] = []

    model_ini_resolved = _map_tvconfigs_to_root(model_ini, root)

    if verbose:
        print(f"[INFO] model_ini: {model_ini}")
        print(f"[INFO] root     : {os.path.abspath(root)}")
        print(f"[INFO] resolved : {model_ini_resolved}")

    try:
        text = _read_text(model_ini_resolved)
    except FileNotFoundError:
        missing.append(f"model.ini not found: {model_ini_resolved}")
        return {
            "model_ini_path": model_ini,
            "model_ini_resolved": model_ini_resolved,
            "raw_value": None,
            "result_text": "N/A",
            "notes": "; ".join(notes),
            "missing": missing,
        }

    raw_val = _find_key_value_in_ini_text(text, "supportedDeviceGooglePairer")
    if raw_val is None:
        missing.append("supportedDeviceGooglePairer not found in model.ini")
        result_text = "N/A"
    else:
        result_text = raw_val  # 不正規化，原值輸出

    return {
        "model_ini_path": model_ini,             # 僅用於決定分頁名，不輸出成欄位
        "model_ini_resolved": model_ini_resolved,
        "raw_value": raw_val,
        "result_text": result_text,
        "notes": "; ".join(notes),
        "missing": missing,
    }

# -----------------------------
# Main
# -----------------------------

def main():
    ap = argparse.ArgumentParser(description="Read supportedDeviceGooglePairer from model.ini and (optionally) append to Excel.")
    ap.add_argument("--model-ini", required=True, help="path to model ini (e.g., model/1_xxx.ini or /tvconfigs/model/1_xxx.ini)")
    ap.add_argument("--root", required=True, help="tvconfigs project root (maps /tvconfigs/* to here)")
    ap.add_argument("-v", "--verbose", action="store_true")
    ap.add_argument("--report", action="store_true", help="append to kipling.xlsx")
    ap.add_argument("--report-xlsx", default=None, help="custom xlsx path (overrides --report default)")
    args = ap.parse_args()

    res = parse_supported_pairer(args.model_ini, args.root, verbose=args.verbose)

    # Console output
    if args.verbose:
        print("----------------------------------------")
    print(f"supportedDeviceGooglePairer: {res.get('result_text', 'N/A')}")
    if res.get("missing"):
        print(f"Missing     : {', '.join(res['missing'])}")
    if res.get("notes"):
        print(f"Notes       : {res['notes']}")

    # Excel
    if args.report or args.report_xlsx:
        xlsx_path = args.report_xlsx if args.report_xlsx else "kipling.xlsx"
        export_report(res, xlsx_path=xlsx_path)
        sheet = _sheet_name_for_model(res.get("model_ini_path",""))
        print(f"[INFO] Report appended to: {xlsx_path} (sheet: {sheet})")

if __name__ == "__main__":
    main()

