#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
check_EWBS.py

規則（依你的說明，並維持原檢查邏輯）：
  1) 在 model.ini 找到（且非註解行）
       - isSupportEWBS = true
       - isSupportNeverEnterSTR = <任意值>
       - isEwbsSettingOn = <任意值>
     三者同時符合 → PASS；否則 FAIL。
  2) 讀取 model.ini 內的 COUNTRY_PATH 指向的檔案內容，若包含以下任何國家名稱則列印：
       "Botswana","Ecuador","Maldives","Philippines","Japan","Peru","Venezuela","Costa Rica"
"""
import argparse
import os
import re
from typing import Optional, List, Any, Dict

# ──────────────────────────────────────────────────────────────────────────────
# Report helpers (aligned with tv_multi_standard_validation.py)
# ──────────────────────────────────────────────────────────────────────────────
def _sheet_name_for_model(model_ini_path: str) -> str:
    base = os.path.basename(model_ini_path or "")
    m = re.match(r"^(\d+)_", base)
    if m:
        return f"PID_{int(m.group(1))}"
    return "others"


def _ensure_openpyxl():
    try:
        import openpyxl  # noqa
    except ImportError:
        raise SystemExit(
            "[ERROR] 需要 openpyxl 以支援報表輸出與附加。\n"
            "  安裝： pip install --user openpyxl\n"
        )


def export_report(res: dict, xlsx_path: str = "kipling.xlsx", num_condition_cols: int = 5) -> None:
    """
    表頭固定為: Rules, Result, condition_1, condition_2, condition_3, ...
    欄位無值時以 'N/A' 填入。依 model.ini 檔名前綴分頁（PID_1、PID_2…；非數字→others），既有資料則附加。
    全欄統一樣式：同寬、換行、垂直置頂（包含表頭）。
    """
    _ensure_openpyxl()
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    COMMON_WIDTH = 80
    COMMON_ALIGN = Alignment(wrap_text=True, vertical="top")
    BOLD = Font(bold=True)

    def _na(s: str) -> str:
        s = (s or "").strip()
        return s if s else "N/A"

    sheet_name = _sheet_name_for_model(res.get("model_ini", ""))

    # 開啟或新建 xlsx
    try:
        wb = load_workbook(xlsx_path)
    except Exception:
        wb = Workbook()

    # 建立或取得 sheet（表頭固定順序）
    header = ["Rules", "Result"] + [f"condition_{i}" for i in range(1, num_condition_cols + 1)]
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row < 1:
            ws.append(header)
    else:
        ws = wb.create_sheet(title=sheet_name)
        ws.append(header)

    # 取值
    rules      = _na(res.get("rules", ""))
    result     = _na(res.get("result", ""))
    conditions = [ _na(x) for x in (res.get("conditions", []) or []) ]
    """
    # 補足 condition_* 欄位數
    if len(conditions) < num_condition_cols:
        conditions += ["N/A"] * (num_condition_cols - len(conditions))
    else:
        conditions = conditions[:num_condition_cols]
    """
    # 寫入一列
    row_values = [rules, result] + conditions
    ws.append(row_values)
    last_row = ws.max_row

    # 給儲存格指派上色
    rules_color = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
    failed_color = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")
    # 上色
    first_cell = ws.cell(row=last_row, column=1)  # 欄位1對應的是 'A' 列
    first_cell.fill = rules_color
    if result == "FAIL":
        ws.cell(row=last_row, column=2).fill = failed_color

    # ── 統一樣式：所有欄位同寬 & 換行 & 垂直置頂（含表頭） ──
    total_cols = 2 + num_condition_cols
    for col_idx in range(1, total_cols + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = COMMON_WIDTH

    # 表頭樣式
    for cell in ws[1]:
        cell.font = BOLD
        cell.alignment = COMMON_ALIGN

    # 資料列樣式（最新一列）
    for col_idx in range(1, total_cols + 1):
        ws.cell(row=last_row, column=col_idx).alignment = COMMON_ALIGN

    # 移除預設空白 Sheet（若存在且非唯一）
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        try:
            wb.remove(wb["Sheet"])
        except Exception:
            pass

    wb.save(xlsx_path)


# ──────────────────────────────────────────────────────────────────────────────
# EWBS core logic
# ──────────────────────────────────────────────────────────────────────────────
COUNTRIES_TO_PRINT = [
    "Botswana", "Ecuador", "Maldives", "Philippines",
    "Japan", "Peru", "Venezuela", "Costa Rica",
]


def _read_lines(path: str) -> List[str]:
    for enc in ("utf-8", "latin-1", "utf-16"):
        try:
            with open(path, "r", encoding=enc, errors="ignore") as f:
                return f.readlines()
        except UnicodeDecodeError:
            continue
        except FileNotFoundError:
            raise
    with open(path, "r") as f:
        return f.readlines()


def _iter_non_comment_lines(path: str):
    for raw in _read_lines(path):
        s = raw.strip()
        if not s or s.startswith("#"):
            continue
        yield s


def _extract_quoted_value(line: str, key: str) -> str:
    """從 'key = "..."' 抽出雙引號字串，失敗回傳空字串。"""
    m = re.search(rf'{re.escape(key)}\s*=\s*"([^"]+)"', line)
    return m.group(1).strip() if m else ""


def _parse_bool_from_line(line: str, key: str) -> Optional[bool]:
    """
    從 'key = true/false' 解析布林值（忽略大小寫與空白），
    也接受 key = "true"/"false"。找不到回傳 None。
    """
    m = re.search(rf'{re.escape(key)}\s*=\s*("?)(true|false)\1', line, flags=re.IGNORECASE)
    if not m:
        return None
    return m.group(2).lower() == "true"


def _resolve_tvconfigs_path(raw: str, root_dir: str) -> str:
    """
    將 /tvconfigs/... 轉為實際檔案路徑：${abs_root}/...
    規則：移除 'tvconfigs/' 前綴，接在 abs(root_dir) 後面。
    其他相對路徑則視為相對於 root_dir。
    """
    root_dir = os.path.abspath(root_dir)
    raw = raw.strip()
    if raw.startswith("/tvconfigs/"):
        rel_path = raw[len("/tvconfigs/"):]
        return os.path.join(root_dir, rel_path)
    if raw.startswith("./") or raw.startswith("../"):
        return os.path.normpath(os.path.join(root_dir, raw))
    if raw.startswith("/"):
        return raw
    return os.path.normpath(os.path.join(root_dir, raw))


def check_ewbs(model_ini_path: str, root_dir: str = ".") -> dict:
    """
    回傳：
      {
        "result": "PASS"/"FAIL",
        "rules": <str>,
        "model_ini": <str>,
        "conditions": [ ... ]
      }
    並在 console 印出檢查過程。
    """
    root_dir = os.path.abspath(root_dir)

    # --- 第一階段：讀取三個旗標 ---
    val_ewbs = val_never_str = val_setting_on = None
    ewbs_line = never_str_line = setting_on_line = None
    country_line = None

    for s in _iter_non_comment_lines(model_ini_path):
        if ewbs_line is None and "isSupportEWBS" in s and "=" in s:
            ewbs_line = s
            val_ewbs = _parse_bool_from_line(s, "isSupportEWBS")

        if never_str_line is None and "isSupportNeverEnterSTR" in s and "=" in s:
            never_str_line = s
            val_never_str = _parse_bool_from_line(s, "isSupportNeverEnterSTR")

        if setting_on_line is None and "isEwbsSettingOn" in s and "=" in s:
            setting_on_line = s
            val_setting_on = _parse_bool_from_line(s, "isEwbsSettingOn")

        if country_line is None and "COUNTRY_PATH" in s and "=" in s:
            country_line = s

        if (ewbs_line is not None and never_str_line is not None and
            setting_on_line is not None and country_line is not None):
            break

    # 組合輸出
    print("=== EWBS Check ===")
    print(f"Model.ini : {model_ini_path}")
    if ewbs_line:        print(f"→ {ewbs_line}")
    if never_str_line:   print(f"→ {never_str_line}")
    if setting_on_line:  print(f"→ {setting_on_line}")

    # 判定 PASS/FAIL（isSupportEWBS 必須為 true；另外兩個鍵只要有宣告即可）
    flags_ok = (val_ewbs is True) and (never_str_line is not None) and (setting_on_line is not None)
    if flags_ok:
        print("→ PASS（isSupportEWBS = true / isSupportNeverEnterSTR set / isEwbsSettingOn set）")
        result = "PASS"
    else:
        reasons = []
        if val_ewbs is not True:        reasons.append("isSupportEWBS != true")
        if never_str_line is None:      reasons.append("isSupportNeverEnterSTR not set")
        if setting_on_line is None:     reasons.append("isEwbsSettingOn not set")
        if not reasons:
            reasons.append("缺少必要欄位（可能未宣告或格式錯誤）")
        print(f"→ FAIL（{', '.join(reasons)}）")
        result = "FAIL"

    # --- 第二階段：開 COUNTRY_PATH，列印指定國家（若存在） ---
    abs_country_path = None
    found_names: List[str] = []
    if country_line:
        raw_country_path = _extract_quoted_value(country_line, "COUNTRY_PATH")
        if raw_country_path:
            abs_country_path = _resolve_tvconfigs_path(raw_country_path, root_dir)
            if os.path.exists(abs_country_path):
                try:
                    with open(abs_country_path, "r", encoding="utf-8", errors="ignore") as cf:
                        content = cf.read()
                    for name in COUNTRIES_TO_PRINT:
                        if re.search(re.escape(name), content, flags=re.IGNORECASE):
                            found_names.append(name)
                    if found_names:
                        print(f"→ COUNTRY_PATH 內包含國家：{', '.join(found_names)}")
                    else:
                        print("→ COUNTRY_PATH 內未找到指定國家")
                except Exception as e:
                    print(f"→ 讀取 COUNTRY_PATH 檔案失敗：{e}")
            else:
                print(f"→ COUNTRY_PATH 檔案不存在：{abs_country_path}")
        else:
            print("→ COUNTRY_PATH 格式錯誤或缺少引號")
    else:
        print("→ COUNTRY_PATH = N/A")

    # 準備報表資料
    rules = f"9. EWBS 驗證 ( 國家要選菲律賓)\n    - isSupportEWBS = true ?\n    - isSupportNeverEnterSTR is set ?\n    - isEwbsSettingOn is set ?\n    - Scan COUNTRY_PATH for specific countries"
    conditions = [
        f"isSupportEWBS = {'true' if val_ewbs is True else ('false' if val_ewbs is False else 'N/A')}", # condition_1
        f"isSupportNeverEnterSTR = {'set' if never_str_line else 'N/A'}",                                # condition_2
        f"isEwbsSettingOn = {'set' if setting_on_line else 'N/A'}",                                      # condition_3
        f"Country Path = {abs_country_path if abs_country_path else 'N/A'}",                             # condition_4
        f"Found Countries = {', '.join(found_names) if found_names else 'N/A'}",                         # condition_5
    ]

    return {
        "result": result,
        "rules": rules,
        "model_ini": model_ini_path,  # 用於分頁，不直接輸出欄位
        "conditions": conditions,
    }

def run(
    model_ini: str,
    root: str = ".",
    standard: Optional[str] = None,
    verbose: bool = False,
    conditions: str = "",
    report_xlsx: Optional[str] = None,
    ctx: Any = None,
    **kwargs,                         # 吸收多餘參數避免 TypeError
) -> Dict[str, Any]:
    res = check_ewbs(model_ini, root)

    # 報表輸出
    if report_xlsx:
        out_xlsx = f"{report_xlsx}.xlsx" if not report_xlsx.endswith(".xlsx") else report_xlsx
        export_report(res, xlsx_path=out_xlsx, num_condition_cols=conditions)
        sheet = _sheet_name_for_model(res.get("model_ini", ""))
        print(f"[INFO] Report appended to: {out_xlsx} (sheet: {sheet})")

# ──────────────────────────────────────────────────────────────────────────────
# CLI
# ──────────────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Check EWBS flags and list specific countries from COUNTRY_PATH, with report output.")
    parser.add_argument("--root", required=True, help="專案根目錄路徑（映射 /tvconfigs/* 至此）")
    parser.add_argument("--model-ini", required=True, help="model.ini 檔案路徑")
    parser.add_argument("--report", action="store_true", help="輸出報表到 xlsx（預設 kipling.xlsx）")
    parser.add_argument("--report-xlsx", metavar="FILE", help="指定報表 xlsx 檔案名稱")
    parser.add_argument("--conditions", type=int, default=5, help="condition_* 欄位數（預設 5）")
    parser.add_argument("-v", "--verbose", action="store_true", help="顯示詳細過程")
    args = parser.parse_args()

    res = check_ewbs(args.model_ini, args.root)

    # 報表輸出
    if args.report or args.report_xlsx:
        xlsx_path = args.report_xlsx if args.report_xlsx else "kipling.xlsx"
        export_report(res, xlsx_path=xlsx_path, num_condition_cols=args.conditions)
        sheet = _sheet_name_for_model(res.get("model_ini", ""))
        print(f"[INFO] Report appended to: {xlsx_path} (sheet: {sheet})")


if __name__ == "__main__":
    main()
