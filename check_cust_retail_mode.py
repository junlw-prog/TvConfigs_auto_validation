#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import argparse
import os
import re
from typing import Optional

def _sheet_name_for_model(model_ini_path: str) -> str:
    base = os.path.basename(model_ini_path or "")
    m = re.match(r"^(\d+)_", base)
    if m:
        return f"PID_{int(m.group(1))}"
    return "others"

def _ensure_openpyxl():
    try:
        import openpyxl  # noqa
    except ImportError:
        raise SystemExit(
            "[ERROR] 需要 openpyxl 套件以支援報表輸出\n"
            "  安裝： pip install --user openpyxl\n"
        )

def export_report(res: dict, xlsx_path: str = "kipling.xlsx", num_condition_cols: int = 3) -> None:
    _ensure_openpyxl()
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    COMMON_WIDTH = 60
    COMMON_ALIGN = Alignment(wrap_text=True, vertical="top")
    BOLD = Font(bold=True)

    def _na(s: str) -> str:
        s = (s or "").strip()
        return s if s else "N/A"

    sheet_name = _sheet_name_for_model(res.get("model_ini", ""))

    try:
        wb = load_workbook(xlsx_path)
    except Exception:
        wb = Workbook()

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)
        headers = ["Rules", "Result"] + [f"condition_{i}" for i in range(1, num_condition_cols + 1)]
        ws.append(headers)

    rules = "CustRetailModeInten defined?"
    result = "PASS" if res.get("passed", False) else "FAIL"
    cust_val = _na(res.get("cust_val", ""))

    conds = [
        f"CustRetailModeInten = {cust_val}",
    ][:num_condition_cols]

    row_values = [rules, result] + conds
    ws.append(row_values)
    last_row = ws.max_row

    total_cols = 2 + num_condition_cols
    for col_idx in range(1, total_cols + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = COMMON_WIDTH

    for cell in ws[1]:
        cell.font = BOLD
        cell.alignment = COMMON_ALIGN

    for cell in ws[last_row]:
        cell.alignment = COMMON_ALIGN

    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        try:
            wb.remove(wb["Sheet"])
        except Exception:
            pass

    wb.save(xlsx_path)

def _read_text(path: str) -> str:
    for enc in ("utf-8", "latin-1", "utf-16"):
        try:
            with open(path, "r", encoding=enc) as f:
                return f.read()
        except UnicodeDecodeError:
            continue
        except FileNotFoundError:
            raise
    with open(path, "r") as f:
        return f.read()

def _strip_comment(line: str) -> str:
    line = line.split("#", 1)[0]
    line = line.split(";", 1)[0]
    return line.strip()

def parse_model_ini_for_cust(model_ini_path: str) -> Optional[str]:
    txt = _read_text(model_ini_path)
    for raw in txt.splitlines():
        line = _strip_comment(raw)
        if not line:
            continue
        m = re.match(r'^\s*CustRetailModeInten\s*=\s*(.+)$', line, re.IGNORECASE)
        if m:
            return m.group(1).strip()
    return None

def main():
    parser = argparse.ArgumentParser(description="Check CustRetailModeInten in model.ini")
    parser.add_argument("--model-ini", required=True, help="Path to model.ini")
    parser.add_argument("--root", default=".", help="Root dir for tvconfigs (not used here)")
    parser.add_argument("--report", default="kipling.xlsx", help="xlsx report path")
    parser.add_argument("-v", "--verbose", action="store_true")
    args = parser.parse_args()

    cust_val = parse_model_ini_for_cust(args.model_ini)
    passed = cust_val is not None

    res = {
        "model_ini": args.model_ini,
        "passed": passed,
        "cust_val": cust_val,
    }

    if args.verbose:
        if passed:
            print(f"[INFO] Found CustRetailModeInten = {cust_val}")
        else:
            print("[INFO] CustRetailModeInten not defined")

    if args.report:
        export_report(res, args.report)

    print("Result :", "PASS" if passed else "FAIL")

if __name__ == "__main__":
    main()
