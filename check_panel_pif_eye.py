#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
check_panel_pif_eye.py

依照 tv_multi_standard_validation.py 的邏輯：
- 路徑映射：把 /tvconfigs/* 轉到 --root 專案根目錄下
- xlsx 輸出格式：表頭固定 Rules, Result, condition_1..N；無值以 N/A 填入
- sheet 命名規則：model.ini 檔名前綴是數字 N → PID_N；否則 others
功能：
- 解析 model.ini，取得 m_pPanelName / PIF_BIN / EYE_DIAGRAM_BIN 三個參數的值（允許引號與註解）
- 如果任一參數缺少，整體 Result=FAIL；全部存在則 PASS
- console 列印三個參數的值
- （可選）輸出到 kipling.xlsx 或指定 --report-xlsx；若 xlsx 已存在則附加資料

Python 3.8+
需要 openpyxl： pip install --user openpyxl
"""
import argparse
import os
import re
from typing import Dict, List, Optional

# -----------------------------
# Utilities for report (aligned with tv_multi_standard_validation.py)
# -----------------------------

def _sheet_name_for_model(model_ini_path: str) -> str:
    """
    根據 model.ini 檔名的前綴決定頁簽：
      - 前綴是數字 N → 'PID_N'
      - 其他 → 'others'
    例: '1_EU_XXX.ini' → 'PID_1'
    """
    base = os.path.basename(model_ini_path or "")
    m = re.match(r"^(\d+)_", base)
    if m:
        return f"PID_{int(m.group(1))}"
    return "others"


def _ensure_openpyxl():
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        raise SystemExit(
            "[ERROR] 需要 openpyxl 以支援報表輸出與附加。\n"
            "  安裝： pip install --user openpyxl\n"
        )


def export_report(res: Dict, xlsx_path: str = "kipling.xlsx", num_condition_cols: int = 6) -> None:
    """
    欄位無值時以 'N/A' 填入。依 model.ini 檔名前綴分頁（PID_1、PID_2…；非數字→others），既有資料則附加。
    表頭固定為: Rules, Result, condition_1, condition_2, ...
    統一：所有欄位同寬、同為自動換行、垂直置頂（包含表頭）。
    """
    _ensure_openpyxl()
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    COMMON_WIDTH = 80
    COMMON_ALIGN = Alignment(wrap_text=True, vertical="top")
    BOLD = Font(bold=True)

    def _na(v: Optional[str]) -> str:
        s = (v or "").strip()
        return s if s else "N/A"

    sheet_name = _sheet_name_for_model(res.get("model_ini", ""))

    # 開啟或新建 xlsx
    try:
        wb = load_workbook(xlsx_path)
    except Exception:
        wb = Workbook()

    # 建立或取得 sheet
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)
        headers = ["Rules", "Result"] + [f"condition_{i}" for i in range(1, num_condition_cols + 1)]
        ws.append(headers)

    # 組裝欄位內容
    rules = "Check m_pPanelName / PIF_BIN / EYE_DIAGRAM_BIN present?"
    result = "PASS" if res.get("passed", False) else "FAIL"

    conds = [
        f"m_pPanelName = {_na(res.get('m_pPanelName'))}",            # condition_1
        f"PIF_BIN      = {_na(res.get('PIF_BIN'))}",                 # condition_2
        f"EYE_DIAGRAM_BIN = {_na(res.get('EYE_DIAGRAM_BIN'))}",      # condition_3
        f"MissingKeys  = {_na(', '.join(res.get('missing_keys', [])))}",  # condition_4
        #f"Model.ini    = {_na(res.get('model_ini'))}",               # condition_5
        #f"Root         = {_na(res.get('root'))}",                    # condition_6
    ][:num_condition_cols]

    ws.append([rules, result] + conds)
    last_row = ws.max_row

    # 套用樣式
    total_cols = 2 + num_condition_cols
    for col_idx in range(1, total_cols + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = COMMON_WIDTH

    for cell in ws[1]:  # header row
        cell.font = BOLD
        cell.alignment = COMMON_ALIGN
    for cell in ws[last_row]:
        cell.alignment = COMMON_ALIGN

    # 移除預設 Sheet
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        try:
            wb.remove(wb["Sheet"])
        except Exception:
            pass

    wb.save(xlsx_path)


# -----------------------------
# Core parsing / validation
# -----------------------------

def _read_text(path: str) -> str:
    for enc in ("utf-8", "latin-1", "utf-16"):
        try:
            with open(path, "r", encoding=enc) as f:
                return f.read()
        except UnicodeDecodeError:
            continue
        except FileNotFoundError:
            raise
    # 若皆失敗，讓原始例外往外拋
    with open(path, "r") as f:
        return f.read()


def _strip_comment(line: str) -> str:
    # 去掉 # 或 ; 之後的註解
    line = line.split("#", 1)[0]
    line = line.split(";", 1)[0]
    return line.strip()


def _resolve_tvconfigs_path(root: str, tvconfigs_like: str) -> str:
    """
    把 "/tvconfigs/xxx/yyy.ini" 映射為 "<root>/xxx/yyy.ini"
    其他相對路徑: 以 root 為基底
    絕對路徑（非 /tvconfigs 開頭）維持不動
    """
    if tvconfigs_like.startswith("/tvconfigs/"):
        rel = tvconfigs_like[len("/tvconfigs/"):]
        return os.path.normpath(os.path.join(root, rel))
    if tvconfigs_like.startswith("./") or tvconfigs_like.startswith("../"):
        return os.path.normpath(os.path.join(root, tvconfigs_like))
    if tvconfigs_like.startswith("/"):
        # 非 /tvconfigs 絕對路徑：原樣返回
        return tvconfigs_like
    # 其他當作相對於 root
    return os.path.normpath(os.path.join(root, tvconfigs_like))


def parse_three_params(model_ini_path: str) -> Dict[str, Optional[str]]:
    """
    從 model.ini 擷取三個參數（大小寫不敏感）：
      - m_pPanelName
      - PIF_BIN
      - EYE_DIAGRAM_BIN
    允許值被單引號或雙引號包住。會忽略 # 或 ; 註解之後的內容。
    """
    txt = _read_text(model_ini_path)
    params = {"m_pPanelName": None, "PIF_BIN": None, "EYE_DIAGRAM_BIN": None}

    # 單行完全比對 key = value
    # e.g. m_pPanelName = "/tvconfigs/panel/xxx.ini"
    pat = {
        "m_pPanelName": re.compile(r'^\s*m_pPanelName\s*=\s*"?([^"#;]+?)"?\s*$', re.IGNORECASE),
        "PIF_BIN": re.compile(r'^\s*PIF_BIN\s*=\s*"?([^"#;]+?)"?\s*$', re.IGNORECASE),
        "EYE_DIAGRAM_BIN": re.compile(r'^\s*EYE_DIAGRAM_BIN\s*=\s*"?([^"#;]+?)"?\s*$', re.IGNORECASE),
    }

    for raw in txt.splitlines():
        line = _strip_comment(raw)
        if not line:
            continue
        for key, regex in pat.items():
            if params[key] is None:
                m = regex.match(line)
                if m:
                    params[key] = m.group(1).strip()

    return params


def build_result(args, model_ini: str, values: Dict[str, Optional[str]]) -> Dict:
    missing_keys = [k for k, v in values.items() if not v]
    passed = len(missing_keys) == 0

    return {
        "passed": passed,
        "model_ini": model_ini,
        "root": os.path.abspath(os.path.normpath(args.root)),
        "m_pPanelName": values.get("m_pPanelName") or "",
        "PIF_BIN": values.get("PIF_BIN") or "",
        "EYE_DIAGRAM_BIN": values.get("EYE_DIAGRAM_BIN") or "",
        "missing_keys": missing_keys,
    }


# -----------------------------
# Main
# -----------------------------

def main():
    parser = argparse.ArgumentParser(description="Check m_pPanelName / PIF_BIN / EYE_DIAGRAM_BIN in model.ini and export report.")
    parser.add_argument("--model-ini", required=True, help="path to model ini (e.g., model/1_xxx.ini)")
    parser.add_argument("--root", required=True, help="tvconfigs project root (maps /tvconfigs/* to here)")
    parser.add_argument("-v", "--verbose", action="store_true", help="verbose logs")

    # 報表參數：--report 與 --report-xlsx（任一存在即輸出；未指定路徑則用 kipling.xlsx）
    parser.add_argument("--report", action="store_true", help="export report to xlsx (default: kipling.xlsx)")
    parser.add_argument("--report-xlsx", metavar="FILE", help="export report to specific xlsx file")

    args = parser.parse_args()

    model_ini = args.model_ini
    if not os.path.exists(model_ini):
        raise SystemExit(f"[ERROR] model ini not found: {model_ini}")

    # 解析三個參數
    values = parse_three_params(model_ini)

    # Console 輸出
    print("=== Parameters from model.ini ===")
    print(f"m_pPanelName    : {values.get('m_pPanelName') or '(missing)'}")
    print(f"PIF_BIN         : {values.get('PIF_BIN') or '(missing)'}")
    print(f"EYE_DIAGRAM_BIN : {values.get('EYE_DIAGRAM_BIN') or '(missing)'}")

    res = build_result(args, model_ini, values)
    print(f"Result          : {'PASS' if res['passed'] else 'FAIL'}")

    # 報表輸出
    if args.report or args.report_xlsx:
        xlsx_path = args.report_xlsx if args.report_xlsx else "kipling.xlsx"
        export_report(res, xlsx_path=xlsx_path)
        sheet = _sheet_name_for_model(model_ini)
        print(f"[INFO] Report appended to: {xlsx_path} (sheet: {sheet})")


if __name__ == "__main__":
    main()
